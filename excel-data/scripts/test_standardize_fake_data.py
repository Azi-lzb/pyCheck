from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook


SCRIPTS_DIR = Path(__file__).resolve().parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import standardize_fake_data as mod


def _build_config(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "config"
    ws.append(["placeholder"])

    definitions = wb.create_sheet("明细定义")
    definitions.append(["序号", "明细名称", "工作簿名关键字", "工作表名关键字", "表头行号", "数据起始行号", "是否禁用", "备注"])
    definitions.append([1, "科技贷款明细", "科技", "科技贷款", 3, 4, None, None])
    definitions.append([2, "养老贷款明细", "综合台账", "养老贷款", 3, 4, None, None])

    params = wb.create_sheet("明细参数")
    params.append(["序号", "明细名称", "明细字段", "字段所在列位置", "匹配方式", "是否禁用", "备注", "添加日期"])
    params.append([1, "科技贷款明细", "列位置_贷款产品类型", "M", "exact", None, None, None])
    params.append([2, "科技贷款明细", "列位置_高技术服务业贷款类型大类编码", "T", "exact", None, None, None])
    params.append([3, "科技贷款明细", "列位置_贷款项目名称", "L", "exact,contains", None, None, None])
    params.append([4, "科技贷款明细", "列位置_贷款客户名称", "E", "exact,regex", None, None, None])
    params.append([5, "科技贷款明细", "列位置_通用两位数字编码", "N", "regex", None, None, None])
    params.append([6, "科技贷款明细", "列位置_通用行业编码", "O", "regex", None, None, None])
    params.append([7, "养老贷款明细", "列位置_贷款产品类型", "M", "exact", None, None, None])

    mapping = wb.create_sheet("字段映射表")
    mapping.append(["序号", "明细名称", "列字段", "原始列值_exact", "原始列值_contains", "原始列值_regex", "映射标准值", "映射标准值_regex", "是否禁用", "追加时间"])
    mapping.append([1, "科技贷款明细", "列位置_贷款产品类型", "经营贷款", None, None, "是", None, None, None])
    mapping.append([2, "科技贷款明细", "列位置_贷款产品类型", "固定资产贷款", None, None, "否", None, None, None])
    mapping.append([3, "科技贷款明细", "列位置_高技术服务业贷款类型大类编码", "06-科技成果转化服务", None, None, "HTS06", None, None, None])
    mapping.append([4, "科技贷款明细", "列位置_贷款项目名称", "项目A", None, None, "项目A标准名", None, None, None])
    mapping.append([5, "科技贷款明细", "列位置_贷款项目名称", None, "技术升级", None, "命中技术升级", None, None, None])
    mapping.append([6, "科技贷款明细", "列位置_贷款客户名称", "晨光新材料有限公司", None, None, "晨光精确客户", None, None, None])
    mapping.append([7, "科技贷款明细", "列位置_贷款客户名称", None, None, r"^晨光.*有限公司$", None, "晨光标准客户", None, None])
    mapping.append([8, "科技贷款明细", "列位置_通用两位数字编码", None, None, r"^(\d{2})[-－].+$", None, "HTS{REGEX_GROUP_1}", None, None])
    mapping.append([9, "科技贷款明细", "列位置_通用行业编码", None, None, r"^([A-Za-z]\d{4})[-－].+$", None, "{REGEX_GROUP_1}", None, None])
    mapping.append([10, "养老贷款明细", "列位置_贷款产品类型", "经营贷款", None, None, "养老经营贷款", None, None, None])

    runtime = wb.create_sheet("runtime")
    runtime.append(["配置项", "值", "说明"])
    runtime.append(["标准化匹配方式", "exact", "字段级匹配方式启用后，此项仅保留兼容"])

    wb.save(path)
    wb.close()


def _rename_to_prefixed_sheet_names(path: Path):
    wb = load_workbook(path)
    wb["明细定义"].title = "标准化配置-明细定义"
    wb["明细参数"].title = "标准化配置-明细参数"
    wb["字段映射表"].title = "标准化配置-字段映射表"
    wb["runtime"].title = "输出报告配置-runtime"
    wb.save(path)
    wb.close()


def _build_source(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "科技贷款"
    ws["M3"] = "贷款类型（流动资金贷款、固定资产贷款、并购贷款、贸易融资等）"
    ws["T3"] = "高技术服务业贷款类型大类编码-对应名称"
    ws["L3"] = "贷款项目名称（贷款合同或相关文件载明的项目名称）"
    ws["E3"] = "贷款客户名称"
    ws["N3"] = "两位数字编码"
    ws["O3"] = "行业编码"
    ws["M4"] = "经营贷款"
    ws["T4"] = "06-科技成果转化服务"
    ws["L4"] = "用于生产经营/技术升级"
    ws["E4"] = "晨光新材料有限公司"
    ws["N4"] = "05-研发与应用"
    ws["O4"] = "C1234-研发与应用"
    ws["M5"] = "固定资产贷款"
    ws["T5"] = "HTS06"
    ws["L5"] = "一般经营周转"
    ws["E5"] = "晨光科技有限公司"
    ws["N5"] = "12-其他服务"
    ws["O5"] = "R8626-数字出版"
    wb.save(path)
    wb.close()


def _build_multi_sheet_source(path: Path):
    wb = Workbook()
    tech = wb.active
    tech.title = "科技贷款"
    tech["E3"] = "贷款客户名称"
    tech["L3"] = "贷款项目名称（贷款合同或相关文件载明的项目名称）"
    tech["M3"] = "贷款类型（流动资金贷款、固定资产贷款、并购贷款、贸易融资等）"
    tech["T3"] = "高技术服务业贷款类型大类编码-对应名称"
    tech["E4"] = "晨光新材料有限公司"
    tech["L4"] = "用于生产经营/技术升级"
    tech["M4"] = "经营贷款"
    tech["T4"] = "06-科技成果转化服务"

    elder = wb.create_sheet("养老贷款")
    elder["M3"] = "贷款类型（流动资金贷款、固定资产贷款、并购贷款、贸易融资等）"
    elder["M4"] = "经营贷款"
    elder["M5"] = "固定资产贷款"
    wb.save(path)
    wb.close()


def test_load_rules_from_config(tmp_path):
    config_path = tmp_path / "config.xlsx"
    _build_config(config_path)

    detail_definition = mod.match_detail_definition(config_path, Path("某科技台账.xlsx"), "科技贷款")
    rules = mod.load_standardization_rules(config_path, detail_definition["明细名称"])

    assert rules["column_letters"] == {
        "列位置_贷款产品类型": "M",
        "列位置_高技术服务业贷款类型大类编码": "T",
        "列位置_贷款项目名称": "L",
        "列位置_贷款客户名称": "E",
        "列位置_通用两位数字编码": "N",
        "列位置_通用行业编码": "O",
    }
    assert rules["field_modes"] == {
        "列位置_贷款产品类型": "exact",
        "列位置_高技术服务业贷款类型大类编码": "exact",
        "列位置_贷款项目名称": "exact,contains",
        "列位置_贷款客户名称": "exact,regex",
        "列位置_通用两位数字编码": "regex",
        "列位置_通用行业编码": "regex",
    }
    assert rules["default_match_mode"] == "exact"
    type_rule = rules["value_maps"]["列位置_贷款产品类型"][0]
    assert type_rule["source"] == "经营贷款"
    assert type_rule["target"] == "是"
    assert type_rule["mode"] == "exact"
    assert [rule["mode"] for rule in rules["value_maps"]["列位置_贷款项目名称"]] == ["exact", "contains"]
    assert [rule["source"] for rule in rules["value_maps"]["列位置_贷款项目名称"]] == ["项目A", "技术升级"]
    assert [rule["mode"] for rule in rules["value_maps"]["列位置_贷款客户名称"]] == ["exact", "regex"]
    assert rules["value_maps"]["列位置_通用两位数字编码"][0]["target"] == "HTS{REGEX_GROUP_1}"
    assert rules["value_maps"]["列位置_通用行业编码"][0]["target"] == "{REGEX_GROUP_1}"


def test_standardize_workbook_appends_normalized_columns_and_report(tmp_path):
    config_path = tmp_path / "config.xlsx"
    source_path = tmp_path / "某科技台账.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_config(config_path)
    _build_source(source_path)

    mod.standardize_workbook(source_path, "科技贷款明细", config_path, output_path, sheet_name="科技贷款")

    wb = load_workbook(output_path, data_only=True)
    ws = wb["科技贷款"]
    headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
    assert "标准化-贷款类型（流动资金贷款、固定资产贷款、并购贷款、贸易融资等）" not in headers
    assert "标准化-高技术服务业贷款类型大类编码-对应名称" not in headers
    assert "标准化-贷款项目名称（贷款合同或相关文件载明的项目名称）" not in headers
    assert "标准化-贷款客户名称" not in headers

    assert ws["M4"].value == "是"
    assert ws["T4"].value == "HTS06"
    assert ws["L4"].value == "命中技术升级"
    assert ws["E4"].value == "晨光精确客户"
    assert ws["N4"].value == "HTS05"
    assert ws["O4"].value == "C1234"
    assert ws["M5"].value == "否"
    assert ws["T5"].value == "HTS06"
    assert ws["L5"].value == "一般经营周转"
    assert ws["E5"].value == "晨光标准客户"
    assert ws["N5"].value == "HTS12"
    assert ws["O5"].value == "R8626"

    report = wb["字段映射表触发情况"]
    report_headers = [report.cell(1, c).value for c in range(1, 12)]
    assert report_headers == [
        "序号",
        "明细名称",
        "列字段",
        "原始列值_exact",
        "原始列值_contains",
        "原始列值_regex",
        "映射标准值",
        "映射标准值_regex",
        "是否禁用",
        "追加时间",
        "触发情况",
    ]
    assert report.cell(2, 1).value == 1
    assert report.cell(2, 2).value == "科技贷款明细"
    assert report.cell(2, 11).value == "贷款明细表中1笔数据映射exact触发本条规则"
    assert report.cell(3, 11).value == "贷款明细表中1笔数据映射exact触发本条规则"
    assert report.cell(4, 11).value == "贷款明细表中1笔数据映射exact触发本条规则"
    assert report.cell(5, 11).value is None
    assert report.cell(6, 11).value == "贷款明细表中1笔数据映射contains触发本条规则"
    assert report.cell(7, 11).value == "贷款明细表中1笔数据映射exact触发本条规则"
    assert report.cell(8, 11).value == "贷款明细表中1笔数据映射regex触发本条规则"
    assert report.cell(9, 11).value == "贷款明细表中2笔数据映射regex触发本条规则"
    assert report.cell(10, 11).value == "贷款明细表中2笔数据映射regex触发本条规则"

    detail_report = wb["明细定义触发情况"]
    detail_headers = [detail_report.cell(1, c).value for c in range(1, 9)]
    assert detail_headers == [
        "序号",
        "明细名称",
        "工作簿名关键字",
        "工作表名关键字",
        "表头行号",
        "数据起始行号",
        "是否禁用",
        "触发情况",
    ]
    assert detail_report.cell(2, 1).value == 1
    assert detail_report.cell(2, 2).value == "科技贷款明细"
    assert detail_report.cell(2, 8).value == "本行配置被工作簿[某科技台账]工作表[科技贷款]应用"
    assert detail_report.cell(3, 2).value == "养老贷款明细"
    assert detail_report.cell(3, 8).value is None

    param_report = wb["明细参数触发情况"]
    param_headers = [param_report.cell(1, c).value for c in range(1, 10)]
    assert param_headers == [
        "序号",
        "明细名称",
        "明细字段",
        "字段所在列位置",
        "匹配方式",
        "是否禁用",
        "备注",
        "添加日期",
        "触发情况",
    ]
    assert param_report.cell(2, 3).value == "列位置_贷款产品类型"
    assert param_report.cell(2, 9).value == "字段映射表中的2条规则被exact触发，规则序号是1和2"
    assert param_report.cell(4, 3).value == "列位置_贷款项目名称"
    assert param_report.cell(4, 9).value == "字段映射表中的1条规则被contains触发，规则序号是5"
    assert param_report.cell(5, 3).value == "列位置_贷款客户名称"
    assert param_report.cell(5, 9).value == "字段映射表中的1条规则被exact触发，规则序号是6；字段映射表中的1条规则被regex触发，规则序号是7"
    assert param_report.cell(6, 3).value == "列位置_通用两位数字编码"
    assert param_report.cell(6, 9).value == "字段映射表中的1条规则被regex触发，规则序号是8"
    wb.close()


def test_match_detail_definition_and_standardize_multi_sheet_workbook(tmp_path):
    config_path = tmp_path / "config.xlsx"
    source_path = tmp_path / "综合台账_科技养老.xlsx"
    output_dir = tmp_path / "output"
    _build_config(config_path)
    _build_multi_sheet_source(source_path)

    outputs = mod.standardize_matching_workbook(source_path, config_path, output_dir)

    assert len(outputs) == 2
    detail_names = {item["detail_name"] for item in outputs}
    assert detail_names == {"科技贷款明细", "养老贷款明细"}

    tech_output = next(item["output_path"] for item in outputs if item["detail_name"] == "科技贷款明细")
    elder_output = next(item["output_path"] for item in outputs if item["detail_name"] == "养老贷款明细")

    tech_wb = load_workbook(tech_output, data_only=True)
    tech_ws = tech_wb["科技贷款"]
    assert tech_ws["E4"].value == "晨光精确客户"
    tech_wb.close()

    elder_wb = load_workbook(elder_output, data_only=True)
    elder_ws = elder_wb["养老贷款"]
    assert elder_ws["M4"].value == "养老经营贷款"
    assert elder_ws["M5"].value == "固定资产贷款"
    elder_wb.close()


def test_load_rules_from_mapping_headers_not_column_order(tmp_path):
    config_path = tmp_path / "config.xlsx"
    source_path = tmp_path / "某科技台账.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_config(config_path)
    _build_source(source_path)

    wb = load_workbook(config_path)
    ws = wb["字段映射表"]
    rows = list(ws.iter_rows(values_only=True))
    original_headers = list(rows[0])
    target_headers = [
        "明细名称",
        "列字段",
        "原始列值_exact",
        "原始列值_contains",
        "原始列值_regex",
        "映射标准值_regex",
        "是否禁用",
        "备注",
        "映射标准值",
    ]
    data_rows = [dict(zip(original_headers, row)) for row in rows[1:]]
    ws.delete_rows(1, ws.max_row)
    ws.append(target_headers)
    for item in data_rows:
        ws.append([item.get(header) for header in target_headers])
    wb.save(config_path)
    wb.close()

    mod.standardize_workbook(source_path, "科技贷款明细", config_path, output_path, sheet_name="科技贷款")

    out_wb = load_workbook(output_path, data_only=True)
    out_ws = out_wb["科技贷款"]
    assert out_ws["M4"].value == "是"
    assert out_ws["N4"].value == "HTS05"
    assert out_ws["O4"].value == "C1234"
    out_wb.close()


def test_standardize_matching_workbook_supports_prefixed_sheet_names(tmp_path):
    config_path = tmp_path / "config.xlsx"
    source_path = tmp_path / "某科技台账.xlsx"
    output_dir = tmp_path / "output"
    _build_config(config_path)
    _rename_to_prefixed_sheet_names(config_path)
    _build_source(source_path)

    outputs = mod.standardize_matching_workbook(source_path, config_path, output_dir)

    assert len(outputs) == 1
    output_path = outputs[0]["output_path"]
    wb = load_workbook(output_path, data_only=True)
    ws = wb["科技贷款"]
    assert ws["M4"].value == "是"
    assert ws["T4"].value == "HTS06"
    wb.close()
