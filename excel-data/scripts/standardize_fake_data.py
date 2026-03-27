import argparse
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG = PROJECT_ROOT / "check-program" / "main" / "config.xlsx"
DEFAULT_OUTPUT_DIR = PROJECT_ROOT / "excel-data" / "output"
DEFAULT_INPUT_DIR = PROJECT_ROOT / "excel-data" / "fakedata"
DEFAULT_INPUTS = {
    "科技产业": PROJECT_ROOT / "excel-data" / "fakedata" / "附表2：科技产业贷款明细（0306改）11：46改.xlsx",
    "数字经济产业": PROJECT_ROOT / "excel-data" / "fakedata" / "附表4：数字经济产业贷款明细（0306改）.xlsx",
    "养老产业": PROJECT_ROOT / "excel-data" / "fakedata" / "附表5：养老产业贷款明细（0306改）.xlsx",
}
HEADER_ROWS = {
    "科技产业": 3,
    "数字经济产业": 4,
    "养老产业": 3,
}
SHEET_NAME_ALIASES = {
    "明细定义": ["标准化配置-明细定义", "明细定义"],
    "明细参数": ["标准化配置-明细参数", "明细参数"],
    "字段映射表": ["标准化配置-字段映射表", "字段映射表"],
    "runtime": ["输出报告配置-runtime", "runtime"],
}


def get_sheet_by_logical_name(workbook, logical_name: str):
    for sheet_name in SHEET_NAME_ALIASES.get(logical_name, [logical_name]):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
    return None


def load_detail_definitions(config_path: Path):
    wb = load_workbook(config_path, data_only=True)
    ws = get_sheet_by_logical_name(wb, "明细定义")
    if ws is None:
        wb.close()
        return []
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_map = {str(value).strip(): idx for idx, value in enumerate(headers) if value is not None}
    definitions = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row = list(row)
        detail_name = row[header_map["明细名称"]] if "明细名称" in header_map and header_map["明细名称"] < len(row) else None
        if not detail_name:
            continue
        disabled_value = row[header_map["是否禁用"]] if "是否禁用" in header_map and header_map["是否禁用"] < len(row) else None
        if str(disabled_value).strip() == "1":
            continue
        workbook_keyword = "" if "工作簿名关键字" not in header_map or header_map["工作簿名关键字"] >= len(row) or row[header_map["工作簿名关键字"]] is None else str(row[header_map["工作簿名关键字"]]).strip()
        worksheet_keyword = "" if "工作表名关键字" not in header_map or header_map["工作表名关键字"] >= len(row) or row[header_map["工作表名关键字"]] is None else str(row[header_map["工作表名关键字"]]).strip()
        header_row = None if "表头行号" not in header_map or header_map["表头行号"] >= len(row) or row[header_map["表头行号"]] is None else int(row[header_map["表头行号"]])
        data_start = None if "数据起始行号" not in header_map or header_map["数据起始行号"] >= len(row) or row[header_map["数据起始行号"]] is None else int(row[header_map["数据起始行号"]])
        definitions.append(
            {
                "序号": row[header_map["序号"]] if "序号" in header_map and header_map["序号"] < len(row) else None,
                "明细名称": str(detail_name).strip(),
                "工作簿名关键字": workbook_keyword,
                "工作表名关键字": worksheet_keyword,
                "表头行号": header_row,
                "数据起始行号": data_start or (header_row + 1 if header_row else None),
                "是否禁用": disabled_value,
            }
        )
    wb.close()
    return definitions


def match_detail_definition(config_path: Path, workbook_path: Path, sheet_name: str):
    workbook_name = workbook_path.stem
    matches = []
    for definition in load_detail_definitions(config_path):
        workbook_keyword = definition["工作簿名关键字"]
        worksheet_keyword = definition["工作表名关键字"]
        if workbook_keyword and workbook_keyword not in workbook_name:
            continue
        if worksheet_keyword and worksheet_keyword not in sheet_name:
            continue
        matches.append(definition)
    if len(matches) > 1:
        raise ValueError(
            f"{workbook_path.name} / {sheet_name} 命中多条明细定义："
            + "、".join(item["明细名称"] for item in matches)
        )
    return matches[0] if matches else None


def find_detail_definition_by_name(config_path: Path, detail_name: str):
    for definition in load_detail_definitions(config_path):
        if definition["明细名称"] == detail_name:
            return definition
    return None


def load_standardization_rules(config_path: Path, detail_name: str):
    wb = load_workbook(config_path, data_only=True)
    mapping_ws = get_sheet_by_logical_name(wb, "字段映射表")
    params_ws = get_sheet_by_logical_name(wb, "明细参数")
    if mapping_ws is None or params_ws is None:
        wb.close()
        raise ValueError(f"{config_path} 缺少 字段映射表 或 明细参数 sheet。")
    default_match_mode = "exact"
    runtime_ws = get_sheet_by_logical_name(wb, "runtime")
    if runtime_ws is not None:
        for row in runtime_ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            if str(row[0]).strip() == "标准化匹配方式" and row[1] is not None:
                candidate = str(row[1]).strip().lower()
                if candidate in {"exact", "contains", "regex"}:
                    default_match_mode = candidate
                break

    param_headers = [params_ws.cell(1, c).value for c in range(1, params_ws.max_column + 1)]
    param_header_map = {str(value).strip(): idx for idx, value in enumerate(param_headers) if value is not None}
    column_letters = {}
    field_modes = {}
    field_mode_orders = {}
    param_report_rows = []
    for row in params_ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row = list(row)
        row_detail_name = row[param_header_map["明细名称"]] if "明细名称" in param_header_map and param_header_map["明细名称"] < len(row) else None
        row_field_name = row[param_header_map["明细字段"]] if "明细字段" in param_header_map and param_header_map["明细字段"] < len(row) else None
        row_column = row[param_header_map["字段所在列位置"]] if "字段所在列位置" in param_header_map and param_header_map["字段所在列位置"] < len(row) else None
        if row_detail_name != detail_name or not row_field_name or not row_column:
            continue
        match_mode = default_match_mode
        mode_order = [default_match_mode]
        if "匹配方式" in param_header_map and param_header_map["匹配方式"] < len(row) and row[param_header_map["匹配方式"]] is not None:
            candidates = [item.strip().lower() for item in str(row[param_header_map["匹配方式"]]).split(",") if item and item.strip()]
            valid_modes = [item for item in candidates if item in {"exact", "contains", "regex"}]
            if valid_modes:
                match_mode = ",".join(valid_modes)
                mode_order = valid_modes
        disabled_value = row[param_header_map["是否禁用"]] if "是否禁用" in param_header_map and param_header_map["是否禁用"] < len(row) else None
        if str(disabled_value).strip() == "1":
            continue
        field_name = str(row_field_name).strip()
        column_letters[field_name] = str(row_column).strip().upper()
        field_modes[field_name] = match_mode
        field_mode_orders[field_name] = mode_order
        row_values = {param_headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(param_headers)) if param_headers[idx] is not None}
        param_report_rows.append({"field_name": field_name, "values": row_values})

    headers = [mapping_ws.cell(1, c).value for c in range(1, mapping_ws.max_column + 1)]
    header_map = {str(value).strip(): idx for idx, value in enumerate(headers) if value is not None}
    has_named_layout = "明细名称" in header_map and "列字段" in header_map
    value_maps = {}
    mode_buckets = {}
    mapping_report_rows = []
    for row in mapping_ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        row = list(row)
        if has_named_layout:
            row_detail_name = row[header_map["明细名称"]] if header_map["明细名称"] < len(row) else None
            row_field_name = row[header_map["列字段"]] if header_map["列字段"] < len(row) else None
        else:
            row_detail_name = row[0] if len(row) > 0 else None
            row_field_name = row[1] if len(row) > 1 else None
        if row_detail_name != detail_name or not row_field_name:
            continue
        field_name = str(row_field_name).strip()
        disabled_value = None
        if has_named_layout and "是否禁用" in header_map and header_map["是否禁用"] < len(row):
            disabled_value = row[header_map["是否禁用"]]
        elif len(row) >= 7:
            disabled_value = row[5]
        if str(disabled_value).strip() == "1":
            continue
        if field_name not in column_letters:
            continue
        report_row_index = len(mapping_report_rows)
        row_values = {}
        if has_named_layout:
            for header_name, idx in header_map.items():
                row_values[header_name] = row[idx] if idx < len(row) else None
        else:
            for idx, value in enumerate(row):
                row_values[f"col_{idx}"] = value
        mapping_report_rows.append({"values": row_values, "hits": []})
        mode_buckets.setdefault(field_name, {mode: [] for mode in field_mode_orders.get(field_name, [default_match_mode])})
        if has_named_layout:
            for mode in field_mode_orders.get(field_name, [default_match_mode]):
                src_col_name = {
                    "exact": "原始列值_exact",
                    "contains": "原始列值_contains",
                    "regex": "原始列值_regex",
                }[mode]
                target_col_name = "映射标准值_regex" if mode == "regex" and "映射标准值_regex" in header_map else "映射标准值"
                src_idx = header_map.get(src_col_name)
                target_idx = header_map.get(target_col_name)
                src = "" if src_idx is None or src_idx >= len(row) or row[src_idx] is None else str(row[src_idx]).strip()
                target = None if target_idx is None or target_idx >= len(row) else row[target_idx]
                if src:
                    mode_buckets[field_name].setdefault(mode, []).append(
                        {
                            "source": src,
                            "target": target,
                            "mode": mode,
                            "report_row_index": report_row_index,
                            "rule_serial": row[header_map["序号"]] if "序号" in header_map and header_map["序号"] < len(row) else None,
                        }
                    )
        else:
            src = "" if row[2] is None else str(row[2]).strip()
            mode = field_mode_orders.get(field_name, [default_match_mode])[0]
            target = row[3]
            if src:
                mode_buckets[field_name].setdefault(mode, []).append(
                    {"source": src, "target": target, "mode": mode, "report_row_index": report_row_index, "rule_serial": None}
                )

    for field_name, buckets in mode_buckets.items():
        ordered_rules = []
        for mode in field_mode_orders.get(field_name, [default_match_mode]):
            ordered_rules.extend(buckets.get(mode, []))
        if ordered_rules:
            value_maps[field_name] = ordered_rules

    wb.close()
    return {
        "column_letters": column_letters,
        "field_modes": field_modes,
        "field_mode_orders": field_mode_orders,
        "value_maps": value_maps,
        "mapping_headers": headers,
        "mapping_report_rows": mapping_report_rows,
        "param_headers": param_headers,
        "param_report_rows": param_report_rows,
        "default_match_mode": default_match_mode,
    }


def apply_mapping_rules(original, rules):
    if original is None:
        return original, None
    text = str(original).strip()
    for rule in rules:
        mode = rule["mode"]
        if mode == "contains":
            if rule["source"] in text:
                return rule["target"], rule
        elif mode == "regex":
            match = re.search(rule["source"], text)
            if match:
                target = rule["target"]
                if isinstance(target, str):
                    group_ref = re.fullmatch(r"REGEX_GROUP_(\d+)", target.strip())
                    if group_ref:
                        group_index = int(group_ref.group(1))
                        try:
                            target = match.group(group_index)
                        except IndexError:
                            target = rule["target"]
                    else:
                        def _replace_group_token(token_match):
                            group_index = int(token_match.group(1))
                            try:
                                return match.group(group_index)
                            except IndexError:
                                return token_match.group(0)

                        target = re.sub(r"\{REGEX_GROUP_(\d+)\}", _replace_group_token, target)
                return target, rule
        else:
            if text == rule["source"]:
                return rule["target"], rule
    return original, None


def standardize_workbook(
    input_path: Path,
    detail_name: str,
    config_path: Path,
    output_path: Path,
    sheet_name=None,
    header_row=None,
    data_start=None,
):
    rules = load_standardization_rules(config_path, detail_name)
    wb = load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    detail_definition = find_detail_definition_by_name(config_path, detail_name)
    if detail_definition:
        header_row = header_row or detail_definition["表头行号"]
        data_start = data_start or detail_definition["数据起始行号"]
    header_row = header_row or HEADER_ROWS.get(detail_name)
    if header_row is None:
        wb.close()
        raise ValueError(f"{detail_name} 缺少表头行号配置。")
    data_start = data_start or (header_row + 1)

    report_hit_counts = {}
    param_hit_counts = {}
    for field_name, mapping in rules["value_maps"].items():
        col_letter = rules["column_letters"][field_name]
        src_col = column_index_from_string(col_letter)

        changed = 0
        filled = 0
        for row_idx in range(data_start, ws.max_row + 1):
            original = ws.cell(row_idx, src_col).value
            normalized, matched_rule = apply_mapping_rules(original, mapping)
            if original not in (None, ""):
                filled += 1
            if matched_rule:
                report_row_index = matched_rule.get("report_row_index")
                if report_row_index is not None:
                    hit_bucket = report_hit_counts.setdefault(report_row_index, {})
                    mode = matched_rule["mode"]
                    hit_bucket[mode] = hit_bucket.get(mode, 0) + 1
                    field_bucket = param_hit_counts.setdefault(field_name, {})
                    field_bucket.setdefault(mode, set()).add(matched_rule.get("rule_serial"))
            if normalized != original:
                changed += 1
                ws.cell(row_idx, src_col).value = normalized

    report_name = "字段映射表触发情况"
    if report_name in wb.sheetnames:
        del wb[report_name]
    report_ws = wb.create_sheet(report_name)
    named_headers = [header for header in rules.get("mapping_headers", []) if header is not None]
    report_headers = named_headers + ["触发情况"]
    report_ws.append(report_headers)
    for c in range(1, len(report_headers) + 1):
        report_ws.cell(1, c).font = Font(name="Arial", bold=True)
    for idx, item in enumerate(rules.get("mapping_report_rows", [])):
        row_values = [item["values"].get(header) for header in named_headers]
        hit_modes = report_hit_counts.get(idx, {})
        trigger_summary = "；".join(f"贷款明细表中{count}笔数据映射{mode}触发本条规则" for mode, count in hit_modes.items())
        report_ws.append(row_values + [trigger_summary])

    detail_report_name = "明细定义触发情况"
    if detail_report_name in wb.sheetnames:
        del wb[detail_report_name]
    detail_ws = wb.create_sheet(detail_report_name)
    detail_headers = ["序号", "明细名称", "工作簿名关键字", "工作表名关键字", "表头行号", "数据起始行号", "是否禁用", "触发情况"]
    detail_ws.append(detail_headers)
    for c in range(1, len(detail_headers) + 1):
        detail_ws.cell(1, c).font = Font(name="Arial", bold=True)
    for definition in load_detail_definitions(config_path):
        trigger = None
        if detail_definition and definition["明细名称"] == detail_definition["明细名称"]:
            trigger = f"本行配置被工作簿[{input_path.stem}]工作表[{ws.title}]应用"
        detail_ws.append(
            [
                definition.get("序号"),
                definition.get("明细名称"),
                definition.get("工作簿名关键字"),
                definition.get("工作表名关键字"),
                definition.get("表头行号"),
                definition.get("数据起始行号"),
                definition.get("是否禁用"),
                trigger,
            ]
        )

    param_report_name = "明细参数触发情况"
    if param_report_name in wb.sheetnames:
        del wb[param_report_name]
    param_ws = wb.create_sheet(param_report_name)
    param_headers = [header for header in rules.get("param_headers", []) if header is not None] + ["触发情况"]
    param_ws.append(param_headers)
    for c in range(1, len(param_headers) + 1):
        param_ws.cell(1, c).font = Font(name="Arial", bold=True)
    for item in rules.get("param_report_rows", []):
        values = [item["values"].get(header) for header in param_headers[:-1]]
        hit_modes = param_hit_counts.get(item["field_name"], {})
        parts = []
        for mode, rule_serials in hit_modes.items():
            serial_list = [str(v) for v in sorted(v for v in rule_serials if v is not None)]
            serial_text = "和".join(serial_list)
            parts.append(f"字段映射表中的{len(rule_serials)}条规则被{mode}触发，规则序号是{serial_text}")
        trigger = "；".join(parts)
        param_ws.append(values + [trigger])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return output_path


def _default_output_path(detail_name: str, input_path: Path):
    return DEFAULT_OUTPUT_DIR / f"标准化_{detail_name}_{input_path.name}"


def _matching_output_path(detail_name: str, workbook_path: Path, sheet_name: str, output_dir: Path):
    safe_sheet_name = re.sub(r'[\\\\/:*?"<>|]+', "_", sheet_name)
    return output_dir / f"标准化_{detail_name}_{workbook_path.stem}_{safe_sheet_name}.xlsx"


def standardize_matching_workbook(input_path: Path, config_path: Path, output_dir: Path):
    wb = load_workbook(input_path, read_only=True)
    outputs = []
    try:
        for sheet_name in wb.sheetnames:
            detail_definition = match_detail_definition(config_path, input_path, sheet_name)
            if not detail_definition:
                continue
            detail_name = detail_definition["明细名称"]
            output_path = _matching_output_path(detail_name, input_path, sheet_name, output_dir)
            standardize_workbook(
                input_path,
                detail_name,
                config_path,
                output_path,
                sheet_name=sheet_name,
                header_row=detail_definition["表头行号"],
                data_start=detail_definition["数据起始行号"],
            )
            outputs.append({"detail_name": detail_name, "sheet_name": sheet_name, "output_path": output_path})
    finally:
        wb.close()
    return outputs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default=str(DEFAULT_CONFIG))
    parser.add_argument("--detail")
    parser.add_argument("--input")
    parser.add_argument("--output")
    args = parser.parse_args()

    config_path = Path(args.config)
    if args.detail:
        if args.input:
            input_path = Path(args.input)
        elif args.detail in DEFAULT_INPUTS:
            input_path = DEFAULT_INPUTS[args.detail]
        else:
            raise ValueError(f"{args.detail} 未提供 --input，且不在默认明细列表中。")
        output_path = Path(args.output) if args.output else _default_output_path(args.detail, input_path)
        out = standardize_workbook(input_path, args.detail, config_path, output_path)
        print(out)
        return

    candidate_inputs = sorted(path for path in DEFAULT_INPUT_DIR.glob("*.xlsx") if path.is_file())
    for input_path in candidate_inputs:
        outputs = standardize_matching_workbook(input_path, config_path, DEFAULT_OUTPUT_DIR)
        if outputs:
            for item in outputs:
                print(item["output_path"])
            continue

        matched_detail = None
        for detail_name, default_input_path in DEFAULT_INPUTS.items():
            if default_input_path == input_path:
                matched_detail = detail_name
                break
        if matched_detail:
            output_path = _default_output_path(matched_detail, input_path)
            out = standardize_workbook(input_path, matched_detail, config_path, output_path)
            print(out)


if __name__ == "__main__":
    main()
