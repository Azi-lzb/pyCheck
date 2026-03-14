"""
以国民经济行业小类编码（参照表 col13）为准，
全面校验并修正 五篇大文章参照表 中
门类/大类/中类 编码与名称的错误。
输出带时间戳的新文件，不覆盖原文件。
"""

import re
import glob
import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook

# excel-data 根目录；参照表与映射表在 data 下
EXCEL_DATA_ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = EXCEL_DATA_ROOT / "data"
REF_FILE = DATA_DIR / "五篇大文章与国民经济行业分类对应参照表20260312.xlsx"

# 优先用最新映射表
map_files = sorted(glob.glob(str(DATA_DIR / "国民经济行业分类映射表_*.xlsx")))
MAP_FILE = Path(map_files[-1]) if map_files else DATA_DIR / "国民经济行业分类映射表.xlsx"


# ── 1. 从映射表构建 数字4位码 → 层级信息 ──────────────────────
def build_lookup(map_path):
    """key: 4位数字字符串(如'8410'), value: dict of codes/names"""
    wb = load_workbook(str(map_path), read_only=True, data_only=True)
    ws = wb.active
    lookup = {}   # numeric4 → info
    full_lookup = {}  # full_code_upper(如'Q8410') → info

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 11:
            continue
        sec_code = str(row[0] or "").strip()   # col1 门类码
        sec_name = str(row[1] or "").strip()   # col2 门类名
        big_code = str(row[3] or "").strip()   # col4 大类码
        big_name = str(row[4] or "").strip()   # col5 大类名
        mid_code = str(row[6] or "").strip()   # col7 中类码
        mid_name = str(row[7] or "").strip()   # col8 中类名
        sml_code = str(row[9] or "").strip()   # col10 小类码
        sml_name = str(row[10] or "").strip()  # col11 小类名

        if not sml_code:
            continue

        star = "*" if "*" in sml_code else ""
        sml_clean = sml_code.replace("*", "").upper()
        big_clean = big_code.replace("*", "").upper()
        mid_clean = mid_code.replace("*", "").upper()

        # 提取纯数字部分
        m = re.search(r"(\d{4,5})", sml_clean)
        if not m:
            continue
        num4 = m.group(1)[:4]  # 只取前4位（统一到小类层级）

        info = {
            "sec_code": sec_code,
            "sec_name": sec_name,
            "big_code": big_clean,
            "big_name": big_name,
            "mid_code": mid_clean,
            "mid_name": mid_name,
            "sml_code": sml_clean,
            "sml_name": sml_name,
            "star": star,
        }
        # 用数字4位作为key（兜底）
        if num4 not in lookup:
            lookup[num4] = info
        # 用完整码作为精确key
        full_lookup[sml_clean] = info

    wb.close()
    return lookup, full_lookup


# ── 2. 比对参照表所有sheet ──────────────────────────────────────
def compare_and_fix(ref_path, lookup, full_lookup, dry_run=False):
    """
    返回: (mismatch_report, wb_modified)
    参照表列（1-based）:
      10=门类码  11=大类码  12=中类码  13=小类码
      14=门类名  15=大类名  16=中类名  17=小类名
    """
    wb = load_workbook(str(ref_path), data_only=True)
    report = []
    total_fixed = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        sheet_fixed = 0

        for r in range(2, ws.max_row + 1):
            sml_raw = ws.cell(r, 13).value
            if not sml_raw:
                continue

            sml_str = str(sml_raw).strip().upper()
            star = "*" if "*" in sml_str else ""
            sml_clean = sml_str.replace("*", "")

            # 提取4位数字
            m = re.search(r"(\d{4,5})", sml_clean)
            if not m:
                continue
            num4 = m.group(1)[:4]

            # 精确匹配 → 数字兜底
            info = full_lookup.get(sml_clean) or lookup.get(num4)
            if not info:
                continue

            # 读当前值
            cur_sec  = str(ws.cell(r, 10).value or "").strip()
            cur_big  = str(ws.cell(r, 11).value or "").strip()
            cur_mid  = str(ws.cell(r, 12).value or "").strip()
            cur_secn = str(ws.cell(r, 14).value or "").strip()
            cur_bign = str(ws.cell(r, 15).value or "").strip()
            cur_midn = str(ws.cell(r, 16).value or "").strip()
            cur_smln = str(ws.cell(r, 17).value or "").strip()

            # 期望值
            exp_sec  = info["sec_code"]
            exp_big  = info["big_code"]
            exp_mid  = info["mid_code"]
            exp_secn = info["sec_name"]
            exp_bign = info["big_name"]
            exp_midn = info["mid_name"]
            exp_smln = info["sml_name"]
            # 小类码本身：若前缀字母错了也需修正
            exp_sml  = info["sml_code"] + info["star"]

            issues = []
            if cur_sec.upper() != exp_sec.upper():
                issues.append(f"门类码: {cur_sec!r}→{exp_sec!r}")
            if cur_big.upper() != exp_big.upper():
                issues.append(f"大类码: {cur_big!r}→{exp_big!r}")
            if cur_mid.upper() != exp_mid.upper():
                issues.append(f"中类码: {cur_mid!r}→{exp_mid!r}")
            if exp_secn and cur_secn != exp_secn:
                issues.append(f"门类名: {cur_secn!r}→{exp_secn!r}")
            if exp_bign and cur_bign != exp_bign:
                issues.append(f"大类名: {cur_bign!r}→{exp_bign!r}")
            if exp_midn and cur_midn != exp_midn:
                issues.append(f"中类名: {cur_midn!r}→{exp_midn!r}")
            if exp_smln and cur_smln != exp_smln:
                issues.append(f"小类名: {cur_smln!r}→{exp_smln!r}")
            # 检查小类码字母前缀是否正确
            sml_letter = re.match(r"[A-Z]", sml_clean)
            if sml_letter and sml_letter.group(0) != exp_sec:
                issues.append(f"小类码前缀: {sml_clean!r}应改为{exp_sml!r}")

            if issues:
                c1_val = ws.cell(r, 1).value
                report.append({
                    "sheet": sn,
                    "row": r,
                    "小类码": sml_str,
                    "issues": issues,
                })

                if not dry_run:
                    ws.cell(r, 10).value = exp_sec
                    ws.cell(r, 11).value = exp_big
                    ws.cell(r, 12).value = exp_mid
                    if exp_secn:
                        ws.cell(r, 14).value = exp_secn
                    if exp_bign:
                        ws.cell(r, 15).value = exp_bign
                    if exp_midn:
                        ws.cell(r, 16).value = exp_midn
                    if exp_smln:
                        ws.cell(r, 17).value = exp_smln
                    # 若小类码字母前缀也错了，一并修正
                    if sml_letter and sml_letter.group(0) != exp_sec:
                        ws.cell(r, 13).value = exp_sml

                sheet_fixed += 1
                total_fixed += 1

        if sheet_fixed:
            print(f"  sheet [{sn}]: {sheet_fixed} 行存在问题")

    return report, wb, total_fixed


# ── 3. 主流程 ───────────────────────────────────────────────────
def main():
    print(f"映射表: {MAP_FILE.name}")
    print(f"参照表: {REF_FILE.name}\n")

    print("正在构建映射表查找字典...")
    lookup, full_lookup = build_lookup(MAP_FILE)
    print(f"  映射表条目数: 精确={len(full_lookup)}, 数字兜底={len(lookup)}\n")

    print("正在扫描参照表（dry_run=True）...")
    report, _, total = compare_and_fix(REF_FILE, lookup, full_lookup, dry_run=True)

    print(f"\n=== 不一致汇总（共 {total} 行） ===")
    by_sheet = {}
    for item in report:
        by_sheet.setdefault(item["sheet"], []).append(item)

    report_lines = []
    for sn, rows in by_sheet.items():
        report_lines.append(f"\n[sheet: {sn}]  ({len(rows)} rows)")
        for item in rows:
            report_lines.append(f"  row={item['row']:5d}  xiao_lei={item['小类码']:12s}  issues: {'; '.join(item['issues'])}")

    report_txt = EXCEL_DATA_ROOT / "fix_reference_report.txt"
    with open(report_txt, "w", encoding="utf-8") as f:
        f.write(f"总不一致行数: {total}\n")
        f.write("\n".join(report_lines))
    print(f"详细报告已保存: {report_txt.name}")

    # 控制台摘要（每sheet最多显示5条）
    for sn, rows in by_sheet.items():
        print(f"\n  [sheet: {sn}]  {len(rows)} rows")
        for item in rows[:5]:
            print(f"    row={item['row']}  {item['小类码']}  {'; '.join(item['issues'])}")
        if len(rows) > 5:
            print(f"    ... {len(rows)-5} more (see report file)")

    if total == 0:
        print("\n所有行层级编码一致，无需修正。")
        return

    print(f"\n开始生成修正文件...")
    _, wb_fixed, fixed = compare_and_fix(REF_FILE, lookup, full_lookup, dry_run=False)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = REF_FILE.with_name(f"{REF_FILE.stem}_{ts}{REF_FILE.suffix}")
    wb_fixed.save(str(out))
    wb_fixed.close()

    print(f"\n✅ 完成！修正行数: {fixed}")
    print(f"✅ 输出文件: {out.name}")


if __name__ == "__main__":
    main()
