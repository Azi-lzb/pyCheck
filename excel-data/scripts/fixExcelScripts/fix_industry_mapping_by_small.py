import re
from datetime import datetime
from pathlib import Path

import openpyxl
import xlrd


# excel-data 根目录；参照表与映射表在 data 下
EXCEL_DATA_ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = EXCEL_DATA_ROOT / "data"
SRC_XLSX = DATA_DIR / "国民经济行业分类映射表.xlsx"
REF_XLS = DATA_DIR / "2017国民经济行业分类注释（网络版）.xls"


def _norm_code(v):
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def _parse_small4(raw):
    s = "" if raw is None else str(raw).strip().upper()
    if not s:
        return "", "", ""
    star = "*" if "*" in s else ""
    s = s.replace("*", "")
    m = re.search(r"([A-Z])?(\d{4})", s)
    if not m:
        return "", "", ""
    return m.group(2), (m.group(1) or ""), star


def parse_reference():
    wb = xlrd.open_workbook(str(REF_XLS))
    ws = wb.sheet_by_name("Sheet1")

    section_names = {}
    big_names = {}
    mid_names = {}
    small_names = {}
    big_to_section = {}
    mid_to_big = {}
    small_to_mid = {}

    cur_section = ""
    cur_big = ""
    cur_mid = ""

    for r in range(1, ws.nrows):  # skip title row
        c0 = _norm_code(ws.cell_value(r, 0)).upper()
        c1 = _norm_code(ws.cell_value(r, 1)).upper()
        name = _norm_code(ws.cell_value(r, 3))

        if c0 and re.fullmatch(r"[A-Z]", c0):
            cur_section = c0
            if name:
                section_names[c0] = name
            continue

        if c0 and re.fullmatch(r"\d{2}", c0):
            cur_big = c0
            if cur_section:
                big_to_section[cur_big] = cur_section
            if name:
                big_names[cur_big] = name
            continue

        if c0 and re.fullmatch(r"\d{3}", c0):
            cur_mid = c0
            if cur_big:
                mid_to_big[cur_mid] = cur_big
            else:
                mid_to_big[cur_mid] = cur_mid[:2]
            if name:
                mid_names[cur_mid] = name
            continue

        if c1 and re.fullmatch(r"\d{4}", c1):
            small = c1
            mid = cur_mid if cur_mid and small.startswith(cur_mid) else small[:3]
            small_to_mid[small] = mid
            if name:
                small_names[small] = name
            if mid not in mid_to_big:
                mid_to_big[mid] = mid[:2]

    return {
        "section_names": section_names,
        "big_names": big_names,
        "mid_names": mid_names,
        "small_names": small_names,
        "big_to_section": big_to_section,
        "mid_to_big": mid_to_big,
        "small_to_mid": small_to_mid,
    }


def main():
    if not SRC_XLSX.exists():
        raise FileNotFoundError(f"找不到源文件: {SRC_XLSX}")
    if not REF_XLS.exists():
        raise FileNotFoundError(f"找不到参考文件: {REF_XLS}")

    ref = parse_reference()
    wb = openpyxl.load_workbook(SRC_XLSX)
    ws = wb.active

    changed_rows = 0
    unresolved = []

    for r in range(2, ws.max_row + 1):
        raw_small = ws.cell(r, 10).value
        small4, raw_prefix, star = _parse_small4(raw_small)
        if not small4:
            continue

        mid3 = ref["small_to_mid"].get(small4, small4[:3])
        big2 = ref["mid_to_big"].get(mid3, mid3[:2])
        section = (
            ref["big_to_section"].get(big2)
            or raw_prefix
            or (str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else "")
        )

        if not section or not re.fullmatch(r"[A-Z]", section):
            unresolved.append((r, raw_small))
            continue

        new_sec_code = section
        new_big_code = f"{section}{big2}"
        new_mid_code = f"{section}{mid3}"
        new_small_code = f"{section}{small4}{star}"

        old_vals = [
            ws.cell(r, 1).value,
            ws.cell(r, 2).value,
            ws.cell(r, 4).value,
            ws.cell(r, 5).value,
            ws.cell(r, 7).value,
            ws.cell(r, 8).value,
            ws.cell(r, 10).value,
            ws.cell(r, 11).value,
        ]

        ws.cell(r, 1).value = new_sec_code
        if ref["section_names"].get(section):
            ws.cell(r, 2).value = ref["section_names"][section]

        ws.cell(r, 4).value = new_big_code
        if ref["big_names"].get(big2):
            ws.cell(r, 5).value = ref["big_names"][big2]

        ws.cell(r, 7).value = new_mid_code
        if ref["mid_names"].get(mid3):
            ws.cell(r, 8).value = ref["mid_names"][mid3]

        ws.cell(r, 10).value = new_small_code
        if ref["small_names"].get(small4):
            ws.cell(r, 11).value = ref["small_names"][small4]

        new_vals = [
            ws.cell(r, 1).value,
            ws.cell(r, 2).value,
            ws.cell(r, 4).value,
            ws.cell(r, 5).value,
            ws.cell(r, 7).value,
            ws.cell(r, 8).value,
            ws.cell(r, 10).value,
            ws.cell(r, 11).value,
        ]
        if old_vals != new_vals:
            changed_rows += 1

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = SRC_XLSX.with_name(f"{SRC_XLSX.stem}_{ts}{SRC_XLSX.suffix}")
    wb.save(out)
    wb.close()

    print(f"输出文件: {out}")
    print(f"总行数(不含表头): {ws.max_row - 1}")
    print(f"发生更新行数: {changed_rows}")
    print(f"未能识别门类行数: {len(unresolved)}")
    if unresolved:
        print("未识别示例(前10):")
        for row_no, raw in unresolved[:10]:
            print(f"  row={row_no}, 小类原值={raw}")


if __name__ == "__main__":
    main()
