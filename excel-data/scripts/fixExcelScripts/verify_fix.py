import openpyxl
import glob
import re
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
files = sorted(glob.glob(str(DATA_DIR / "国民经济行业分类映射表_*.xlsx")))
if not files:
    print("未找到输出文件")
    exit()
f = files[-1]
print(f"检验文件: {f}\n")
wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
ws = wb.active

# ── 1. M73/M741/M7310 专项检查 ──────────────────────────────
print("=== M73 / M741 / M7310 专项行 ===")
targets = {"M73", "M731", "M7310", "M741"}
found_m = False
for r in range(2, ws.max_row + 1):
    v10 = ws.cell(r, 10).value
    if not v10:
        continue
    s = str(v10).upper().replace("*", "")
    if any(s == t or s.startswith(t) for t in targets):
        sec = ws.cell(r, 1).value
        big = ws.cell(r, 4).value
        mid = ws.cell(r, 7).value
        sml = ws.cell(r, 10).value
        bn  = ws.cell(r, 5).value
        mn  = ws.cell(r, 8).value
        smn = ws.cell(r, 11).value
        print(f"  row={r:4d} | 门类={sec} | 大类={big}({bn}) | 中类={mid}({mn}) | 小类={sml}({smn})")
        found_m = True
if not found_m:
    print("  （映射表中未发现 M73/M741/M7310 编码行）")

# ── 2. 全表层级一致性校验 ────────────────────────────────────
print("\n=== 全表层级一致性校验 ===")
mismatch_big = []
mismatch_mid = []

for r in range(2, ws.max_row + 1):
    big = str(ws.cell(r, 4).value or "").strip().replace("*", "")
    mid = str(ws.cell(r, 7).value or "").strip().replace("*", "")
    sml = str(ws.cell(r, 10).value or "").strip().replace("*", "")
    if not sml:
        continue
    # sml like A0111, big like A01, mid like A011
    m = re.fullmatch(r"([A-Z])(\d{2})(\d{1})(\d{1,2})", sml)
    if not m:
        continue
    expected_big = m.group(1) + m.group(2)   # A01
    expected_mid = m.group(1) + m.group(2) + m.group(3)  # A011
    if big and big != expected_big:
        mismatch_big.append((r, big, expected_big, sml))
    if mid and mid != expected_mid:
        mismatch_mid.append((r, mid, expected_mid, sml))

print(f"  大类与小类不一致: {len(mismatch_big)} 行")
for row, got, exp, sml in mismatch_big[:5]:
    print(f"    row={row}, 大类={got}, 应为={exp}, 小类={sml}")

print(f"  中类与小类不一致: {len(mismatch_mid)} 行")
for row, got, exp, sml in mismatch_mid[:10]:
    print(f"    row={row}, 中类={got}, 应为={exp}, 小类={sml}")

# ── 3. 策略一兼容性检查 ───────────────────────────────────────
print("\n=== 策略一兼容性 ===")
print("  build_strategy1.py 使用 re.search(r'[A-Z]\\d{4}', text) 提取行业码（extract_industry4）")
print("  映射表读取时, J列(col10)作为key, 若编码是 A0111, key=A0111, 同时 key[1:]='0111' 也被注册")
sample_codes = []
for r in range(2, ws.max_row + 1):
    v = ws.cell(r, 10).value
    if v:
        sample_codes.append(str(v).strip())
    if len(sample_codes) >= 5:
        break

print(f"  前5个小类码样本: {sample_codes}")

# 模拟 load_industry_desc_map 逻辑
d = {}
for r in range(2, ws.max_row + 1):
    code_raw = ws.cell(r, 10).value
    desc_raw = ws.cell(r, 11).value
    if code_raw is None:
        continue
    display_code = str(code_raw).strip()
    code = display_code.upper()
    if not code:
        continue
    desc = "" if desc_raw is None else str(desc_raw)
    d[code] = (display_code, desc)
    if len(code) == 5 and code[0].isalpha() and code[1:].isdigit():
        d[code[1:]] = (display_code, desc)

test_keys = ["A0111", "0111", "C2661", "2661", "M7310"]
print("  查字典测试:")
for k in test_keys:
    hit = d.get(k.upper())
    status = "✓" if hit else "✗"
    print(f"    [{status}] key={k!r} -> {hit[0] if hit else '无匹配'}")

wb.close()
print("\n验证完成。")
