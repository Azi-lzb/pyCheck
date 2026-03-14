import openpyxl
import re
import glob
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
files = sorted(glob.glob(str(DATA_DIR / "国民经济行业分类映射表_*.xlsx")))
f = files[-1]
wb = openpyxl.load_workbook(f, data_only=True)
ws = wb.active

d = {}
for r in range(2, ws.max_row + 1):
    code_raw = ws.cell(r, 10).value
    desc_raw = ws.cell(r, 11).value
    if not code_raw:
        continue
    display_code = str(code_raw).strip()
    code = display_code.upper()
    d[code] = (display_code, str(desc_raw or "")[:12])
    if len(code) == 5 and code[0].isalpha() and code[1:].isdigit():
        d[code[1:]] = (display_code, str(desc_raw or "")[:12])

test_keys = ["A0111", "0111", "C2661", "2661", "M7310", "7310", "M7410", "7410"]
print("=== 策略一字典查询测试 ===")
for k in test_keys:
    hit = d.get(k.upper())
    status = "OK  " if hit else "MISS"
    disp = hit[0] if hit else "无匹配"
    print(f"  [{status}] key={k!r:10s} -> {disp}")

print()
print("=== extract_industry4 正则 [A-Z]\\d{4} 兼容测试 ===")
samples = ["A0111科技", "C2661某行业", "M7310研发", "0111农业（纯数字无字母前缀）"]
for s in samples:
    m = re.search(r"[A-Z]\d{4}", s.upper())
    result = m.group(0) if m else "无匹配（数字开头码无法提取）"
    print(f"  input={s!r:30s} -> {result}")

print()
print("=== M73/M741/M7310 层级关系 ===")
for r in range(2, ws.max_row + 1):
    sml = str(ws.cell(r, 10).value or "").strip().upper().replace("*", "")
    if sml in ("M7310", "M7320", "M7330", "M7340", "M7350", "M7410"):
        sec = ws.cell(r, 1).value
        big = ws.cell(r, 4).value
        mid = ws.cell(r, 7).value
        bn  = str(ws.cell(r, 5).value or "")
        mn  = str(ws.cell(r, 8).value or "")
        smn = str(ws.cell(r, 11).value or "")
        print(f"  row={r} | 门={sec} | 大={big}({bn[:8]}) | 中={mid}({mn[:8]}) | 小={sml}({smn[:12]})")

wb.close()
