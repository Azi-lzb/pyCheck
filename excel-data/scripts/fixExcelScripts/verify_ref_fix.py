import openpyxl
import re
import glob
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
files = sorted(glob.glob(str(DATA_DIR / "五篇大文章*_2026*.xlsx")))
f = files[-1]
print("验证文件:", f)
wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

checks = {
    "84大类行": [],
    "P门类下8xxx": [],
    "K门类下7xxx": [],
}

for sn in wb.sheetnames:
    ws = wb[sn]
    for r in range(2, ws.max_row + 1):
        sec = str(ws.cell(r, 10).value or "").strip().upper()
        big = str(ws.cell(r, 11).value or "").strip().upper()
        mid = str(ws.cell(r, 12).value or "").strip().upper()
        sml = str(ws.cell(r, 13).value or "").strip().upper()
        sn14 = ws.cell(r, 14).value
        sn15 = ws.cell(r, 15).value
        sn17 = ws.cell(r, 17).value

        if re.search(r"[A-Z]84", big):
            checks["84大类行"].append((sn, r, sec, big, mid, sml, sn14, sn15))

        if sec == "P" and re.search(r"[A-Z]8[0-9]\d{2}", sml):
            checks["P门类下8xxx"].append((sn, r, sec, big, sml))

        if sec == "K" and re.search(r"[A-Z]7[0-9]\d{2}", sml):
            checks["K门类下7xxx"].append((sn, r, sec, big, sml))

wb.close()

print("\n=== 84大类行 ===")
if checks["84大类行"]:
    for sn, r, sec, big, mid, sml, sn14, sn15 in checks["84大类行"]:
        print(f"  [{sn}] row={r} | 门={sec} 大={big} 中={mid} 小={sml} | 门名={sn14} 大名={sn15}")
else:
    print("  (无)")

print("\n=== P门类下8xxx (教育门类不应含84xx) ===")
if checks["P门类下8xxx"]:
    for x in checks["P门类下8xxx"]:
        print(f"  {x}")
else:
    print("  OK - P门类下无8xxx")

print("\n=== K门类下7xxx (金融不应含71xx房地产) ===")
if checks["K门类下7xxx"]:
    for x in checks["K门类下7xxx"][:10]:
        print(f"  {x}")
else:
    print("  OK - K门类下无7xxx")

print("\n验证完成。")
