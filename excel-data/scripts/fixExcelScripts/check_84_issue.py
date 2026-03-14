import openpyxl
import re
import glob
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data"
REF_FILE = DATA_DIR / "五篇大文章与国民经济行业分类对应参照表20260312.xlsx"
wb = openpyxl.load_workbook(str(REF_FILE), read_only=True, data_only=True)

# 列头确认
ws0 = wb.worksheets[0]
headers = [ws0.cell(1, c).value for c in range(1, 22)]
print("=== 第一sheet列头 ===")
for i, h in enumerate(headers, 1):
    print(f"  col{i}: {h}")

# 所有sheet中找84大类行
print("\n=== 大类含'84'的行（前20条） ===")
count = 0
for sn in wb.sheetnames:
    ws = wb[sn]
    for r in range(2, ws.max_row + 1):
        c10 = str(ws.cell(r, 10).value or "")
        c11 = str(ws.cell(r, 11).value or "")
        c12 = str(ws.cell(r, 12).value or "")
        c13 = str(ws.cell(r, 13).value or "")
        if "84" in c11 and re.search(r"[A-Z]84", c11.upper()):
            sec  = ws.cell(r, 10).value
            big  = ws.cell(r, 11).value
            mid  = ws.cell(r, 12).value
            sml  = ws.cell(r, 13).value
            sn14 = ws.cell(r, 14).value
            sn15 = ws.cell(r, 15).value
            sn17 = ws.cell(r, 17).value
            print(f"  sheet={sn} row={r} | 门={sec} | 大={big}({sn15}) | 中={mid} | 小={sml}({sn17})")
            count += 1
            if count >= 20:
                break
    if count >= 20:
        break
wb.close()
