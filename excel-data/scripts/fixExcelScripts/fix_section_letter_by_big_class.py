"""
根据《大类-门类对应.xlsx》中的“大类编码→门类字母”映射，
修正《五篇大文章与国民经济行业分类对应参照表20260312.xlsx》中
J/K/L/M/U 列里国民经济行业编码的门类字母前缀。

修正规则（逐行）：
1) 以 K 列（对应国民经济行业编码大类，如 C39 / M76）提取2位“大类数字编码”；
   若 K 为空，则依次从 L/M 列提取；
2) 用“大类数字编码”在映射表中找到期望门类字母（如 76 → N）；
3) 修正：
   - J 列：直接改为期望门类字母
   - K/L/M 列：若以字母开头，则将首字母替换为期望门类字母
   - U 列：将文本中与当前大类数字编码匹配的形如 [A-Z]\\d{2,5} 的码的首字母替换为期望门类字母

输出：生成带时间戳的新文件，不覆盖原文件；并输出一个简要 CSV 报告。
"""

from __future__ import annotations

import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

# excel-data 根目录；参照表与映射表在 data 下
EXCEL_DATA_ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = EXCEL_DATA_ROOT / "data"
REF_FILE = DATA_DIR / "五篇大文章与国民经济行业分类对应参照表20260312.xlsx"
MAP_FILE = DATA_DIR / "大类-门类对应.xlsx"
REPO_ROOT = EXCEL_DATA_ROOT.parent

# 参照表列（1-based）
COL_J_SECTION = 10
COL_K_BIG = 11
COL_L_MID = 12
COL_M_SMALL = 13
COL_U_RAW_TYPE = 21

_RE_CODE_LEADING = re.compile(r"^([A-Z])(.+)$")
_RE_DIGITS = re.compile(r"(\d{2,})")
_RE_INLINE_CODE = re.compile(r"\b([A-Z])(\d{2,5})\b")
_RE_SECTION_PAREN = re.compile(r"(门类\s*[（(])\s*([A-Z])\s*([)）])", re.IGNORECASE)


def _norm_big_code(big_code: str) -> str | None:
    s = (big_code or "").strip()
    m = _RE_DIGITS.search(s)
    if not m:
        return None
    digits = m.group(1)
    if len(digits) < 2:
        return None
    return digits[:2].zfill(2)


def load_big_to_section(map_path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(map_path, data_only=True)
    ws = wb.active
    mapping: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        big = ws.cell(r, 1).value
        sec = ws.cell(r, 2).value
        if big is None or sec is None:
            continue
        big_s = str(big).strip()
        sec_s = str(sec).strip().upper()
        if not big_s:
            continue
        big2 = _norm_big_code(big_s)
        if not big2:
            continue
        if not re.fullmatch(r"[A-Z]", sec_s):
            continue
        mapping[big2] = sec_s
    wb.close()
    return mapping


def _replace_leading_letter(value: object, expected_letter: str) -> tuple[object, bool]:
    if value is None:
        return value, False
    s = str(value).strip()
    if not s:
        return value, False
    m = _RE_CODE_LEADING.match(s.upper())
    if not m:
        return value, False
    cur_letter, rest = m.group(1), m.group(2)
    if cur_letter == expected_letter:
        return value, False
    return f"{expected_letter}{rest}", True


def _fix_u_text(
    value: object,
    *,
    expected_letter_from_big2: str | None,
    expected_letter_from_j: str | None,
    big2: str | None,
) -> tuple[object, bool]:
    if value is None:
        return value, False
    s = str(value)
    if not s.strip():
        return value, False

    changed = False

    # 1) 修正 “门类(B)” 这类不带数字的情况：优先用 J 列门类字母
    sec_for_u = expected_letter_from_j or expected_letter_from_big2
    out = s
    if sec_for_u:

        def repl_section_paren(m: re.Match[str]) -> str:
            nonlocal changed
            prefix, letter, suffix = m.group(1), m.group(2), m.group(3)
            if letter.upper() != sec_for_u:
                changed = True
                return f"{prefix}{sec_for_u}{suffix}"
            return m.group(0)

        out = _RE_SECTION_PAREN.sub(repl_section_paren, out)

    # 2) 修正 “小类(C3911)” 这类带数字的码：仅在能确定 big2 且有映射时处理
    if not big2 or not expected_letter_from_big2:
        return out, changed

    def repl_inline(m: re.Match[str]) -> str:
        nonlocal changed
        letter, digits = m.group(1), m.group(2)
        if digits.startswith(big2) and letter.upper() != expected_letter_from_big2:
            changed = True
            return f"{expected_letter_from_big2}{digits}"
        return m.group(0)

    out2 = _RE_INLINE_CODE.sub(repl_inline, out)
    return out2, changed


def fix_reference_letters(ref_path: Path, mapping: dict[str, str]) -> tuple[Path, Path, int]:
    wb = openpyxl.load_workbook(ref_path, data_only=False)
    changes: list[dict[str, object]] = []
    total_cells_changed = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            # J 列门类字母（弱依赖：即使大类数字取不到，也可用于修正 U 的“门类(X)”）
            cur_j = ws.cell(r, COL_J_SECTION).value
            cur_j_s = str(cur_j).strip().upper() if cur_j is not None else ""
            expected_from_j = cur_j_s if re.fullmatch(r"[A-Z]", cur_j_s or "") else None

            # 以 K→L→M 取大类两位数字
            big2 = (
                _norm_big_code(str(ws.cell(r, COL_K_BIG).value or ""))
                or _norm_big_code(str(ws.cell(r, COL_L_MID).value or ""))
                or _norm_big_code(str(ws.cell(r, COL_M_SMALL).value or ""))
            )
            expected_from_big2 = mapping.get(big2) if big2 else None

            row_changed = False

            # J
            if expected_from_big2 and cur_j_s != expected_from_big2:
                ws.cell(r, COL_J_SECTION).value = expected_from_big2
                total_cells_changed += 1
                row_changed = True

            # K/L/M
            if expected_from_big2:
                for c in (COL_K_BIG, COL_L_MID, COL_M_SMALL):
                    cell = ws.cell(r, c)
                    new_v, ch = _replace_leading_letter(cell.value, expected_from_big2)
                    if ch:
                        cell.value = new_v
                        total_cells_changed += 1
                        row_changed = True

            # U
            cell_u = ws.cell(r, COL_U_RAW_TYPE)
            new_u, ch_u = _fix_u_text(
                cell_u.value,
                expected_letter_from_big2=expected_from_big2,
                expected_letter_from_j=expected_from_j,
                big2=big2,
            )
            if ch_u:
                cell_u.value = new_u
                total_cells_changed += 1
                row_changed = True

            if row_changed:
                changes.append(
                    {
                        "sheet": sheet_name,
                        "row": r,
                        "big2": big2 or "",
                        "expected_section": expected_from_big2 or expected_from_j or "",
                    }
                )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = DATA_DIR / f"{ref_path.stem}_门类修正U_{ts}{ref_path.suffix}"
    out_csv = DATA_DIR / f"{ref_path.stem}_门类修正U_{ts}_report.csv"
    wb.save(out_xlsx)
    wb.close()

    with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["sheet", "row", "big2", "expected_section"])
        w.writeheader()
        w.writerows(changes)

    return out_xlsx, out_csv, total_cells_changed


def _pick_default_input() -> Path:
    # 先用 data 下最新“门类修正”结果，再尝试仓库根目录，最后兜底原始参照表
    pat = "五篇大文章与国民经济行业分类对应参照表20260312_门类修正_*.xlsx"
    data_candidates = sorted(DATA_DIR.glob(pat), key=lambda p: p.stat().st_mtime)
    if data_candidates:
        return data_candidates[-1]
    root_candidates = sorted(REPO_ROOT.glob(pat), key=lambda p: p.stat().st_mtime)
    if root_candidates:
        return root_candidates[-1]
    return REF_FILE


def main() -> None:
    if not MAP_FILE.exists():
        raise FileNotFoundError(f"映射表不存在: {MAP_FILE}")

    in_path = _pick_default_input()
    args = sys.argv[1:]
    if len(args) >= 2 and args[0] == "--in":
        in_path = Path(args[1])

    if not in_path.exists():
        raise FileNotFoundError(f"参照表不存在: {in_path}")

    mapping = load_big_to_section(MAP_FILE)
    print(f"映射表条目数: {len(mapping)}")

    print(f"输入文件：{in_path}")
    out_xlsx, out_csv, total_cells_changed = fix_reference_letters(in_path, mapping)
    print(f"完成：改动单元格数 = {total_cells_changed}")
    print(f"输出文件：{out_xlsx}")
    print(f"报告文件：{out_csv}")


if __name__ == "__main__":
    main()
