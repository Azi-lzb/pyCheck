"""
调试：找 HT2026000/JD2026000 这条借据，分析 HTS06 为何被判为未填报
（临时脚本，结果文件在 main/ 下）
"""
import os
from pathlib import Path
import openpyxl
import re

# 结果文件在 main/ 目录
MAIN_DIR = Path(__file__).resolve().parent.parent / "main"
RESULT_FILE = MAIN_DIR / "策略一核查结果.xlsx"
wb = openpyxl.load_workbook(str(RESULT_FILE), read_only=True, data_only=True)

for sn in wb.sheetnames:
    ws = wb[sn]
    # 先找表头行
    header_row = None
    headers = []
    for r in range(1, 10):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        strs = [str(v or "") for v in vals]
        if any("借据" in s or "合同" in s for s in strs):
            header_row = r
            headers = strs
            break
    if header_row is None:
        continue

    # 找 HT2026000 或 JD2026000
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        row_str = " ".join(str(v or "") for v in vals)
        if "HT2026000" in row_str or "JD2026000" in row_str:
            print(f"\n=== sheet={sn} row={r} ===")
            for i, (h, v) in enumerate(zip(headers, vals), 1):
                if v is not None and str(v).strip():
                    h_short = str(h)[:30]
                    print(f"  col{i:2d} [{h_short}] = {str(v)[:80]}")

wb.close()
