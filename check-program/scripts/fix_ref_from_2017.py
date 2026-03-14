import re
from datetime import datetime
from pathlib import Path

import openpyxl
import xlrd


# 参照表与 2017 注释在项目根下的 excel-data/data（本脚本在 check-program/scripts）
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
EXCEL_DATA_DIR = PROJECT_ROOT / "excel-data" / "data"
SRC_REF = EXCEL_DATA_DIR / "五篇大文章与国民经济行业分类对应参照表20260312.xlsx"
REF_2017_XLS = EXCEL_DATA_DIR / "2017国民经济行业分类注释（网络版）.xls"


def norm_code(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s


def build_small_mapping():
    """
    从 2017 国民经济行业分类注释（网络版）.xls 构建：
    4位数字小类码 -> {sec, big2, mid3, name_small, name_mid, name_big}
    """
    wb = xlrd.open_workbook(str(REF_2017_XLS))
    ws = wb.sheet_by_name("Sheet1")

    section = ""
    big2 = ""
    mid3 = ""
    big_name = ""
    mid_name = ""

    small_map = {}

    for r in range(1, ws.nrows):
        c0 = norm_code(ws.cell_value(r, 0)).upper()
        name = norm_code(ws.cell_value(r, 3))

        if c0 and re.fullmatch(r"[A-Z]", c0):
            section = c0
            continue

        if c0 and re.fullmatch(r"\d{2}", c0):
            big2 = c0
            big_name = name
            continue

        if c0 and re.fullmatch(r"\d{3}", c0):
            mid3 = c0
            mid_name = name
            continue

        # 小类码在第2列（见之前调试）
        c1 = norm_code(ws.cell_value(r, 1)).upper()
        if c1 and re.fullmatch(r"\d{4}", c1):
            small4 = c1
            small_name = name
            sec = section
            b2 = big2 or small4[:2]
            m3 = mid3 or small4[:3]
            small_map[small4] = {
                "sec": sec,
                "big2": b2,
                "mid3": m3,
                "name_small": small_name,
                "name_mid": mid_name,
                "name_big": big_name,
            }

    wb.release_resources()
    return small_map


def main():
    if not SRC_REF.exists():
        raise FileNotFoundError(f"找不到源参照表: {SRC_REF}")
    if not REF_2017_XLS.exists():
        raise FileNotFoundError(f"找不到 2017 注释文件: {REF_2017_XLS}")

    small_map = build_small_mapping()
    print("小类映射条数:", len(small_map))

    wb = openpyxl.load_workbook(str(SRC_REF))
    total_changed = 0

    for sn in wb.sheetnames:
        ws = wb[sn]
        sheet_changed = 0
        for r in range(2, ws.max_row + 1):
            sml_raw = ws.cell(r, 13).value
            if not sml_raw:
                continue
            sml_str = str(sml_raw).strip().upper()
            m = re.search(r"(\d{4})", sml_str)
            if not m:
                continue
            small4 = m.group(1)
            info = small_map.get(small4)
            if not info:
                continue

            sec = info["sec"]
            big2 = info["big2"]
            mid3 = info["mid3"]

            if not sec or not big2 or not mid3:
                continue

            # 期望编码
            exp_sec = sec
            exp_big = sec + big2
            exp_mid = sec + mid3
            exp_sml = sec + small4

            cur_sec = str(ws.cell(r, 10).value or "").strip().upper()
            cur_big = str(ws.cell(r, 11).value or "").strip().upper()
            cur_mid = str(ws.cell(r, 12).value or "").strip().upper()
            cur_sml = str(ws.cell(r, 13).value or "").strip().upper()

            changed = False

            if cur_sec != exp_sec:
                ws.cell(r, 10).value = exp_sec
                changed = True
            if cur_big != exp_big:
                ws.cell(r, 11).value = exp_big
                changed = True
            if cur_mid != exp_mid:
                ws.cell(r, 12).value = exp_mid
                changed = True
            if cur_sml != exp_sml:
                ws.cell(r, 13).value = exp_sml
                changed = True

            # 名称列：14=门类名,15=大类名,16=中类名,17=小类名
            if info["name_big"]:
                ws.cell(r, 15).value = info["name_big"]
            if info["name_mid"]:
                ws.cell(r, 16).value = info["name_mid"]
            if info["name_small"]:
                ws.cell(r, 17).value = info["name_small"]

            # 原始映射类别（U列=21）：只改括号里的字母，不动数字和中文
            raw_map = ws.cell(r, 21).value
            if raw_map:
                s = str(raw_map)

                def repl(mo):
                    # mo.group(1): 旧字母, group(2): 可能的数字
                    digits = mo.group(2) or ""
                    return f"({exp_sec}{digits})"

                new_s = re.sub(r"\(([A-Z])(\d*)\)", repl, s)
                if new_s != s:
                    ws.cell(r, 21).value = new_s
                    changed = True

            if changed:
                sheet_changed += 1
                total_changed += 1

        print(f"sheet [{sn}] 修改行数: {sheet_changed}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = SRC_REF.with_name(f"{SRC_REF.stem}_{ts}{SRC_REF.suffix}")
    wb.save(str(out))
    wb.close()
    print("总修改行数:", total_changed)
    print("已保存为:", out)


if __name__ == "__main__":
    main()

