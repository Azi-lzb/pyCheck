from pathlib import Path
import sys

import openpyxl
from openpyxl import Workbook


MAIN_DIR = Path(__file__).resolve().parents[1]
if str(MAIN_DIR) not in sys.path:
    sys.path.insert(0, str(MAIN_DIR))

import build_strategy1 as strategy1


def _make_runtime_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "标准化配置-明细定义"
    ws.append(strategy1.DETAIL_DEFINITION_ALL_HEADERS)
    ws.append(
        [
            1,
            "科技产业",
            "fake",
            "科技产业贷款",
            3,
            4,
            "",
            "",
            "科技产业贷款明细",
            "fake.xlsx",
            "科技",
            "万元",
            4,
            2,
            13,
            19,
            21,
        ]
    )
    detail_params = wb.create_sheet("标准化配置-明细参数")
    detail_params.append(strategy1.DETAIL_PARAM_HEADERS)
    for idx, row in enumerate(
        [
            ("科技产业", "贷款客户行业列序号", 6),
            ("科技产业", "贷款投向行业列序号", 15),
            ("科技产业", "贷款余额列序号", 8),
            ("科技产业", "机构报送产业分类列序号", "17,20,23,26"),
            ("科技产业", "报送列-类别映射", "17:HTP,20:HTS,23:SE,26:PA"),
        ],
        start=1,
    ):
        detail_params.append([idx, row[0], row[1], row[2], "exact", "", "", "2026-03-27"])
    runtime = wb.create_sheet("runtime")
    runtime.append(["配置项", "值", "说明"])
    runtime.append(["启用核查口径", "实际投向行业,客户主营行业", ""])
    runtime.append(["启用结果类型", "错报,漏报,疑似漏报,多报,疑似正确", ""])
    runtime.append(["汇总报告单位", "亿元", ""])
    runtime.append(["模板行-线索合计统计项", "错报,多报,疑似多报,疑似正确", ""])
    runtime.append(["模板行-按投向有误统计项", "错报,多报", ""])
    runtime.append(["模板行-按投向疑似无误统计项", "疑似正确,疑似多报", ""])
    runtime.append(["模板行-按主营有误统计项", "错报,多报", ""])
    runtime.append(["模板行-按主营疑似无误统计项", "", ""])
    runtime.append(["模板行-关键字", "关键字", ""])
    wb.save(path)
    wb.close()


def test_load_runtime_settings_reads_switches_and_units(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    _make_runtime_workbook(cfg)
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))

    settings = strategy1.load_runtime_settings()

    assert settings["enabled_bases"] == ["actual", "customer"]
    assert settings["enabled_result_types"] == ["错报", "漏报", "疑似漏报", "多报", "疑似正确"]
    assert settings["summary_unit"] == "亿元"
    assert settings["template_row_labels"]["线索合计"] == ["错报", "多报", "疑似多报", "疑似正确"]
    assert settings["template_row_labels"]["按投向有误"] == ["错报", "多报"]
    assert settings["template_row_labels"]["按投向疑似无误"] == ["疑似正确", "疑似多报"]
    assert settings["template_row_labels"]["按主营有误"] == ["错报", "多报"]
    assert settings["template_row_labels"]["按主营疑似无误"] == []
    assert settings["template_row_labels"]["关键字"] == ["关键字"]


def test_load_runtime_settings_supports_prefixed_runtime_sheet(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    runtime = wb.create_sheet("输出报告配置-runtime")
    runtime.append(["配置项", "值", "说明"])
    runtime.append(["启用核查口径", "实际投向行业", ""])
    runtime.append(["启用结果类型", "错报,多报", ""])
    runtime.append(["汇总报告单位", "亿元", ""])
    wb.save(cfg)
    wb.close()
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))

    settings = strategy1.load_runtime_settings()

    assert settings["enabled_bases"] == ["actual"]
    assert settings["enabled_result_types"] == ["错报", "多报"]


def test_load_config_rows_reads_customer_and_balance_columns(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    _make_runtime_workbook(cfg)
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))

    rows = strategy1.load_config_rows()

    assert rows[0]["贷款客户行业列序号"] == 6
    assert rows[0]["贷款投向行业列序号"] == 15
    assert rows[0]["贷款余额列序号"] == 8
    assert rows[0]["贷款余额原始单位"] == "万元"


def test_load_config_rows_supports_prefixed_config_sheet(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "线索规则配置-config"
    ws.append(strategy1.CONFIG_HEADERS)
    ws.append(
        [
            "科技产业贷款明细",
            "fake.xlsx",
            "科技",
            3,
            4,
            6,
            15,
            8,
            "万元",
            "17,20,23,26",
            "17:HTP,20:HTS,23:SE,26:PA",
            4,
            2,
            13,
            19,
            21,
        ]
    )
    wb.save(cfg)
    wb.close()
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))

    rows = strategy1.load_config_rows()

    assert rows[0]["输出工作表名称"] == "科技产业贷款明细"


def test_load_config_rows_can_build_runtime_rows_from_detail_sheets_only(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "标准化配置-明细定义"
    ws.append(strategy1.DETAIL_DEFINITION_HEADERS + strategy1.CONFIG_HEADERS[:3] + strategy1.CONFIG_HEADERS[8:])
    ws.append(
        [
            1,
            "科技产业",
            "fake",
            "科技产业贷款",
            3,
            4,
            "",
            "",
            "科技产业贷款明细",
            "fake.xlsx",
            "科技",
            "万元",
            4,
            2,
            13,
            19,
            21,
        ]
    )
    ws_param = wb.create_sheet("标准化配置-明细参数")
    ws_param.append(strategy1.DETAIL_PARAM_HEADERS)
    ws_param.append([1, "科技产业", "贷款客户行业列序号", 6, "exact", "", "", "2026-03-27"])
    ws_param.append([2, "科技产业", "贷款投向行业列序号", 15, "exact", "", "", "2026-03-27"])
    ws_param.append([3, "科技产业", "贷款余额列序号", 8, "exact", "", "", "2026-03-27"])
    ws_param.append([4, "科技产业", "机构报送产业分类列序号", "17,20,23,26", "exact", "", "", "2026-03-27"])
    ws_param.append([5, "科技产业", "报送列-类别映射", "17:HTP,20:HTS,23:SE,26:PA", "exact", "", "", "2026-03-27"])
    wb.save(cfg)
    wb.close()
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))

    rows = strategy1.load_config_rows()

    assert rows[0]["输出工作表名称"] == "科技产业贷款明细"
    assert rows[0]["来源台账文件"] == "fake.xlsx"
    assert rows[0]["台账类型"] == "tech"
    assert rows[0]["贷款余额原始单位"] == "万元"
    assert rows[0]["机构报送产业分类列序号"] == [17, 20, 23, 26]


def test_write_template_summary_sheets_uses_runtime_row_configs():
    wb = Workbook()
    wb.remove(wb.active)
    summary_amounts = {
        ("科技产业", strategy1._summary_group_key("01 高技术制造业")): {
            "amount": {
                "total_amount": 10000,
                "line": 6000,
                "actual_error": 5000,
                "suspect_ok": 1000,
                "customer_error": 0,
                "customer_all": 0,
            },
            "count": {
                "total_count": 4,
                "line": 2,
                "actual_error": 1,
                "suspect_ok": 1,
                "customer_error": 0,
                "customer_all": 0,
            },
        },
        ("科技产业", "HTP01"): {
            "amount": {
                "total_amount": 10000,
                "line": 6000,
                "actual_error": 5000,
                "suspect_ok": 1000,
                "customer_error": 0,
                "customer_all": 0,
            },
            "count": {
                "total_count": 4,
                "line": 2,
                "actual_error": 1,
                "suspect_ok": 1,
                "customer_error": 0,
                "customer_all": 0,
            },
        }
    }
    runtime = {
        "summary_unit": "万元",
        "template_row_labels": {
            "线索合计": ["错报", "多报", "疑似多报", "疑似正确"],
            "按投向有误": ["错报", "多报"],
            "按投向疑似无误": ["疑似正确", "疑似多报"],
            "按主营有误": ["错报", "多报"],
            "按主营疑似无误": [],
            "关键字": ["关键字"],
        },
    }

    strategy1.write_template_summary_sheets(wb, summary_amounts, runtime)

    tech_ws = wb["科技产业"]
    assert [tech_ws.title, wb.sheetnames[1], wb.sheetnames[2]] == ["科技产业", "数字经济产业", "养老产业"]
    assert tech_ws["B4"].value == "总余额"
    assert tech_ws["B5"].value == "线索合计"
    assert tech_ws["B6"].value == "按投向有误"
    assert tech_ws["B7"].value == "按投向疑似无误"
    assert tech_ws["B8"].value == "按主营有误"
    assert tech_ws["B9"].value == "按主营疑似无误"
    assert tech_ws["C4"].value == 10000
    assert tech_ws["D4"].value == 10000
    assert tech_ws["C5"].value == 6000
    assert tech_ws["D5"].value == 6000
    assert tech_ws["C6"].value == 5000
    assert tech_ws["D6"].value == 5000
    assert tech_ws["C7"].value == 1000
    assert tech_ws["D7"].value == 1000
    assert tech_ws["C8"].value == 0
    assert tech_ws["D8"].value == 0
    assert tech_ws["C9"].value == 0
    assert tech_ws["D9"].value == 0
    assert tech_ws["B10"].value == "关键字"
    assert tech_ws["A11"].value == "占比"
    assert tech_ws["B17"].value == "总笔数"
    assert tech_ws["A24"].value == "占比"
    assert tech_ws.max_row == 29

    digital_ws = wb["数字经济产业"]
    assert digital_ws.max_column == 8
    assert [digital_ws.cell(2, c).value for c in range(3, 9)] == ["合计", "DE01", "DE02", "DE03", "DE04", "DE05"]


def test_classify_template_stat_keys_dedupes_customer_when_actual_hits():
    runtime = {
        "enabled_result_types": ["错报", "疑似无误"],
        "template_row_labels": {
            "线索合计": ["错报", "疑似无误"],
            "按投向有误": ["错报"],
            "按投向疑似无误": ["疑似无误"],
            "按主营有误": ["错报"],
            "按主营疑似无误": ["疑似无误"],
        },
    }

    assert strategy1.classify_template_stat_keys("错报", "疑似无误", runtime) == {"line", "actual_error"}
    assert strategy1.classify_template_stat_keys("", "疑似无误", runtime) == {"line", "suspect_ok", "customer_all"}


def test_should_count_keyword_summary_only_when_actual_and_customer_do_not_hit():
    assert strategy1.should_count_keyword_summary({"是否疑似线索": "是"}, set()) is True
    assert strategy1.should_count_keyword_summary({"是否疑似线索": "是"}, {"line"}) is False
    assert strategy1.should_count_keyword_summary({"是否疑似线索": "是"}, {"customer_all", "suspect_ok"}) is False
    assert strategy1.should_count_keyword_summary({"是否疑似线索": ""}, set()) is False


def test_keyword_template_stat_keys_can_add_line_summary():
    runtime = {"template_row_labels": {"线索合计": ["错报", "关键字"], "关键字": ["关键字"]}}

    assert strategy1.keyword_template_stat_keys(runtime) == {"keyword", "line"}
    assert strategy1.keyword_template_stat_keys({"template_row_labels": {"线索合计": ["错报"]}}) == set()


def test_write_template_summary_group_total_does_not_sum_duplicate_codes():
    wb = Workbook()
    wb.remove(wb.active)
    summary_amounts = {
        ("科技产业", strategy1._summary_group_key("01 高技术制造业")): {
            "amount": {"total_amount": 10000},
            "count": {"total_count": 1},
        },
        ("科技产业", "HTP01"): {
            "amount": {"total_amount": 10000},
            "count": {"total_count": 1},
        },
        ("科技产业", "HTP02"): {
            "amount": {"total_amount": 10000},
            "count": {"total_count": 1},
        },
    }
    runtime = {
        "summary_unit": "万元",
        "template_row_labels": {key: [] for key in ["线索合计", "按投向有误", "按投向疑似无误", "按主营有误", "按主营疑似无误", "关键字"]},
    }

    strategy1.write_template_summary_sheets(wb, summary_amounts, runtime)

    tech_ws = wb["科技产业"]
    assert tech_ws["C4"].value == 10000
    assert tech_ws["D4"].value == 10000
    assert tech_ws["E4"].value == 10000


def test_build_result_headers_respects_runtime_switches():
    runtime = {
        "enabled_bases": ["actual", "customer"],
        "enabled_result_types": ["AA", "错报"],
    }

    headers = strategy1.build_result_headers(["HTP"], runtime)

    assert headers == [
        "投向-报送-高技术制造业",
        "投向-匹配-高技术制造业",
        "投向-AA-高技术制造业",
        "投向-错报-高技术制造业",
        "投向-是否线索-高技术制造业",
        "投向-是否疑似线索-高技术制造业",
        "投向-备注-高技术制造业",
        "客户-报送-高技术制造业",
        "客户-匹配-高技术制造业",
        "客户-AA-高技术制造业",
        "客户-错报-高技术制造业",
        "客户-是否线索-高技术制造业",
        "客户-是否疑似线索-高技术制造业",
        "客户-备注-高技术制造业",
        "行业小类描述",
    ]


def test_build_result_headers_can_enable_keyword_columns():
    runtime = {
        "enabled_bases": ["actual"],
        "enabled_result_types": ["错报", "关键字"],
    }

    headers = strategy1.build_result_headers(["HTP"], runtime)

    assert headers[-1] == "行业小类描述"


def test_write_column_validation_report_generates_excel(tmp_path, monkeypatch):
    monkeypatch.setattr(strategy1, "LOCAL_OUTPUT_DIR", str(tmp_path))
    source_wb = tmp_path / "source.xlsx"
    ref_wb = tmp_path / "ref.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "数字经济产业贷款明细"
    ws["F4"] = "贷款客户行业分类（5位编码-名称）"
    ws["O4"] = "贷款实际投向行业（5位编码-名称）"
    ws["P4"] = "数字经济产业贷款类型（大类代码-名称）"
    ws["F5"] = "C1442-乳粉制造"
    ws["O5"] = "C1442-乳粉制造"
    ws["P5"] = "DE05"
    wb.save(source_wb)
    wb.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"
    ws["B1"] = "产业大类编码"
    ws["M1"] = "对应国名经济行业编码小类"
    ws["S1"] = "是否带星号"
    ws["U1"] = "原始映射类别"
    ws["B2"] = "DE05"
    ws["M2"] = "C1442"
    ws["S2"] = "是"
    ws["U2"] = "门类(C)"
    wb.save(ref_wb)
    wb.close()

    config_rows = [{
        "输出工作表名称": "数字经济产业贷款明细",
        "来源台账文件": str(source_wb),
        "台账类型": "digital",
        "表头行号": 4,
        "数据起始行号": 5,
        "贷款客户行业列序号": 6,
        "贷款投向行业列序号": 15,
        "贷款余额列序号": 8,
        "贷款余额原始单位": "万元",
        "机构报送产业分类列序号": [16],
        "报送列-类别映射": [(16, "DE")],
        "参照表工作表序号": "汇总",
        "参照表产业分类代码列序号": 2,
        "参照表行业4位码列序号": 13,
        "参照表星标列序号": 19,
        "参照表原始映射列序号": 21,
    }]
    source_file_map = {source_wb.name: str(source_wb)}
    detail_param_lookup = {
        "数字经济产业": {
            "列位置_贷款客户行业分类": {"column_letter": "F"},
            "列位置_贷款实际投向行业分类": {"column_letter": "O"},
            "列位置_产业大类编码": {"column_letter": "P"},
        }
    }
    monkeypatch.setattr(
        strategy1,
        "_load_field_config_rows",
        lambda logical_name: [
            {"适用明细": "数字经济产业", "来源Sheet": "汇总", "字段名称": "列位置_产业大类编码", "列字母": "B", "列序号": 2},
            {"适用明细": "数字经济产业", "来源Sheet": "汇总", "字段名称": "列位置_对应国名经济行业编码小类", "列字母": "M", "列序号": 13},
            {"适用明细": "数字经济产业", "来源Sheet": "汇总", "字段名称": "列位置_是否带星号", "列字母": "S", "列序号": 19},
            {"适用明细": "数字经济产业", "来源Sheet": "汇总", "字段名称": "列位置_原始映射类别", "列字母": "U", "列序号": 21},
        ] if logical_name == "参照表字段配置" else [],
    )

    report_path = strategy1.write_column_validation_report(config_rows, source_file_map, str(ref_wb), detail_param_lookup)

    assert Path(report_path).exists()
    report_wb = openpyxl.load_workbook(report_path, data_only=True)
    assert "源文件字段校验" in report_wb.sheetnames
    assert "参照表字段校验" in report_wb.sheetnames
    source_ws = report_wb["源文件字段校验"]
    ref_ws = report_wb["参照表字段校验"]
    assert source_ws.max_row >= 2
    assert ref_ws.max_row >= 2
    assert source_ws["A1"].value == "明细名称"
    assert ref_ws["A1"].value == "适用对象"
    report_wb.close()


def test_classify_file_type_prefers_header_keywords_for_tech_even_with_28_columns():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for c in range(1, 29):
        ws.cell(1, c).value = ""
    ws["A1"] = "附表2 银行科技产业贷款明细"
    ws["Q3"] = "是否高技术制造业贷款"
    ws["T3"] = "高技术服务业贷款类型大类编码-对应名称"

    file_type, reported_cols, industry_col = strategy1.classify_file_type(ws)

    assert file_type == "tech"
    assert reported_cols == [17, 20, 23, 26]
    assert industry_col == 15


def test_build_config_rows_from_files_skips_unrecognized_workbook(tmp_path, capsys):
    path = tmp_path / "unknown.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "未知台账"
    for c in range(1, 6):
        ws.cell(1, c).value = f"列{c}"
    wb.save(path)
    wb.close()

    rows = strategy1.build_config_rows_from_files([str(path)])

    assert rows == []
    captured = capsys.readouterr()
    assert "跳过（无法识别台账类型" in captured.out


def test_evaluate_keyword_rules_matches_contains_exclude_groups_and_industry_context():
    rules = [
        {
            "序号": 1,
            "规则名称": "科技研发用途",
            "明细名称": "科技产业",
            "列字段": "列位置_贷款用途",
            "字段所在列位置": "N",
            "匹配方式": "contains",
            "包含关键词": "研发,技术",
            "排除关键词": "养老",
            "同时命中关键词组": "平台,系统|研发,设计",
            "至少命中组数": 2,
            "匹配投向行业": "是",
            "匹配主营行业": "",
            "行业比对层级": "小类,说明",
            "命中说明": "用途包含科技研发语义",
            "是否疑似线索": "是",
            "是否禁用": "",
        }
    ]
    field_catalog = {
        "列位置_贷款用途": {"value": "用于软件开发研发平台设计和系统建设", "label": "贷款用途", "column_letter": "N"},
        "贷款实际投向行业": {"value": "I6510-软件开发", "label": "贷款实际投向行业", "column_letter": "O"},
        "贷款客户行业分类": {"value": "I6510-软件开发", "label": "贷款客户行业分类", "column_letter": "F"},
    }
    industry_lookup = {
        "I6510": {
            "大类名称": "软件和信息技术服务业",
            "中类名称": "软件开发",
            "小类名称": "软件开发",
            "小类说明": "指基础软件、应用软件开发等活动",
        }
    }

    result = strategy1.evaluate_keyword_rules_for_row("科技产业", field_catalog, rules, industry_lookup)

    assert result["是否命中"] == "是"
    assert result["是否疑似线索"] == "是"
    assert result["规则名称"] == "科技研发用途"
    assert "用途包含科技研发语义" in result["命中说明"]
    assert "投向行业" in result["命中说明"]


def test_evaluate_keyword_rules_skips_rule_when_excluded_keyword_present():
    rules = [
        {
            "序号": 1,
            "规则名称": "养老服务用途",
            "明细名称": "养老产业",
            "列字段": "列位置_贷款用途",
            "字段所在列位置": "N",
            "匹配方式": "contains",
            "包含关键词": "养老,护理",
            "排除关键词": "医院",
            "同时命中关键词组": "",
            "至少命中组数": 0,
            "匹配投向行业": "",
            "匹配主营行业": "",
            "行业比对层级": "",
            "命中说明": "养老服务用途",
            "是否疑似线索": "是",
            "是否禁用": "",
        }
    ]
    field_catalog = {
        "列位置_贷款用途": {"value": "养老医院护理中心建设", "label": "贷款用途", "column_letter": "N"},
        "列位:N": {"value": "养老医院护理中心建设", "label": "贷款用途", "column_letter": "N"},
    }

    result = strategy1.evaluate_keyword_rules_for_row("养老产业", field_catalog, rules, {})

    assert result["是否命中"] == ""
    assert result["规则名称"] == ""
    assert result["是否疑似线索"] == ""


def test_load_runtime_settings_filters_enabled_result_types_by_clue_main_labels(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    _make_runtime_workbook(cfg)
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))
    monkeypatch.setattr(
        strategy1,
        "load_clue_rules",
        lambda: [
            {"主标签": "正确"},
            {"主标签": "AA"},
            {"主标签": "错报"},
        ],
    )

    settings = strategy1.load_runtime_settings()

    assert settings["enabled_result_types"] == ["错报"]


def test_resolve_source_workbook_path_prefers_mapping_file_path():
    mapped = strategy1.resolve_source_workbook_path(
        r"伪数据\附科技.xlsx",
        {"附科技.xlsx": r"C:\tmp\标准化前\附科技.xlsx"},
    )

    assert mapped == r"C:\tmp\标准化前\附科技.xlsx"


def test_standardize_source_workbooks_uses_mapping_paths_and_output_dir(tmp_path, monkeypatch):
    source_file = tmp_path / "科技.xlsx"
    source_file.write_text("placeholder", encoding="utf-8")
    config_rows = [{"来源台账文件": "科技.xlsx", "输出工作表名称": "科技产业贷款明细"}]
    calls = []

    class _FakeStandardizeModule:
        @staticmethod
        def standardize_matching_workbook(input_path, config_path, output_dir):
            calls.append((input_path, config_path, output_dir))
            return [{"sheet_name": "科技产业贷款明细", "output_path": output_dir / "标准化_科技.xlsx"}]

    monkeypatch.setattr(strategy1, "_load_standardize_module", lambda: _FakeStandardizeModule())
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(tmp_path / "config.xlsx"))
    monkeypatch.setattr(strategy1, "STANDARDIZE_OUTPUT_DIR", str(tmp_path / "output"))

    standardized = strategy1.standardize_source_workbooks(config_rows, {"科技.xlsx": str(source_file)})

    assert len(calls) == 1
    assert calls[0][0] == source_file
    assert calls[0][1] == Path(strategy1.CONFIG_FILE)
    assert calls[0][2] == Path(strategy1.STANDARDIZE_OUTPUT_DIR)
    assert standardized[strategy1._normalized_path_key(source_file)][0]["sheet_name"] == "科技产业贷款明细"


def test_resolve_standardized_workbook_path_prefers_matching_sheet(tmp_path):
    original = tmp_path / "科技.xlsx"
    standardized = tmp_path / "标准化_科技.xlsx"

    resolved = strategy1.resolve_standardized_workbook_path(
        str(original),
        "科技产业贷款明细",
        {
            strategy1._normalized_path_key(original): [
                {"sheet_name": "数字经济产业贷款明细", "output_path": tmp_path / "数字.xlsx"},
                {"sheet_name": "科技产业贷款明细", "output_path": standardized},
            ]
        },
    )

    assert resolved == str(standardized)


def test_build_strategy_output_path_points_to_excel_data_output(tmp_path, monkeypatch):
    monkeypatch.setattr(strategy1, "STANDARDIZE_OUTPUT_DIR", str(tmp_path / "output"))

    output_path = strategy1.build_strategy_output_path()

    assert Path(output_path).parent == tmp_path / "output"
    assert Path(output_path).name.startswith("策略一核查结果_")


def test_reconcile_config_rows_with_mapping_falls_back_to_inferred_rows(tmp_path):
    source_file = tmp_path / "科技.xlsx"
    source_file.write_text("placeholder", encoding="utf-8")
    invalid_rows = [{"来源台账文件": "科技产业", "输出工作表名称": "1"}]
    inferred_rows = [{"来源台账文件": str(source_file), "输出工作表名称": "科技产业贷款明细"}]

    original_builder = strategy1.build_config_rows_from_files
    strategy1.build_config_rows_from_files = lambda _: inferred_rows
    try:
        rows = strategy1.reconcile_config_rows_with_mapping(invalid_rows, {"科技.xlsx": str(source_file)})
    finally:
        strategy1.build_config_rows_from_files = original_builder

    assert rows == inferred_rows


def test_standardize_only_from_mapping_returns_output_paths(monkeypatch):
    monkeypatch.setattr(strategy1, "init_config_file_if_missing", lambda: None)
    monkeypatch.setattr(strategy1, "ensure_config_schema", lambda: None)
    monkeypatch.setattr(strategy1, "get_source_file_map_from_mapping", lambda: {"a.xlsx": r"C:\tmp\a.xlsx"})
    monkeypatch.setattr(strategy1, "load_config_rows", lambda: [{"来源台账文件": "a.xlsx"}])
    monkeypatch.setattr(strategy1, "reconcile_config_rows_with_mapping", lambda rows, source_map: rows)
    monkeypatch.setattr(
        strategy1,
        "standardize_source_workbooks",
        lambda rows, source_map: {
            strategy1._normalized_path_key(r"C:\tmp\a.xlsx"): [
                {"output_path": Path(r"C:\tmp\out1.xlsx")},
                {"output_path": Path(r"C:\tmp\out2.xlsx")},
            ]
        },
    )

    outputs = strategy1.standardize_only_from_mapping()

    assert outputs == [r"C:\tmp\out1.xlsx", r"C:\tmp\out2.xlsx"]


def test_initialize_config_workbook_rebuilds_all_runtime_and_standardize_sheets(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))
    monkeypatch.setattr(
        strategy1,
        "_read_existing_mapping_rows",
        lambda path: [
            ["参照表路径", r"C:\ref.xlsx"],
            ["映射表路径", r"C:\map.xlsx"],
            ["源文件", "源文件路径"],
            ["科技.xlsx", r"C:\ledger\科技.xlsx"],
        ],
    )
    monkeypatch.setattr(
        strategy1,
        "_build_initial_config_row_dicts",
        lambda path: [
            {
                "输出工作表名称": "科技产业贷款明细",
                "来源台账文件": r"..\..\excel-data\fakedata\科技.xlsx",
                "台账类型": "tech",
                "表头行号": 3,
                "数据起始行号": 4,
                "贷款客户行业列序号": 6,
                "贷款投向行业列序号": 15,
                "贷款余额列序号": 8,
                "机构报送产业分类列序号": [17, 20, 23, 26],
                "报送列-类别映射": [(17, "HTP"), (20, "HTS"), (23, "SE"), (26, "PA")],
                "参照表工作表序号": 4,
                "参照表产业分类代码列序号": 2,
                "参照表行业4位码列序号": 13,
                "参照表星标列序号": 19,
                "参照表原始映射列序号": 21,
            }
        ],
    )

    rows = strategy1.initialize_config_workbook(str(cfg))

    assert len(rows) == 1
    wb = openpyxl.load_workbook(cfg, data_only=True)
    assert wb.sheetnames == [
        "文件路径配置",
        "log",
        "线索规则配置_产业行业映射",
        "线索规则配置-关键字映射",
        "输出报告配置",
        "源文件明细定义",
        "源文件字段配置",
        "标准化配置-字段映射表",
    ]
    assert wb["输出报告配置"]["B2"].value == "实际投向行业,客户主营行业"
    assert wb["输出报告配置"].max_row == 10
    assert wb["输出报告配置"]["A5"].value == "模板行-线索合计统计项"
    assert wb["源文件明细定义"].max_column >= 9
    assert wb["源文件明细定义"]["I2"].value == "科技产业贷款明细"
    assert wb["线索规则配置-关键字映射"]["B1"].value == "规则名称"
    assert wb["线索规则配置-关键字映射"]["E1"].value == "字段所在列位置"
    assert wb["源文件明细定义"]["B2"].value == "科技产业"
    wb.close()


def test_ensure_config_schema_adds_missing_standardize_sheets(tmp_path, monkeypatch):
    cfg = tmp_path / "config.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "config"
    ws.append(strategy1.CONFIG_HEADERS)
    ws.append(
        [
            "科技产业贷款明细",
            "科技.xlsx",
            "科技",
            3,
            4,
            6,
            15,
            8,
            "万元",
            "17,20,23,26",
            "17:HTP,20:HTS,23:SE,26:PA",
            4,
            2,
            13,
            19,
            21,
        ]
    )
    wb.create_sheet("mapping").append(["参照表路径", ""])
    wb.create_sheet("log").append(strategy1.LOG_HEADERS)
    wb.create_sheet("clue").append(strategy1.CLUE_HEADERS)
    runtime_ws = wb.create_sheet("runtime")
    runtime_ws.append(strategy1.RUNTIME_HEADERS)
    runtime_ws.append(["启用核查口径", "实际投向行业,客户主营行业", ""])
    runtime_ws.append(["启用结果类型", "错报", ""])
    runtime_ws.append(["汇总报告单位", "万元", ""])
    wb.save(cfg)
    wb.close()

    monkeypatch.setattr(strategy1, "CONFIG_FILE", str(cfg))

    strategy1.ensure_config_schema()

    result_wb = openpyxl.load_workbook(cfg, data_only=True)
    assert "源文件明细定义" in result_wb.sheetnames
    assert "源文件字段配置" in result_wb.sheetnames
    assert "标准化配置-字段映射表" in result_wb.sheetnames
    assert "线索规则配置-关键字映射" in result_wb.sheetnames
    runtime_map = {result_wb["输出报告配置"].cell(r, 1).value: result_wb["输出报告配置"].cell(r, 2).value for r in range(2, result_wb["输出报告配置"].max_row + 1)}
    assert runtime_map["模板行-按投向有误统计项"] == "错报,多报"
    assert runtime_map["模板行-按主营疑似无误统计项"] in {"", None}
    assert "明细贷款余额原始单位" not in runtime_map
    assert "线索规则配置-config" not in result_wb.sheetnames
    assert result_wb["源文件明细定义"].max_column >= 9
    result_wb.close()


def test_detect_text_contradictions_flags_evidence_and_internal_conflict():
    rule_rows = [
        {
            "规则编号": "TXT001",
            "适用明细": "科技产业",
            "规则类型": "证据不足",
            "适用字段": "佐证摘要",
            "匹配模式": "empty_or_short",
            "关键词": "",
            "反向关键词": "",
            "命中说明": "佐证摘要为空或过短",
            "风险等级": "medium",
        },
        {
            "规则编号": "TXT002",
            "适用明细": "科技产业",
            "规则类型": "文本内部冲突",
            "适用字段": "贷款类型|贷款用途",
            "匹配模式": "keyword_pair",
            "关键词": "流动资金贷款|经营贷款",
            "反向关键词": "建设|购置固定资产|厂房",
            "命中说明": "贷款类型与用途不一致",
            "风险等级": "high",
        },
    ]
    text_fields = {
        "贷款类型": "流动资金贷款",
        "贷款用途": "用于建设养老中心并购置固定资产",
        "佐证摘要": "依据",
    }

    hits = strategy1.detect_text_contradictions("科技产业", text_fields, rule_rows)

    assert [item["规则编号"] for item in hits] == ["TXT001", "TXT002"]


def test_score_text_semantics_prefers_matching_detail():
    stopwords = {"用于生产经营", "依据参照表映射"}
    semantic_rules = [
        {"适用明细": "科技产业", "适用字段": "贷款用途", "关键词": "研发|技术|信息服务", "反向关键词": "养老|护理", "字段权重": 2.0},
        {"适用明细": "科技产业", "适用字段": "佐证摘要", "关键词": "高技术|科技成果", "反向关键词": "", "字段权重": 3.0},
        {"适用明细": "养老产业", "适用字段": "贷款用途", "关键词": "养老|照护|护理", "反向关键词": "研发|芯片", "字段权重": 2.0},
    ]
    text_fields = {
        "贷款用途": "用于研发平台技术升级和信息服务系统建设",
        "佐证摘要": "属于高技术服务业和科技成果转化",
    }

    scores = strategy1.score_text_semantics("科技产业", text_fields, semantic_rules, stopwords)

    assert scores["科技产业"]["support_score"] > scores["养老产业"]["support_score"]
    assert scores["科技产业"]["positive_hits"]


def test_detect_text_contradictions_supports_detail_param_and_column_tokens():
    field_catalog = {
        "列位置_是否高技术服务业贷款": {"value": "是", "label": "列位置_是否高技术服务业贷款(S)"},
        "列位置_高技术服务业贷款类型大类编码": {"value": "HTS01", "label": "列位置_高技术服务业贷款类型大类编码(T)"},
        "列位:U[纳入高技术服务业贷款依据]": {"value": "依据参照表映射", "label": "纳入高技术服务业贷款依据", "column_letter": "U"},
        "列位:U": {"value": "依据参照表映射", "label": "纳入高技术服务业贷款依据", "column_letter": "U"},
    }
    rules = [
        {
            "规则编号": "TXT900",
            "适用明细": "科技产业",
            "规则类型": "布尔标记与文本冲突",
            "适用字段": "列位置_是否高技术服务业贷款|列位置_高技术服务业贷款类型大类编码|列位:U[纳入高技术服务业贷款依据]",
            "匹配模式": "flag_requires_keywords",
            "关键词": "研发|信息服务",
            "反向关键词": "",
            "命中说明": "高技术服务业标记缺少文本支撑",
            "风险等级": "medium",
        }
    ]
    detail_param_lookup = {
        "科技产业": {
            "列位置_是否高技术服务业贷款": {"column_letter": "S", "label": "列位置_是否高技术服务业贷款"},
            "列位置_高技术服务业贷款类型大类编码": {"column_letter": "T", "label": "列位置_高技术服务业贷款类型大类编码"},
        }
    }

    hits = strategy1.detect_text_contradictions("科技产业", field_catalog, rules, detail_param_lookup)

    assert hits[0]["规则编号"] == "TXT900"
    assert "纳入高技术服务业贷款依据" in hits[0]["命中字段"]


def test_build_text_verdict_returns_obvious_conflict_for_high_risk_hit():
    verdict_rules = [
        {"最小分值": 0, "最大分值": 29.99, "冲突等级": "high", "结论": "明显矛盾", "复核建议": "优先核对用途与佐证摘要"},
        {"最小分值": 30, "最大分值": 69.99, "冲突等级": "medium", "结论": "建议复核", "复核建议": "补充结构化佐证"},
        {"最小分值": 70, "最大分值": 100, "冲突等级": "none", "结论": "建议纳统", "复核建议": "可进入人工抽查"},
    ]

    verdict = strategy1.build_text_verdict(
        target_detail="科技产业",
        semantic_scores={"科技产业": {"support_score": 82.0, "positive_hits": ["研发"], "negative_hits": []}},
        contradiction_hits=[{"风险等级": "high", "命中说明": "贷款类型与用途不一致"}],
        verdict_rules=verdict_rules,
    )

    assert verdict["结论"] == "明显矛盾"
    assert "贷款类型与用途不一致" in verdict["结论原因"]


def test_write_text_analysis_results_outputs_expected_report_sheets(tmp_path, monkeypatch):
    source_path = tmp_path / "科技.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "科技产业贷款明细"
    ws.cell(4, 1).value = 1
    ws.cell(4, 2).value = "HT001"
    ws.cell(4, 3).value = "JD001"
    ws.cell(4, 5).value = "晨光科技有限公司"
    ws.cell(4, 6).value = "C3951-电视机制造"
    ws.cell(4, 13).value = "流动资金贷款"
    ws.cell(4, 14).value = "用于建设养老中心"
    ws.cell(4, 16).value = "HTP01"
    ws.cell(4, 17).value = 1
    ws.cell(4, 18).value = "依据"
    wb.save(source_path)
    wb.close()

    cfg_row = {
        "输出工作表名称": "科技产业贷款明细",
        "来源台账文件": "科技.xlsx",
        "台账类型": "tech",
        "表头行号": 3,
        "数据起始行号": 4,
        "贷款客户行业列序号": 6,
        "贷款投向行业列序号": 15,
        "贷款余额列序号": 8,
        "机构报送产业分类列序号": [17],
        "报送列-类别映射": [(17, "HTP")],
        "参照表工作表序号": 4,
        "参照表产业分类代码列序号": 2,
        "参照表行业4位码列序号": 13,
        "参照表星标列序号": 19,
        "参照表原始映射列序号": 21,
    }

    monkeypatch.setattr(strategy1, "init_config_file_if_missing", lambda: None)
    monkeypatch.setattr(strategy1, "ensure_config_schema", lambda: None)
    monkeypatch.setattr(strategy1, "get_source_file_map_from_mapping", lambda: {"科技.xlsx": str(source_path)})
    monkeypatch.setattr(strategy1, "load_config_rows", lambda: [cfg_row])
    monkeypatch.setattr(strategy1, "reconcile_config_rows_with_mapping", lambda rows, source_map: rows)
    monkeypatch.setattr(strategy1, "standardize_source_workbooks", lambda rows, source_map: {})
    monkeypatch.setattr(strategy1, "find_mapping_file", lambda: str(tmp_path / "mapping.xlsx"))
    monkeypatch.setattr(strategy1, "build_mapping_by_config", lambda mapping_file, cfg: {})
    monkeypatch.setattr(strategy1, "load_clue_rules", lambda: [])
    monkeypatch.setattr(strategy1, "load_runtime_settings", lambda: {"enabled_bases": ["actual"], "enabled_result_types": [], "summary_unit": "万元"})
    monkeypatch.setattr(strategy1, "load_industry_desc_map", lambda: {})
    monkeypatch.setattr(
        strategy1,
        "load_text_analysis_settings",
        lambda: {
            "rules": [
                {
                    "规则编号": "TXT001",
                    "适用明细": "科技产业",
                    "规则类型": "证据不足",
                    "适用字段": "佐证摘要",
                    "匹配模式": "empty_or_short",
                    "关键词": "8",
                    "反向关键词": "",
                    "命中说明": "佐证摘要为空或过短",
                    "风险等级": "medium",
                },
                {
                    "规则编号": "TXT002",
                    "适用明细": "科技产业",
                    "规则类型": "文本内部冲突",
                    "适用字段": "贷款类型|贷款用途",
                    "匹配模式": "keyword_pair",
                    "关键词": "流动资金贷款",
                    "反向关键词": "建设|养老中心",
                    "命中说明": "贷款类型与用途不一致",
                    "风险等级": "high",
                },
                {
                    "规则编号": "TXT101",
                    "适用明细": "科技产业",
                    "规则类型": "语义支持",
                    "适用字段": "佐证摘要",
                    "匹配模式": "keyword_bag",
                    "关键词": "研发|高技术",
                    "反向关键词": "养老",
                    "命中说明": "",
                    "风险等级": "",
                    "字段权重": 3.0,
                },
            ],
            "verdicts": [
                {"最小分值": 0, "最大分值": 100, "冲突等级": "high", "结论": "明显矛盾", "复核建议": "优先核对文本"},
                {"最小分值": 0, "最大分值": 100, "冲突等级": "medium", "结论": "建议复核", "复核建议": "补充佐证"},
                {"最小分值": 0, "最大分值": 100, "冲突等级": "none", "结论": "建议纳统", "复核建议": "可抽查"},
            ],
            "stopwords": [],
        },
    )
    monkeypatch.setattr(strategy1, "LOCAL_OUTPUT_DIR", str(tmp_path / "output"))

    output_path = strategy1.write_text_analysis_results()

    result_wb = openpyxl.load_workbook(output_path, data_only=True)
    assert result_wb.sheetnames == ["文本辅助核查", "规则命中明细", "词典配置快照"]
    assert result_wb["文本辅助核查"]["P2"].value == "明显矛盾"
    assert result_wb["规则命中明细"].max_row == 3
    result_wb.close()
