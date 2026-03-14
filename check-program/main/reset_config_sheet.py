import os

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# config/clue 默认数据以 build_strategy1 中 DEFAULT_CONFIG_ROWS、CLUE_DEFAULT_ROWS 为准，重置时无现有表则写回该内置数据
from build_strategy1 import (
    ROOT,
    CONFIG_HEADERS,
    CLUE_HEADERS,
    CLUE_DEFAULT_ROWS,
    TYPE_KEY_TO_LABEL,
    DEFAULT_CONFIG_ROWS,
    get_source_file_map_from_mapping,
    build_config_rows_from_files,
    find_ledger_files,
)

CONFIG_FILE = os.path.join(ROOT, "config.xlsx")

# 与 build_strategy1 中 log 表头一致，用于无现有 config 时或 log 无表头时
LOG_HEADERS = [
    "运行时间",
    "参照表路径",
    "参照表sheet数量",
    "参照表sheet列表",
    "结果文件",
    "明细sheet名称",
    "来源台账文件",
    "台账类型",
    "使用参照表sheet序号",
    "使用参照表sheet名称",
    "错报数量",
    "漏报数量",
    "疑似漏报数量",
]


def main():
    cfg_path = CONFIG_FILE
    if os.path.exists(cfg_path):
        wb = openpyxl.load_workbook(cfg_path)
        config_rows = []
        if "config" in wb.sheetnames:
            ws_cfg_old = wb["config"]
            for row in ws_cfg_old.iter_rows(min_row=1, max_row=ws_cfg_old.max_row, values_only=True):
                config_rows.append([c if c is not None else "" for c in row])
        mapping_rows = []
        if "mapping" in wb.sheetnames:
            ws_m = wb["mapping"]
            for row in ws_m.iter_rows(min_row=1, max_row=ws_m.max_row, values_only=True):
                mapping_rows.append(list(row))
        log_rows = []
        if "log" in wb.sheetnames:
            ws_l = wb["log"]
            for row in ws_l.iter_rows(min_row=1, max_row=ws_l.max_row, values_only=True):
                log_rows.append(list(row))
        clue_rows = []
        if "clue" in wb.sheetnames:
            ws_clue_old = wb["clue"]
            for row in ws_clue_old.iter_rows(min_row=1, max_row=ws_clue_old.max_row, values_only=True):
                clue_rows.append(list(row))
        for s in list(wb.sheetnames):
            wb.remove(wb[s])
    else:
        # 根目录没有 config 则新建
        wb = Workbook()
        wb.remove(wb.active)  # 去掉默认空表，后面会建 config/mapping/log
        config_rows = []
        mapping_rows = []
        log_rows = []
        clue_rows = []
        print("未找到 config.xlsx，已新建并初始化。")

    # 1. 重建 config 表：有现有数据则整表写回，否则使用 build_strategy1 中内置 DEFAULT_CONFIG_ROWS（与 config 表实际数据一致）
    ws_cfg = wb.create_sheet("config")
    if config_rows and len(config_rows) >= 1:
        for row in config_rows:
            ws_cfg.append(row)
        for c in range(1, ws_cfg.max_column + 1):
            if ws_cfg.cell(1, c).value:
                ws_cfg.cell(1, c).font = Font(name="Arial", bold=True)
    else:
        source_map = get_source_file_map_from_mapping() if os.path.exists(cfg_path) else {}
        paths = list(source_map.values()) if source_map else []
        if not paths:
            paths = find_ledger_files()
        rows = build_config_rows_from_files(paths)
        if not rows:
            rows = DEFAULT_CONFIG_ROWS
            print("未从源文件推断到配置行，已使用内置默认配置（与 build_strategy1 中 DEFAULT_CONFIG_ROWS 一致，科技/数字/养老三表）。")
        ws_cfg.append(CONFIG_HEADERS)
        for c in range(1, len(CONFIG_HEADERS) + 1):
            ws_cfg.cell(1, c).font = Font(name="Arial", bold=True)
        for r in rows:
            type_label = TYPE_KEY_TO_LABEL.get(r["台账类型"], r["台账类型"])
            ws_cfg.append(
                [
                    r["输出工作表名称"],
                    r["来源台账文件"],
                    type_label,
                    r["表头行号"],
                    r["数据起始行号"],
                    r["贷款投向行业列序号"],
                    ",".join(str(x) for x in r["机构报送产业分类列序号"]),
                    ",".join(f"{c}:{t}" for c, t in r["报送列-类别映射"]),
                    r["参照表工作表序号"],
                    r["参照表产业分类代码列序号"],
                    r["参照表行业4位码列序号"],
                    r["参照表星标列序号"],
                    r.get("参照表原始映射列序号", 21),
                ]
            )
        desc_row = [
            "说明：结果文件中的 sheet 名称，例如“科技产业贷款明细核查”",
            "说明：来源台账文件路径，相对项目根目录或绝对路径",
            "说明：台账类型，只能填“科技/数字/养老”三种之一；脚本会据此推断默认的参照表 sheet 以及“报送列-类别映射”等参数；若填其他值（如 tech/digital/elder），该行配置会被跳过，整张台账不会执行策略一，也不会套用默认规则",
            "说明：表头行号，填数字；若填 3，则第 3 行作为表头",
            "说明：数据起始行号，填数字；若填 4，则从第 4 行开始读取数据",
            "说明：贷款投向行业列序号，对应源数据中“贷款实际投向行业（行业小类）”那一列，单元格内容需包含类似 A1234 的行业码，程序用正则从整格文本中提取；若列号填错或该列没有标准行业码，则无法在参照表中匹配到行业，整行不会产生多报/漏报结果",
            "说明：机构报送产业分类列序号，可填多个列号，用英文逗号分隔，例如 17,20,23；若写少了会漏判该列报送，写错列号会把无关列当作报送代码导致误判",
            "说明：报送列与产业类别代码映射，例如 17:HTP,20:HTS,23:SE,26:PA；若类别映射错位（如 17 映射到 HTS），则该列报送会被归入错误产业，导致该产业多报/漏报统计全部偏移",
            "说明：参照表工作表序号，可填数字序号（从1开始）或参照表 sheet 名称（需与参照表内名称完全一致）",
            "说明：参照表中“产业分类代码”所在列号，例如 HTP01/DE02 等",
            "说明：参照表中“行业4位码”所在列号，例如 C2345，将与贷款投向行业提取出的 4 位码匹配",
            "说明：参照表中“是否标*”所在列号，用于识别星标并判断疑似多报/疑似漏报",
            "说明：参照表中“原始映射内容”所在列号（默认 U 列=21），一般为小类/中类/大类/门类说明，用于结果中【code:原始映射】展示",
        ]
        if ws_cfg.max_row < 4:
            while ws_cfg.max_row < 4:
                ws_cfg.append([])
            ws_cfg.append(desc_row)
        else:
            ws_cfg.insert_rows(5)
            for idx, val in enumerate(desc_row, start=1):
                ws_cfg.cell(5, idx).value = val

    # 2. 重建 mapping
    ws_m = wb.create_sheet("mapping")
    if mapping_rows:
        for row in mapping_rows:
            ws_m.append(row)
    else:
        ws_m.append(["参照表路径", ""])
        ws_m.append(["映射表路径", ""])

    # 5. 重建 log：表头优先用现有 config 中 log 的表头，无则用默认 LOG_HEADERS；有历史则保留数据行
    log_headers = list(log_rows[0]) if (log_rows and len(log_rows) > 0) else LOG_HEADERS
    ws_l = wb.create_sheet("log")
    ws_l.append(log_headers)
    for c in range(1, len(log_headers) + 1):
        if ws_l.cell(1, c).value:
            ws_l.cell(1, c).font = Font(name="Arial", bold=True)
    if log_rows and len(log_rows) > 1:
        has_header = log_rows[0] and str(log_rows[0][0]).strip() == "运行时间"
        data_start = 1 if has_header else 0
        for row in log_rows[data_start:]:
            ws_l.append(row)

    # 4. 重建 clue 表：有现有数据则整表写回，否则使用 build_strategy1 中 CLUE_DEFAULT_ROWS（12 条规则，与 clue 表实际数据一致）
    ws_clue = wb.create_sheet("clue")
    if clue_rows and len(clue_rows) >= 1:
        for row in clue_rows:
            ws_clue.append(row)
        for c in range(1, max(len(CLUE_HEADERS), ws_clue.max_column) + 1):
            if ws_clue.cell(1, c).value:
                ws_clue.cell(1, c).font = Font(name="Arial", bold=True)
    else:
        ws_clue.append(CLUE_HEADERS)
        for c in range(1, len(CLUE_HEADERS) + 1):
            ws_clue.cell(1, c).font = Font(name="Arial", bold=True)
        for row in CLUE_DEFAULT_ROWS:
            ws_clue.append(row)

    wb.save(cfg_path)
    wb.close()
    if config_rows and len(config_rows) >= 1:
        print(f"config 工作表已重建，已保留 config/clue 整表及 log 表头；config 表共 {len(config_rows)} 行。")
    else:
        print(f"config 工作表已重建，共 {len(rows)} 行配置。")


if __name__ == "__main__":
    main()

