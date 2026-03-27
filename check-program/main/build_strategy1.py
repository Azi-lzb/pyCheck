import importlib.util
import os
import re
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

# 可选：用于文件选择对话框（留空输入时弹出，Windows/带图形环境可用）
try:
    import tkinter as _tk
    from tkinter import filedialog as _filedialog
except ImportError:
    _tk = None
    _filedialog = None


# 程序所在目录（check-program/main）；打包为 exe 时为 exe 所在目录
if getattr(sys, "frozen", False):
    ROOT = os.path.dirname(sys.executable)
else:
    ROOT = os.path.dirname(os.path.abspath(__file__))
# 项目根目录（pyCheck，含 excel-data、check-program），文件选择器默认从此处打开；打包为 exe 时与 ROOT 一致
if getattr(sys, "frozen", False):
    PROJECT_ROOT = ROOT
else:
    PROJECT_ROOT = os.path.dirname(os.path.dirname(ROOT))
OUTPUT_FILE = os.path.join(ROOT, "策略一核查结果.xlsx")
CONFIG_FILE = os.path.join(ROOT, "config.xlsx")
LOCAL_OUTPUT_DIR = os.path.join(ROOT, "output")
STANDARDIZE_SCRIPT = os.path.join(PROJECT_ROOT, "excel-data", "scripts", "standardize_fake_data.py")
STANDARDIZE_OUTPUT_DIR = LOCAL_OUTPUT_DIR
STANDARDIZE_REPORT_SHEETS = {"字段映射表触发情况", "明细定义触发情况", "明细参数触发情况"}
_STANDARDIZE_MODULE = None
_CLUE_RULES_CACHE = {"mtime_ns": None, "rules": None}
# 术语：参照表 = 五篇大文章与国民经济行业分类对应参照表（config mapping 表 A2）；映射表 = 国民经济行业分类映射表（config mapping 表 B2）
# 映射表（国民经济行业分类映射表）：用于按行业小类取描述；路径优先从 config mapping B2 读取，否则用此处默认（可改为 ../excel-data/data/国民经济行业分类映射表.xlsx）
INDUSTRY_DESC_FILE = os.path.join(ROOT, "国民经济行业分类映射表.xlsx")
# 映射表内列：小类代码列（1-based J=10）、小类描述列（1-based L=12）
INDUSTRY_MAP_CODE_COL = 10
INDUSTRY_MAP_DESC_COL = 12


def _load_standardize_module():
    global _STANDARDIZE_MODULE
    if _STANDARDIZE_MODULE is not None:
        return _STANDARDIZE_MODULE
    if not os.path.exists(STANDARDIZE_SCRIPT):
        raise FileNotFoundError(f"未找到标准化脚本：{STANDARDIZE_SCRIPT}")
    spec = importlib.util.spec_from_file_location("strategy1_standardize_fake_data", STANDARDIZE_SCRIPT)
    if spec is None or spec.loader is None:
        raise ImportError(f"无法加载标准化脚本：{STANDARDIZE_SCRIPT}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    _STANDARDIZE_MODULE = module
    return module


def _ask_open_file(title, filetypes=None, initialdir=None):
    """弹出系统文件选择对话框选择单个文件。无图形环境时返回 None。"""
    if _tk is None or _filedialog is None:
        return None
    if filetypes is None:
        filetypes = (
            ("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("所有文件", "*.*"),
        )
    root = _tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    path = _filedialog.askopenfilename(
        title=title,
        filetypes=filetypes,
        initialdir=initialdir or PROJECT_ROOT,
    )
    try:
        root.destroy()
    except Exception:
        pass
    return path if path else None


def _ask_open_files(title, filetypes=None, initialdir=None):
    """弹出系统文件选择对话框选择多个文件。无图形环境时返回 ()。"""
    if _tk is None or _filedialog is None:
        return ()
    if filetypes is None:
        filetypes = (
            ("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("所有文件", "*.*"),
        )
    root = _tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    paths = _filedialog.askopenfilenames(
        title=title,
        filetypes=filetypes,
        initialdir=initialdir or PROJECT_ROOT,
    )
    try:
        root.destroy()
    except Exception:
        pass
    return tuple(paths) if paths else ()


def resolve_sheet_name(sheetnames, logical_name):
    aliases = SHEET_NAME_ALIASES.get(logical_name, [logical_name])
    for alias in aliases:
        if alias in sheetnames:
            return alias
    return None


def get_sheet_by_logical_name(workbook, logical_name):
    sheet_name = resolve_sheet_name(workbook.sheetnames, logical_name)
    return workbook[sheet_name] if sheet_name else None

CONFIG_HEADERS = [
    "输出工作表名称",
    "来源台账文件",
    "台账类型",
    "表头行号",
    "数据起始行号",
    "贷款客户行业列序号",
    "贷款投向行业列序号",
    "贷款余额列序号",
    "贷款余额原始单位",
    "机构报送产业分类列序号",
    "报送列-类别映射",
    "参照表工作表序号",
    "参照表产业分类代码列序号",
    "参照表行业4位码列序号",
    "参照表星标列序号",
    "参照表原始映射列序号",
]

TYPE_LABEL_TO_KEY = {"科技": "tech", "数字": "digital", "养老": "elder"}
TYPE_KEY_TO_LABEL = {"tech": "科技", "digital": "数字", "elder": "养老"}
BASIS_RUNTIME_TO_KEY = {
    "actual": "actual",
    "customer": "customer",
    "实际投向行业": "actual",
    "客户主营行业": "customer",
}
SHEET_NAME_ALIASES = {
    "config": ["线索规则配置-config", "config"],
    "mapping": ["文件路径配置", "文件路径配置-mapping", "mapping"],
    "log": ["log"],
    "clue": ["线索规则配置_产业行业映射", "线索规则配置-clue", "clue"],
    "关键字映射": ["线索规则配置-关键字映射", "关键字映射"],
    "runtime": ["输出报告配置", "输出报告配置-runtime", "runtime"],
    "明细定义": ["源文件明细定义", "标准化配置-明细定义", "明细定义"],
    "明细参数": ["源文件字段配置", "标准化配置-明细参数", "明细参数"],
    "字段映射表": ["标准化配置-字段映射表", "字段映射表"],
    "参照表字段配置": ["参照表字段配置"],
    "国民经济映射字段配置": ["国民经济映射字段配置"],
}
PREFERRED_SHEET_NAMES = {logical_name: aliases[0] for logical_name, aliases in SHEET_NAME_ALIASES.items()}

# 重置 config 时若无源文件可推断，则使用以下默认配置（与 config 表实际数据一致，防止 config 丢失需重新配置）
DEFAULT_CONFIG_ROWS = [
    {
        "输出工作表名称": "科技产业贷款明细",
        "来源台账文件": r"伪数据\附科技",
        "台账类型": "科技",
        "表头行号": 3,
        "数据起始行号": 4,
        "贷款客户行业列序号": 6,
        "贷款投向行业列序号": 15,
        "贷款余额列序号": 8,
        "贷款余额原始单位": "万元",
        "机构报送产业分类列序号": [17, 20, 23, 26],
        "报送列-类别映射": [(17, "HTP"), (20, "HTS"), (23, "SE"), (26, "PA")],
        "参照表工作表序号": 4,
        "参照表产业分类代码列序号": 2,
        "参照表行业4位码列序号": 13,
        "参照表星标列序号": 19,
        "参照表原始映射列序号": 21,
    },
    {
        "输出工作表名称": "数字经济产业贷款明细",
        "来源台账文件": r"伪数据\附数字",
        "台账类型": "数字",
        "表头行号": 4,
        "数据起始行号": 5,
        "贷款客户行业列序号": 6,
        "贷款投向行业列序号": 15,
        "贷款余额列序号": 8,
        "贷款余额原始单位": "万元",
        "机构报送产业分类列序号": [16],
        "报送列-类别映射": [(16, "DE")],
        "参照表工作表序号": 2,
        "参照表产业分类代码列序号": 2,
        "参照表行业4位码列序号": 13,
        "参照表星标列序号": 19,
        "参照表原始映射列序号": 21,
    },
    {
        "输出工作表名称": "养老产业贷款明细",
        "来源台账文件": r"伪数据\附养老",
        "台账类型": "养老",
        "表头行号": 3,
        "数据起始行号": 4,
        "贷款客户行业列序号": 6,
        "贷款投向行业列序号": 15,
        "贷款余额列序号": 8,
        "贷款余额原始单位": "万元",
        "机构报送产业分类列序号": [16],
        "报送列-类别映射": [(16, "EC")],
        "参照表工作表序号": 3,
        "参照表产业分类代码列序号": 2,
        "参照表行业4位码列序号": 13,
        "参照表星标列序号": 19,
        "参照表原始映射列序号": 21,
    },
]

CATEGORY_NAME_MAP = {
    "HTP": "高技术制造业",
    "HTS": "高技术服务业",
    "SE": "战略性新兴产业",
    "PA": "知识产权密集型产业",
    "DE": "数字经济产业",
    "EC": "养老产业",
}

BASIS_NAME_MAP = {
    "actual": "投向",
    "customer": "客户",
}

KEYWORD_RESULT_LABEL = "关键字"
KEYWORD_RESULT_ALIASES = {KEYWORD_RESULT_LABEL, "关键字命中"}
RESULT_TYPE_ORDER = ["疑似正确", "错报", "漏报", "疑似漏报", "多报"]
SUMMARY_RESULT_LABELS = ["疑似正确", "错报", "多报", "疑似多报", "漏报", "疑似漏报"]

RUNTIME_HEADERS = ["配置项", "值", "说明"]
RUNTIME_DEFAULT_ROWS = [
    ["启用核查口径", "实际投向行业,客户主营行业", "可选 实际投向行业、客户主营行业；兼容旧值 actual、customer"],
    ["启用结果类型", "疑似正确,错报,漏报,疑似漏报,多报", "控制结果列与汇总页展示哪些结果类型；如需启用关键字列，请额外写入“关键字”"],
    ["汇总报告单位", "万元", "汇总sheet展示单位，可填 元/万元/亿元"],
    ["模板行-线索合计统计项", "错报,多报,疑似多报,疑似正确", "模板汇总sheet“线索合计”统计哪些结果标签"],
    ["模板行-按投向有误统计项", "错报,多报", "模板汇总sheet“按投向有误”统计哪些结果标签"],
    ["模板行-按投向疑似无误统计项", "疑似正确,疑似多报", "模板汇总sheet“按投向疑似无误”统计哪些结果标签"],
    ["模板行-按主营有误统计项", "错报,多报", "模板汇总sheet“按主营有误”统计哪些结果标签"],
    ["模板行-按主营疑似无误统计项", "", "模板汇总sheet“按主营疑似无误”统计哪些结果标签，可留空"],
    ["模板行-关键字", "关键字", "模板汇总sheet“关键字”统计哪些结果标签；当前建议填“关键字”"],
]

SUMMARY_TEMPLATE_RUNTIME_KEYS = {
    "线索合计": "模板行-线索合计统计项",
    "按投向有误": "模板行-按投向有误统计项",
    "按投向疑似无误": "模板行-按投向疑似无误统计项",
    "按主营有误": "模板行-按主营有误统计项",
    "按主营疑似无误": "模板行-按主营疑似无误统计项",
    "关键字": "模板行-关键字",
}
SUMMARY_TEMPLATE_SECTIONS = [
    ("金额", [("总余额", "total_amount"), ("线索合计", "line"), ("按投向有误", "actual_error"), ("按投向疑似无误", "suspect_ok"), ("按主营有误", "customer_error"), ("按主营疑似无误", "customer_all"), ("关键字", "keyword")]),
    ("占比", [("线索占比", "line"), ("按投向有误", "actual_error"), ("按投向疑似无误", "suspect_ok"), ("按主营有误", "customer_error"), ("按主营疑似无误", "customer_all"), ("关键字", "keyword")]),
    ("笔数", [("总笔数", "total_count"), ("线索合计", "line"), ("按投向有误", "actual_error"), ("按投向疑似无误", "suspect_ok"), ("按主营有误", "customer_error"), ("按主营疑似无误", "customer_all"), ("关键字", "keyword")]),
    ("占比", [("线索占比", "line"), ("按投向有误", "actual_error"), ("按投向疑似无误", "suspect_ok"), ("按主营有误", "customer_error"), ("按主营疑似无误", "customer_all"), ("关键字", "keyword")]),
]
SUMMARY_TEMPLATE_SPECS = {
    "科技产业": {
        "groups": [
            ("01 高技术制造业", [("HTP01", "医药制造业"), ("HTP02", "航空、航天器及设备制造业"), ("HTP03", "电子及通信设备制造业"), ("HTP04", "计算机及办公设备制造业"), ("HTP05", "医疗仪器设备及仪器仪表制造业"), ("HTP06", "信息化学品制造业")]),
            ("02 高技术服务业", [("HTS01", "信息服务"), ("HTS02", "电子商务服务"), ("HTS03", "检验检测服务"), ("HTS04", "专业技术服务业的高技术服务"), ("HTS05", "研发与设计服务"), ("HTS06", "科技成果转化服务"), ("HTS07", "知识产权及相关法律服务"), ("HTS08", "环境监测及治理服务")]),
            ("03 战略性新兴产业", [("SE01", "新一代信息技术产业"), ("SE02", "高端装备制造产业"), ("SE03", "新材料产业"), ("SE04", "生物产业"), ("SE05", "新能源汽车产业"), ("SE06", "新能源产业"), ("SE07", "节能环保产业"), ("SE08", "数字创意产业"), ("SE09", "相关服务业")]),
            ("04 知识产权密集型产业", [("PA01", "信息通信技术制造业"), ("PA02", "信息通信技术服务业"), ("PA03", "新装备制造业"), ("PA04", "新材料制造业"), ("PA05", "医药医疗产业"), ("PA06", "环保产业"), ("PA07", "研发、设计和技术服务业")]),
        ],
    },
    "数字经济产业": {
        "groups": [
            ("数字经济产业", [("DE01", "数字产品制造业"), ("DE02", "数字产品服务业"), ("DE03", "数字技术应用业"), ("DE04", "数字要素驱动业"), ("DE05", "数字化效率提升业")]),
        ],
    },
    "养老产业": {
        "groups": [
            ("养老产业", [("EC01", "养老照护服务"), ("EC02", "老年医疗卫生服务"), ("EC03", "老年健康促进与社会参与"), ("EC04", "老年社会保障"), ("EC05", "养老教育培训和人力资源服务"), ("EC06", "养老金融服务"), ("EC07", "养老科技和智慧养老服务"), ("EC08", "养老公共管理"), ("EC09", "其他养老服务"), ("EC10", "老年用品及相关产品制造"), ("EC11", "老年用品及相关产品销售和租赁"), ("EC12", "养老设施建设")]),
        ],
    },
}


def resolve_summary_group_name(sheet_name, category_code):
    spec = SUMMARY_TEMPLATE_SPECS.get(sheet_name, {})
    for group_name, codes in spec.get("groups", []):
        if any(code.startswith(category_code) for code, _ in codes):
            return group_name
    return ""


DETAIL_DEFINITION_HEADERS = ["序号", "明细名称", "工作簿名关键字", "工作表名关键字", "表头行号", "数据起始行号", "是否禁用", "备注"]
DETAIL_PARAM_HEADERS = ["序号", "明细名称", "明细字段", "字段所在列位置", "匹配方式", "是否禁用", "备注", "添加日期"]
DETAIL_DEFINITION_CONFIG_HEADERS = [
    "输出工作表名称",
    "来源台账文件",
    "台账类型",
    "贷款余额原始单位",
    "参照表工作表序号",
    "参照表产业分类代码列序号",
    "参照表行业4位码列序号",
    "参照表星标列序号",
    "参照表原始映射列序号",
]
DETAIL_PARAM_CONFIG_FIELD_NAMES = [
    "贷款客户行业列序号",
    "贷款投向行业列序号",
    "贷款余额列序号",
    "机构报送产业分类列序号",
    "报送列-类别映射",
]
DETAIL_DEFINITION_ALL_HEADERS = DETAIL_DEFINITION_HEADERS + DETAIL_DEFINITION_CONFIG_HEADERS
FIELD_MAPPING_HEADERS = ["序号", "明细名称", "列字段", "原始列值_exact", "原始列值_contains", "原始列值_regex", "映射标准值", "映射标准值_regex", "是否禁用", "备注"]
KEYWORD_MAPPING_HEADERS = ["序号", "规则名称", "明细名称", "列字段", "字段所在列位置", "匹配方式", "包含关键词", "排除关键词", "同时命中关键词组", "至少命中组数", "匹配投向行业", "匹配主营行业", "行业比对层级", "命中说明", "是否疑似线索", "是否禁用", "备注"]
TEXT_RULE_HEADERS = ["规则编号", "适用明细", "规则类型", "适用字段", "匹配模式", "关键词", "反向关键词", "命中说明", "风险等级", "字段权重"]
TEXT_VERDICT_HEADERS = ["序号", "最小分值", "最大分值", "冲突等级", "结论", "复核建议"]
TEXT_STOPWORD_HEADERS = ["序号", "词语", "是否禁用", "备注"]
TEXT_RULE_COLUMN_PATTERN = re.compile(r"^列位:([A-Z]+)(?:\[(.+?)\])?$")

# 线索规则表（config.xlsx 的 clue 工作表）：按条件匹配得到主标签、副标签、是否线索、是否疑似线索，通用六大产业
# 表头共 9 列（见 使用说明.md 第七节）
CLUE_HEADERS = [
    "编号",
    "匹配到的产业数量",
    "行业是否含星",
    "机构报送产业为空?",
    "命中情况",
    "主标签",
    "是否线索",
    "是否疑似线索",
    "备注/副标签",
]
# 默认 12 行规则（与线索默认规则表一致）；匹配顺序从上到下
# 表中「是否机构报送」= 是 → 机构有报送 → 本表「机构报送产业为空?」= 否；「是否机构报送」= 否 → 本表「机构报送产业为空?」= 是
CLUE_DEFAULT_ROWS = [
    [1, "0", "-", "是", "-", "正确", "否", "否", "行业未匹配到任何产业且机构报送为空，正确"],
    [2, "0", "-", "否", "-", "多报", "是", "是", "行业未匹配到任何产业且机构报送为不为空，多报"],
    [3, "1", "否", "否", "是", "正确", "否", "否", "行业匹配到的产业与报送产业一一对应，正确"],
    [4, "1", "是", "否", "是", "疑似正确", "否", "否", "行业*匹配到的产业与报送产业一一对应，疑似正确"],
    [5, "1", "否", "是", "-", "漏报", "是", "是", "行业匹配到产业但是机构未报送，漏报"],
    [6, "1", "是", "是", "-", "疑似漏报", "是", "是", "行业*匹配到产业但是机构未报送，疑似漏报"],
    [7, "1", "-", "否", "否", "错报", "是", "是", "行业（*）匹配到产业与机构报送产业不同，错报"],
    [8, ">1", "否", "否", "是", "正确", "否", "否", "行业匹配到的产业包含了机构报送的产业，正确"],
    [9, ">1", "是", "否", "是", "疑似正确", "否", "否", "行业*匹配到的产业包含了机构报送的产业，疑似正确"],
    [10, ">1", "否", "是", "-", "漏报", "是", "是", "行业匹配到多个产业但是机构未报送，漏报"],
    [11, ">1", "是", "是", "-", "疑似漏报", "是", "是", "行业*匹配到多个产业但是机构未报送，疑似漏报"],
    [12, ">1", "-", "否", "否", "错报", "是", "是", "行业（*）匹配到多个产业与机构报送产业不同，错报"],
]


def extract_industry4(text):
    if text is None:
        return ""
    s = str(text).upper()
    m = re.search(r"[A-Z]\d{4}", s)
    return m.group(0) if m else ""


def get_industry_desc_file_from_config():
    """从 config.xlsx 的 mapping 工作表读取 B2，即映射表（国民经济行业分类映射表）的文件路径。"""
    if not os.path.exists(CONFIG_FILE):
        return None
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        wb.close()
        return None
    path_cell = ws.cell(2, 2).value  # B2 = 映射表路径
    wb.close()
    if not path_cell:
        return None
    raw = str(path_cell).strip()
    if not raw:
        return None
    if os.path.isabs(raw):
        p = raw
    else:
        p = os.path.join(ROOT, raw)
    return p if os.path.exists(p) else None


def _load_field_config_rows(logical_name):
    rows = _load_named_sheet_rows(logical_name)
    normalized = []
    for row in rows:
        enabled = str(row.get("是否启用") or "").strip()
        if enabled in {"否", "0", "false", "False"}:
            continue
        normalized.append(row)
    return normalized


def _field_config_column_index(rows, field_name, preferred_sheet=None):
    target_field = f"列位置_{field_name}" if not str(field_name).startswith("列位置_") else str(field_name)
    for row in rows:
        sheet_name = str(row.get("来源Sheet") or "").strip()
        if preferred_sheet and sheet_name and sheet_name != preferred_sheet:
            continue
        if str(row.get("字段名称") or "").strip() != target_field:
            continue
        return parse_int(row.get("列序号"), 0)
    return 0


def load_industry_desc_map():
    """从映射表（国民经济行业分类映射表）读取 小类代码->(显示用代码, 小类描述)，描述去掉空格和换行。

    例如 J 列为 `C2345*`，L 列为 `某某行业`，则字典中存储为：
    key='C2345*'，value=('C2345*', '某某行业')
    最终“行业小类描述”列会展示为 `C2345*：某某行业`。
    路径优先从 config 的 mapping 工作表 B2 读取。
    """
    file_path = get_industry_desc_file_from_config() or INDUSTRY_DESC_FILE
    if not os.path.exists(file_path):
        return {}
    d = {}
    field_rows = _load_field_config_rows("国民经济映射字段配置")
    code_idx = (_field_config_column_index(field_rows, "小类编码") or INDUSTRY_MAP_CODE_COL) - 1
    desc_idx = (_field_config_column_index(field_rows, "小类描述") or INDUSTRY_MAP_DESC_COL) - 1
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or max(code_idx, desc_idx) >= len(row):
                continue
            code_raw = row[code_idx]
            desc_raw = row[desc_idx]
            if code_raw is None:
                continue
            display_code = str(code_raw).strip()  # 保留原样（含*等）
            code = display_code.upper()
            if not code:
                continue
            desc = "" if desc_raw is None else str(desc_raw)
            desc = re.sub(r"[\r\n\t]+", " ", desc)
            desc = re.sub(r"\s+", " ", desc).strip()
            value = (display_code, desc)
            d[code] = value
            # 若是类似 A0111，则同时支持用 0111 作为 key 查询
            if len(code) == 5 and code[0].isalpha() and code[1:].isdigit():
                d[code[1:]] = value
        wb.close()
    except Exception:
        pass
    return d


def load_industry_hierarchy_map():
    file_path = get_industry_desc_file_from_config() or INDUSTRY_DESC_FILE
    if not os.path.exists(file_path):
        return {}
    result = {}
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        field_rows = _load_field_config_rows("国民经济映射字段配置")
        if not field_rows:
            wb.close()
            return {}
        code_idx = (_field_config_column_index(field_rows, "小类编码") or 10) - 1
        big_idx = (_field_config_column_index(field_rows, "大类名称") or 5) - 1
        mid_idx = (_field_config_column_index(field_rows, "中类名称") or 8) - 1
        small_name_idx = (_field_config_column_index(field_rows, "小类名称") or 11) - 1
        small_desc_idx = (_field_config_column_index(field_rows, "小类描述") or 12) - 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            small_code = str((row[code_idx] if code_idx < len(row) else "") or "").strip().upper()
            if not small_code:
                continue
            entry = {
                "大类名称": str((row[big_idx] if big_idx < len(row) else "") or "").strip(),
                "中类名称": str((row[mid_idx] if mid_idx < len(row) else "") or "").strip(),
                "小类名称": str((row[small_name_idx] if small_name_idx < len(row) else "") or "").strip(),
                "小类说明": _clean_text_value(row[small_desc_idx] if small_desc_idx < len(row) else None),
            }
            result[small_code] = entry
            if len(small_code) == 5 and small_code[0].isalpha() and small_code[1:].isdigit():
                result[small_code[1:]] = entry
        wb.close()
    except Exception:
        return {}
    return result


def extract_codes(text):
    if text is None:
        return []
    s = str(text).upper()
    codes = re.findall(r"[A-Z]{2,4}\d{2,6}", s)
    out = []
    seen = set()
    for code in codes:
        if code not in seen:
            seen.add(code)
            out.append(code)
    return out


def is_star_value(v):
    if v is None:
        return False
    s = str(v).strip().upper()
    if not s:
        return False
    if s in {"是", "Y", "YES", "TRUE", "1", "*"}:
        return True
    if s in {"否", "N", "NO", "FALSE", "0"}:
        return False
    return False


def parse_int(v, default=0):
    try:
        return int(str(v).strip())
    except Exception:
        return default


def parse_idx_list(v):
    if v is None:
        return []
    parts = re.split(r"[，,、\s]+", str(v).strip())
    out = []
    for p in parts:
        if p.isdigit():
            out.append(int(p))
    return out


def parse_col_category_map(v):
    """
    解析示例: 17:HTP,20:HTS,23:SE,26:PA
    返回: [(17,'HTP'), (20,'HTS'), ...]
    """
    if v is None:
        return []
    s = str(v).strip()
    if not s:
        return []
    out = []
    for part in re.split(r"[，,、\s]+", s):
        if not part or ":" not in part:
            continue
        col_s, cat_s = part.split(":", 1)
        col_s = col_s.strip()
        cat = cat_s.strip().upper()
        if col_s.isdigit() and cat:
            out.append((int(col_s), cat))
    return out


def parse_sheet_selector(v, default=1):
    """
    解析参照表工作表序号：既支持数字（序号），也支持字符串（sheet名称）。
    """
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    if s.isdigit():
        return int(s)
    return s


def dedup_pairs_keep_order(pairs):
    out = []
    seen = set()
    for col, cat in pairs:
        key = (col, cat)
        if key not in seen:
            seen.add(key)
            out.append(key)
    return out


def dedup_keep_order(items):
    out = []
    seen = set()
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _split_csv_values(text):
    if text is None:
        return []
    parts = re.split(r"[，,、\s]+", str(text).strip())
    return [p for p in parts if p]


def _today_text():
    return datetime.now().strftime("%Y-%m-%d")


def get_default_detail_definition_rows():
    return [
        [1, "科技产业", "科技", "科技产业贷款", 3, 4, "", "当前伪数据命中：附表2 / 科技产业贷款明细", "科技产业贷款明细", r"伪数据\附科技", "科技", "万元", 4, 2, 13, 19, 21],
        [2, "数字经济产业", "数字经济", "数字经济产业贷款", 4, 5, "", "当前伪数据命中：附表4 / 数字经济产业贷款明细", "数字经济产业贷款明细", r"伪数据\附数字", "数字", "万元", 2, 2, 13, 19, 21],
        [3, "养老产业", "养老", "养老产业贷款", 3, 4, "", "当前伪数据命中：附表5 / 养老产业贷款明细", "养老产业贷款明细", r"伪数据\附养老", "养老", "万元", 3, 2, 13, 19, 21],
        [4, "绿色金融标准化", "绿色", "绿色金融", 3, 4, 1, "示例模板：启用后需补 明细参数 与 字段映射表", "", "", "", "万元", "", "", "", "", ""],
        [5, "普惠金融标准化", "普惠", "普惠金融", 3, 4, 1, "示例模板：启用后需补 明细参数 与 字段映射表", "", "", "", "万元", "", "", "", "", ""],
    ]


def get_default_detail_param_rows():
    rows = [
        [1, "科技产业", "列位置_贷款客户行业分类", "F", "exact,regex", "", "行业分类字段：建议先 exact，再 regex 提取 1位字母+4位数字代码。", None],
        [2, "科技产业", "列位置_贷款实际投向行业分类", "O", "exact,regex", "", "贷款实际投向行业字段：建议先 exact，再 regex 提取 1位字母+4位数字代码。", "2026-03-25"],
        [3, "科技产业", "列位置_是否高技术制造业贷款", "P", "exact", "", "布尔字段：0/1 映射为 否/是，建议仅用 exact。", "2026-03-25"],
        [4, "科技产业", "列位置_高技术制造业贷款类型大类编码", "Q", "exact,regex", "", "代码字段：支持 HTP03 或 03-xxx 统一为 HTP03。", "2026-03-25"],
        [5, "科技产业", "列位置_是否高技术服务业贷款", "S", "exact", "", "布尔字段：0/1 映射为 否/是，建议仅用 exact。", "2026-03-25"],
        [6, "科技产业", "列位置_高技术服务业贷款类型大类编码", "T", "exact,regex", "", "代码字段：支持 HTS01 或 01-xxx 统一为 HTS01。", "2026-03-25"],
        [7, "科技产业", "列位置_是否战略性新兴产业贷款", "V", "exact", "", "布尔字段：0/1 映射为 否/是，建议仅用 exact。", "2026-03-25"],
        [8, "科技产业", "列位置_战略性新兴产业贷款类型大类编码", "W", "exact,regex", "", "代码字段：支持 SE06 或 06-xxx 统一为 SE06。", "2026-03-25"],
        [9, "科技产业", "列位置_是否知识产权（专利）密集型产业贷款", "Y", "exact", "", "布尔字段：0/1 映射为 否/是，建议仅用 exact。", "2026-03-25"],
        [10, "科技产业", "列位置_知识产权（专利）密集型产业贷款类型大类编码", "Z", "exact,regex", "", "代码字段：支持 PA01 或 01-xxx 统一为 PA01。", "2026-03-25"],
        [11, "数字经济产业", "列位置_贷款客户行业分类", "F", "exact,regex", "", "行业分类字段：建议先 exact，再 regex 提取 1位字母+4位数字代码。", "2026-03-25"],
        [12, "数字经济产业", "列位置_贷款实际投向行业分类", "O", "exact,regex", "", "贷款实际投向行业字段：建议提取 C1442 这类代码。", "2026-03-25"],
        [13, "数字经济产业", "列位置_产业大类编码", "P", "exact,regex", "", "大类编码字段：支持 DE05 或 05-xxx 统一为 DE05。", "2026-03-25"],
        [14, "数字经济产业", "列位置_产业小类编码", "Q", "exact,regex", "", "156小类编码字段：支持 050908-xxx 或 050908-xxx* 提取编码。", "2026-03-25"],
        [15, "养老产业", "列位置_贷款客户行业分类", "F", "exact,regex", "", "行业分类字段：建议先 exact，再 regex 提取 1位字母+4位数字代码。", "2026-03-25"],
        [16, "养老产业", "列位置_贷款实际投向行业分类", "O", "exact,regex", "", "贷款实际投向行业字段：建议提取 R8626 / P8413 这类代码。", "2026-03-25"],
        [17, "养老产业", "列位置_产业大类编码", "P", "exact,regex", "", "大类编码字段：支持 EC09 或 09-xxx 统一为 EC09。", "2026-03-25"],
        [18, "养老产业", "列位置_产业小类编码", "Q", "exact,regex", "", "79小类编码字段：支持 0110-xxx 或 0110-xxx* 提取编码。", "2026-03-25"],
    ]
    next_seq = len(rows) + 1
    for detail_name, customer_col, actual_col, balance_col, reported_cols, col_map in [
        ("科技产业", 6, 15, 8, "17,20,23,26", "17:HTP,20:HTS,23:SE,26:PA"),
        ("数字经济产业", 6, 15, 8, "16", "16:DE"),
        ("养老产业", 6, 15, 8, "16", "16:EC"),
    ]:
        for field_name, value in [
            ("贷款客户行业列序号", customer_col),
            ("贷款投向行业列序号", actual_col),
            ("贷款余额列序号", balance_col),
            ("机构报送产业分类列序号", reported_cols),
            ("报送列-类别映射", col_map),
        ]:
            rows.append([next_seq, detail_name, field_name, value, "exact", "", "由运行配置合并而来", _today_text()])
            next_seq += 1
    return rows


def _detail_definition_row_from_config(row_dict, base_row=None):
    row = list(base_row) if base_row else [""] * len(DETAIL_DEFINITION_ALL_HEADERS)
    if len(row) < len(DETAIL_DEFINITION_ALL_HEADERS):
        row += [""] * (len(DETAIL_DEFINITION_ALL_HEADERS) - len(row))
    row[8] = row_dict["输出工作表名称"]
    row[9] = row_dict["来源台账文件"]
    row[10] = TYPE_KEY_TO_LABEL.get(row_dict["台账类型"], row_dict["台账类型"])
    row[11] = row_dict.get("贷款余额原始单位", "万元")
    row[12] = row_dict["参照表工作表序号"]
    row[13] = row_dict["参照表产业分类代码列序号"]
    row[14] = row_dict["参照表行业4位码列序号"]
    row[15] = row_dict["参照表星标列序号"]
    row[16] = row_dict.get("参照表原始映射列序号", 21)
    return row


def _runtime_detail_rows_from_config_rows(config_rows):
    detail_rows = [list(DETAIL_DEFINITION_ALL_HEADERS)]
    param_rows = [list(DETAIL_PARAM_HEADERS)]
    detail_name_to_row = {str(row[1]).strip(): list(row) for row in get_default_detail_definition_rows() if len(row) > 1 and str(row[1]).strip()}
    param_seq = 1
    for seq, cfg in enumerate(config_rows, start=1):
        detail_name = detail_name_from_cfg(cfg)
        base_row = detail_name_to_row.get(detail_name, [seq, detail_name, "", "", cfg["表头行号"], cfg["数据起始行号"], "", ""])
        if len(base_row) < len(DETAIL_DEFINITION_ALL_HEADERS):
            base_row += [""] * (len(DETAIL_DEFINITION_ALL_HEADERS) - len(base_row))
        base_row[0] = seq
        base_row[1] = detail_name
        base_row[4] = cfg["表头行号"]
        base_row[5] = cfg["数据起始行号"]
        detail_rows.append(_detail_definition_row_from_config(cfg, base_row))
        for field_name, value in [
            ("贷款客户行业列序号", cfg.get("贷款客户行业列序号", 6)),
            ("贷款投向行业列序号", cfg["贷款投向行业列序号"]),
            ("贷款余额列序号", cfg.get("贷款余额列序号", 8)),
            ("机构报送产业分类列序号", ",".join(str(x) for x in cfg["机构报送产业分类列序号"])),
            ("报送列-类别映射", ",".join(f"{c}:{t}" for c, t in cfg["报送列-类别映射"])),
        ]:
            param_rows.append([param_seq, detail_name, field_name, value, "exact", "", "由运行配置合并而来", _today_text()])
            param_seq += 1
    return detail_rows, param_rows


def _extend_headers_if_needed(existing_headers, required_headers):
    headers = list(existing_headers)
    for header in required_headers:
        if header not in headers:
            headers.append(header)
    return headers


def get_default_field_mapping_rows():
    return [
        [1, "科技产业", "列位置_是否高技术制造业贷款", "0", "", "", "否", "", "", "布尔字段 exact：0 -> 否"],
        [2, "科技产业", "列位置_是否高技术制造业贷款", "1", "", "", "是", "", "", "布尔字段 exact：1 -> 是"],
        [3, "科技产业", "列位置_是否高技术服务业贷款", "0", "", "", "否", "", "", "布尔字段 exact：0 -> 否"],
        [4, "科技产业", "列位置_是否高技术服务业贷款", "1", "", "", "是", "", "", "布尔字段 exact：1 -> 是"],
        [5, "科技产业", "列位置_是否战略性新兴产业贷款", "0", "", "", "否", "", "", "布尔字段 exact：0 -> 否"],
        [6, "科技产业", "列位置_是否战略性新兴产业贷款", "1", "", "", "是", "", "", "布尔字段 exact：1 -> 是"],
        [7, "科技产业", "列位置_是否知识产权（专利）密集型产业贷款", "0", "", "", "否", "", "", "布尔字段 exact：0 -> 否"],
        [8, "科技产业", "列位置_是否知识产权（专利）密集型产业贷款", "1", "", "", "是", "", "", "布尔字段 exact：1 -> 是"],
        [9, "科技产业", "列位置_贷款客户行业分类", "", "", r"^([A-Za-z]\d{4})(?:[-－].+)?(\*)?$", "", "{REGEX_GROUP_1}", "", "提取 1位字母+4位数字，不保留*"],
        [10, "科技产业", "列位置_贷款实际投向行业分类", "", "", r"^([A-Za-z]\d{4})(?:[-－].+)?(\*)?$", "", "{REGEX_GROUP_1}", "", "提取 1位字母+4位数字，不保留*"],
        [11, "科技产业", "列位置_高技术制造业贷款类型大类编码", "", "", r"^(HTP\d{2})(?:[-－].+)?$", "", "{REGEX_GROUP_1}", "", "若原值已为 HTP03 或 HTP03-文字，则提取 HTP03"],
        [12, "科技产业", "列位置_高技术制造业贷款类型大类编码", "", "", r"^(\d{2})(?:[-－].+)?$", "", "HTP{REGEX_GROUP_1}", "", "提取前两位数字并拼接 HTP；如 03-计算机及办公设备制造业 -> HTP03"],
        [13, "科技产业", "列位置_高技术服务业贷款类型大类编码", "", "", r"^(HTS\d{2})(?:[-－].+)?$", "", "{REGEX_GROUP_1}", "", "若原值已为 HTS01 或 HTS01-文字，则提取 HTS01"],
        [14, "科技产业", "列位置_高技术服务业贷款类型大类编码", "", "", r"^(\d{2})(?:[-－].+)?$", "", "HTS{REGEX_GROUP_1}", "", "提取前两位数字并拼接 HTS；如 01-信息服务 -> HTS01"],
        [15, "科技产业", "列位置_战略性新兴产业贷款类型大类编码", "", "", r"^(SE\d{2})(?:[-－].+)?$", "", "{REGEX_GROUP_1}", "", "若原值已为 SE06 或 SE06-文字，则提取 SE06"],
        [16, "科技产业", "列位置_战略性新兴产业贷款类型大类编码", "", "", r"^(\d{2})(?:[-－].+)?(\*)?$", "", "SE{REGEX_GROUP_1}", "", "提取前两位数字并拼接 SE，不保留*"],
        [17, "科技产业", "列位置_知识产权（专利）密集型产业贷款类型大类编码", "", "", r"^(PA\d{2})(?:[-－].+)?$", "", "{REGEX_GROUP_1}", "", "若原值已为 PA01 或 PA01-文字，则提取 PA01"],
        [18, "科技产业", "列位置_知识产权（专利）密集型产业贷款类型大类编码", "", "", r"^(\d{2})(?:[-－].+)?$", "", "PA{REGEX_GROUP_1}", "", "提取前两位数字并拼接 PA；如 01-某类 -> PA01"],
        [19, "数字经济产业", "列位置_贷款客户行业分类", "", "", r"^([A-Za-z]\d{4})(?:[-－].+)?(\*)?$", "", "{REGEX_GROUP_1}", "", "提取 1位字母+4位数字，不保留*"],
        [20, "数字经济产业", "列位置_贷款实际投向行业分类", "", "", r"^([A-Za-z]\d{4})(?:[-－].+)?(\*)?$", "", "{REGEX_GROUP_1}", "", "提取 1位字母+4位数字，不保留*"],
        [21, "数字经济产业", "列位置_产业大类编码", "", "", r"^(DE\d{2})(?:[-－].+)?$", "", "{REGEX_GROUP_1}", "", "若原值已为 DE05 或 DE05-文字，则提取 DE05"],
        [22, "数字经济产业", "列位置_产业大类编码", "", "", r"^(\d{2})(?:[-－].+)?$", "", "DE{REGEX_GROUP_1}", "", "提取前两位数字并拼接 DE；如 05-数字化效率提升业 -> DE05"],
        [23, "数字经济产业", "列位置_产业小类编码", "", "", r"^(\d{6})(?:[-－].+)?(\*)?$", "", "DE{REGEX_GROUP_1}", "", "提取 6位编码，允许前面带1位字母，不保留*"],
        [24, "养老产业", "列位置_贷款客户行业分类", "", "", r"^([A-Za-z]\d{4})(?:[-－].+)?(\*)?$", "", "{REGEX_GROUP_1}", "", "提取 1位字母+4位数字，不保留*"],
        [25, "养老产业", "列位置_贷款实际投向行业分类", "", "", r"^([A-Za-z]\d{4})(?:[-－].+)?(\*)?$", "", "{REGEX_GROUP_1}", "", "提取 1位字母+4位数字，不保留*"],
        [26, "养老产业", "列位置_产业大类编码", "", "", r"^(EC\d{2})(?:[-－].+)?$", "", "{REGEX_GROUP_1}", "", "若原值已为 EC09 或 EC09-文字，则提取 EC09"],
        [27, "养老产业", "列位置_产业大类编码", "", "", r"^(\d{2})(?:[-－].+)?$", "", "EC{REGEX_GROUP_1}", "", "提取前两位数字并拼接 EC；如 09-其他养老服务 -> EC09"],
        [28, "养老产业", "列位置_产业小类编码", "", "", r"^(\d{4})(?:[-－].+)?(\*)?$", "", "{REGEX_GROUP_1}", "", "提取 4位编码，不保留*"],
    ]


def get_default_keyword_mapping_rows():
    return [
        [1, "科技企业名称示例", "科技产业", "客户名称", "E", "contains", "科技,软件,信息", "", "", 0, "", "", "", "企业名称命中科技相关词", "是", 1, "示例规则，默认禁用"],
        [2, "数字用途示例", "数字经济产业", "贷款用途", "N", "contains", "数字,平台,系统", "", "数据,平台|系统,智能", 2, "是", "", "中类,小类,说明", "贷款用途命中数字经济相关词", "是", 1, "示例规则，默认禁用"],
        [3, "养老用途示例", "养老产业", "贷款用途", "N", "contains", "养老,护理,康养", "", "", 0, "", "是", "小类,说明", "贷款用途命中养老相关词", "是", 1, "示例规则，默认禁用"],
    ]


def get_default_text_rule_rows():
    return [
        ["TXT001", "科技产业", "证据不足", "列位:R[纳入高技术制造业贷款依据]|列位:U[纳入高技术服务业贷款依据]|列位:X[纳入战略性新兴产业贷款依据]|列位:AA[纳入知识产权（专利）密集型产业贷款依据]", "empty_or_short", "8", "", "科技产业依据摘要为空或过短", "medium", ""],
        ["TXT002", "数字经济产业", "证据不足", "列位:R[纳入数字经济产业贷款依据]", "empty_or_short", "8", "", "数字经济产业依据摘要为空或过短", "medium", ""],
        ["TXT003", "养老产业", "证据不足", "列位:R[纳入养老产业贷款依据]", "empty_or_short", "8", "", "养老产业依据摘要为空或过短", "medium", ""],
        ["TXT004", "全部", "文本内部冲突", "列位:M[贷款类型]|列位:N[贷款合同中约定的贷款用途]", "keyword_pair", "流动资金贷款|经营贷款", "建设|购置固定资产|厂房|土建", "贷款类型与用途不一致", "high", ""],
        ["TXT005", "全部", "文本内部冲突", "列位:M[贷款类型]|列位:N[贷款合同中约定的贷款用途]", "keyword_pair", "固定资产贷款", "日常经营|补充流动资金", "固定资产贷款与用途不一致", "medium", ""],
        ["TXT006", "科技产业", "字段间冲突", "列位:N[贷款合同中约定的贷款用途]|列位:R[纳入高技术制造业贷款依据]|列位:U[纳入高技术服务业贷款依据]|列位:X[纳入战略性新兴产业贷款依据]|列位:AA[纳入知识产权（专利）密集型产业贷款依据]", "negative_keyword", "", "养老|护理|养老机构|医院|病房", "科技产业文本出现明显养老/医疗语义", "high", ""],
        ["TXT007", "数字经济产业", "字段间冲突", "列位:N[贷款合同中约定的贷款用途]|列位:R[纳入数字经济产业贷款依据]", "negative_keyword", "", "养老|护理|养老机构|医院|病房", "数字经济文本出现明显养老/医疗语义", "high", ""],
        ["TXT008", "养老产业", "字段间冲突", "列位:N[贷款合同中约定的贷款用途]|列位:R[纳入养老产业贷款依据]", "negative_keyword", "", "芯片|半导体|研发平台|软件开发|信息服务", "养老产业文本出现明显科技/数字语义", "high", ""],
        ["TXT009", "科技产业", "布尔标记与文本冲突", "列位置_是否高技术服务业贷款|列位置_高技术服务业贷款类型大类编码|列位:U[纳入高技术服务业贷款依据]", "flag_requires_keywords", "研发|信息服务|技术服务|成果转化|检验检测", "", "高技术服务业标记缺少文本支撑", "medium", ""],
        ["TXT010", "科技产业", "布尔标记与文本冲突", "列位置_是否高技术制造业贷款|列位置_高技术制造业贷款类型大类编码|列位:R[纳入高技术制造业贷款依据]", "flag_requires_keywords", "制造|设备|生产线|工艺|材料", "", "高技术制造业标记缺少文本支撑", "medium", ""],
        ["TXT101", "科技产业", "语义支持", "列位:N[贷款合同中约定的贷款用途]", "keyword_bag", "研发|技术|科技|信息服务|知识产权|专利|成果转化|高技术", "养老|护理|养老机构|医院", "", "", 2.0],
        ["TXT102", "科技产业", "语义支持", "列位:R[纳入高技术制造业贷款依据]|列位:U[纳入高技术服务业贷款依据]|列位:X[纳入战略性新兴产业贷款依据]|列位:AA[纳入知识产权（专利）密集型产业贷款依据]", "keyword_bag", "高技术|科技成果|研发|知识产权|专利|信息服务", "养老|护理|养老机构|医院", "", "", 3.0],
        ["TXT103", "科技产业", "语义支持", "列位:E[贷款客户名称]", "keyword_bag", "科技|信息|软件|电子", "养老|护理", "", "", 1.0],
        ["TXT201", "数字经济产业", "语义支持", "列位:N[贷款合同中约定的贷款用途]", "keyword_bag", "数字|数据|软件|平台|系统|互联网|信息服务|电子商务|智能", "养老|护理|病房", "", "", 2.0],
        ["TXT202", "数字经济产业", "语义支持", "列位:R[纳入数字经济产业贷款依据]", "keyword_bag", "数字|数据|软件|平台|系统|互联网|核心产业|智能", "养老|护理|病房", "", "", 3.0],
        ["TXT203", "数字经济产业", "语义支持", "列位:E[贷款客户名称]", "keyword_bag", "数据|网络|软件|信息|电子", "养老|护理", "", "", 1.0],
        ["TXT301", "养老产业", "语义支持", "列位:N[贷款合同中约定的贷款用途]", "keyword_bag", "养老|老年|护理|照护|康养|养老机构|居家养老|助餐", "芯片|研发|软件|信息服务", "", "", 2.0],
        ["TXT302", "养老产业", "语义支持", "列位:R[纳入养老产业贷款依据]", "keyword_bag", "养老|老年|护理|照护|康养|民政|养老机构|适老", "芯片|研发|软件|信息服务", "", "", 3.0],
        ["TXT303", "养老产业", "语义支持", "列位:E[贷款客户名称]", "keyword_bag", "养老|健康|康养|护理", "科技|软件|信息", "", "", 1.0],
    ]


def get_default_text_verdict_rows():
    return [
        [1, 0, 100, "high", "明显矛盾", "优先核对贷款用途、贷款类型与佐证摘要"],
        [2, 0, 100, "medium", "建议复核", "补充更具体的结构化字段和佐证材料"],
        [3, 0, 69.99, "none", "建议复核", "文本支持度不足，建议补充用途和依据描述"],
        [4, 70, 100, "none", "建议纳统", "文本与结构化信息基本一致，可进入人工抽查"],
    ]


def get_default_text_stopword_rows():
    return [
        [1, "用于生产经营", "", "弱信息词"],
        [2, "支持企业发展", "", "弱信息词"],
        [3, "依据参照表映射", "", "模板话术"],
        [4, "根据统计口径", "", "模板话术"],
        [5, "项目建设", "", "泛化表述"],
    ]


LOG_HEADERS = [
    "运行时间",
    "参照表路径",
    "参照表sheet数量",
    "参照表sheet列表",
    "结果文件",
    "明细sheet名称",
    "来源台账文件",
    "台账类型",
    "使用参照表sheet序号",
    "使用参照表sheet名称",
    "多报数量",
    "漏报数量",
    "疑似多报数量",
    "疑似漏报数量",
]


def _config_row_values_from_row_dict(row_dict):
    type_label = TYPE_KEY_TO_LABEL.get(row_dict["台账类型"], row_dict["台账类型"])
    return [
        row_dict["输出工作表名称"],
        row_dict["来源台账文件"],
        type_label,
        row_dict["表头行号"],
        row_dict["数据起始行号"],
        row_dict.get("贷款客户行业列序号", 6),
        row_dict["贷款投向行业列序号"],
        row_dict.get("贷款余额列序号", 8),
        row_dict.get("贷款余额原始单位", "万元"),
        ",".join(str(x) for x in row_dict["机构报送产业分类列序号"]),
        ",".join(f"{c}:{t}" for c, t in row_dict["报送列-类别映射"]),
        row_dict["参照表工作表序号"],
        row_dict["参照表产业分类代码列序号"],
        row_dict["参照表行业4位码列序号"],
        row_dict["参照表星标列序号"],
        row_dict.get("参照表原始映射列序号", 21),
    ]


def _read_existing_mapping_rows(cfg_path):
    if not os.path.exists(cfg_path):
        return [["参照表路径", ""], ["映射表路径", ""], ["源文件", "源文件路径"]]
    wb = openpyxl.load_workbook(cfg_path, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        wb.close()
        return [["参照表路径", ""], ["映射表路径", ""], ["源文件", "源文件路径"]]
    rows = [[cell for cell in row] for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]
    wb.close()
    if not rows:
        return [["参照表路径", ""], ["映射表路径", ""], ["源文件", "源文件路径"]]
    has_source_header = any(str((row[0] if len(row) > 0 else "") or "").strip() == "源文件" and str((row[1] if len(row) > 1 else "") or "").strip() == "源文件路径" for row in rows)
    if not has_source_header:
        rows.append(["源文件", "源文件路径"])
    return rows


def _build_initial_config_row_dicts(cfg_path):
    source_map = get_source_file_map_from_mapping() if os.path.exists(cfg_path) else {}
    paths = list(source_map.values()) if source_map else []
    if not paths:
        paths = find_ledger_files()
    rows = build_config_rows_from_files(paths)
    return rows or DEFAULT_CONFIG_ROWS


def initialize_config_workbook(cfg_path=CONFIG_FILE):
    rows = _build_initial_config_row_dicts(cfg_path)
    mapping_rows = _read_existing_mapping_rows(cfg_path)
    detail_rows, param_rows = _runtime_detail_rows_from_config_rows(rows)

    wb = Workbook()
    ws_m = wb.active
    ws_m.title = PREFERRED_SHEET_NAMES["mapping"]
    for row in mapping_rows:
        ws_m.append(row)

    ws_l = wb.create_sheet(PREFERRED_SHEET_NAMES["log"])
    ws_l.append(LOG_HEADERS)
    for c in range(1, len(LOG_HEADERS) + 1):
        ws_l.cell(1, c).font = Font(name="Arial", bold=True)

    ws_clue = wb.create_sheet(PREFERRED_SHEET_NAMES["clue"])
    ws_clue.append(CLUE_HEADERS)
    for c in range(1, len(CLUE_HEADERS) + 1):
        ws_clue.cell(1, c).font = Font(name="Arial", bold=True)
    for row in CLUE_DEFAULT_ROWS:
        ws_clue.append(row)

    ws_keyword_mapping = wb.create_sheet(PREFERRED_SHEET_NAMES["关键字映射"])
    ws_keyword_mapping.append(KEYWORD_MAPPING_HEADERS)
    for c in range(1, len(KEYWORD_MAPPING_HEADERS) + 1):
        ws_keyword_mapping.cell(1, c).font = Font(name="Arial", bold=True)
    for row in get_default_keyword_mapping_rows():
        ws_keyword_mapping.append(row)

    ws_runtime = wb.create_sheet(PREFERRED_SHEET_NAMES["runtime"])
    for row in _merge_runtime_rows([]):
        ws_runtime.append(row)
    for c in range(1, len(RUNTIME_HEADERS) + 1):
        ws_runtime.cell(1, c).font = Font(name="Arial", bold=True)

    ws_detail = wb.create_sheet(PREFERRED_SHEET_NAMES["明细定义"])
    for row in detail_rows:
        ws_detail.append(row)
    for c in range(1, len(detail_rows[0]) + 1):
        ws_detail.cell(1, c).font = Font(name="Arial", bold=True)

    ws_param = wb.create_sheet(PREFERRED_SHEET_NAMES["明细参数"])
    for row in param_rows:
        ws_param.append(row)
    for c in range(1, len(param_rows[0]) + 1):
        ws_param.cell(1, c).font = Font(name="Arial", bold=True)

    ws_mapping = wb.create_sheet(PREFERRED_SHEET_NAMES["字段映射表"])
    ws_mapping.append(FIELD_MAPPING_HEADERS)
    for c in range(1, len(FIELD_MAPPING_HEADERS) + 1):
        ws_mapping.cell(1, c).font = Font(name="Arial", bold=True)
    for row in get_default_field_mapping_rows():
        ws_mapping.append(row)

    os.makedirs(os.path.dirname(cfg_path), exist_ok=True)
    wb.save(cfg_path)
    wb.close()
    return rows


def _append_sheet_with_headers(ws, headers):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(name="Arial", bold=True)


def _sheet_rows_or_defaults(wb, sheet_name, default_headers, default_rows):
    ws = get_sheet_by_logical_name(wb, sheet_name)
    if ws is None:
        return [default_headers] + [list(row) for row in default_rows]
    rows = [
        [cell for cell in row]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
    ]
    if not rows:
        return [default_headers] + [list(row) for row in default_rows]
    return rows


def _extract_runtime_rows_from_detail_sheets(wb):
    detail_ws = get_sheet_by_logical_name(wb, "明细定义")
    param_ws = get_sheet_by_logical_name(wb, "明细参数")
    if detail_ws is None or param_ws is None:
        return []

    detail_headers = [detail_ws.cell(1, i).value for i in range(1, detail_ws.max_column + 1)]
    param_headers = [param_ws.cell(1, i).value for i in range(1, param_ws.max_column + 1)]
    detail_headers = [str(x).strip() if x is not None else "" for x in detail_headers]
    param_headers = [str(x).strip() if x is not None else "" for x in param_headers]
    if not all(header in detail_headers for header in DETAIL_DEFINITION_CONFIG_HEADERS):
        return []
    param_header_map = {name: idx for idx, name in enumerate(param_headers) if name}
    if not all(header in param_header_map for header in ["明细名称", "明细字段", "字段所在列位置", "是否禁用"]):
        return []

    param_lookup = defaultdict(dict)
    for row in param_ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        detail_name = str(row[param_header_map["明细名称"]] or "").strip()
        field_name = str(row[param_header_map["明细字段"]] or "").strip()
        disabled = row[param_header_map["是否禁用"]]
        if not detail_name or not field_name or str(disabled or "").strip() in {"1", "是", "true", "True"}:
            continue
        value = row[param_header_map["字段所在列位置"]]
        param_lookup[detail_name][field_name] = value

    detail_header_map = {name: idx for idx, name in enumerate(detail_headers) if name}
    rows = []
    for row in detail_ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        detail_name = str(row[detail_header_map["明细名称"]] or "").strip()
        output_sheet = str(row[detail_header_map["输出工作表名称"]] or "").strip()
        source_file = str(row[detail_header_map["来源台账文件"]] or "").strip()
        type_label = str(row[detail_header_map["台账类型"]] or "").strip()
        file_type = TYPE_LABEL_TO_KEY.get(type_label, "")
        if not detail_name or not output_sheet or not source_file or not file_type:
            continue
        params = param_lookup.get(detail_name, {})
        reported_cols = parse_idx_list(params.get("机构报送产业分类列序号"))
        col_cat_map = dedup_pairs_keep_order(parse_col_category_map(params.get("报送列-类别映射")))
        if not col_cat_map:
            col_cat_map = _default_col_category_map(file_type, reported_cols)
        if file_type in {"digital", "elder"} and len(reported_cols) > 1:
            reported_cols = [reported_cols[-1]]
        rows.append(
            {
                "输出工作表名称": output_sheet,
                "来源台账文件": source_file,
                "台账类型": file_type,
                "表头行号": parse_int(row[detail_header_map["表头行号"]], 1),
                "数据起始行号": parse_int(row[detail_header_map["数据起始行号"]], 2),
                "贷款客户行业列序号": parse_int(params.get("贷款客户行业列序号"), 6),
                "贷款投向行业列序号": parse_int(params.get("贷款投向行业列序号"), 15),
                "贷款余额列序号": parse_int(params.get("贷款余额列序号"), 8),
                "贷款余额原始单位": str(row[detail_header_map["贷款余额原始单位"]] or "万元").strip() or "万元",
                "机构报送产业分类列序号": reported_cols,
                "报送列-类别映射": col_cat_map,
                "参照表工作表序号": parse_sheet_selector(row[detail_header_map["参照表工作表序号"]], 4),
                "参照表产业分类代码列序号": parse_int(row[detail_header_map["参照表产业分类代码列序号"]], 2),
                "参照表行业4位码列序号": parse_int(row[detail_header_map["参照表行业4位码列序号"]], 13),
                "参照表星标列序号": parse_int(row[detail_header_map["参照表星标列序号"]], 19),
                "参照表原始映射列序号": parse_int(row[detail_header_map["参照表原始映射列序号"]], 21),
            }
        )
    return rows


def _normalize_keyword_mapping_rows(rows):
    if not rows:
        return [KEYWORD_MAPPING_HEADERS] + [list(row) for row in get_default_keyword_mapping_rows()]
    headers = [str(item or "").strip() for item in rows[0]]
    if headers[: len(KEYWORD_MAPPING_HEADERS)] == KEYWORD_MAPPING_HEADERS:
        return rows
    header_map = {name: idx for idx, name in enumerate(headers) if name}
    normalized = [KEYWORD_MAPPING_HEADERS]
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        seq = row[header_map["序号"]] if "序号" in header_map and header_map["序号"] < len(row) else None
        detail_name = row[header_map["明细名称"]] if "明细名称" in header_map and header_map["明细名称"] < len(row) else None
        field_name = row[header_map["列字段"]] if "列字段" in header_map and header_map["列字段"] < len(row) else None
        contains_value = row[header_map["原始列值_contains"]] if "原始列值_contains" in header_map and header_map["原始列值_contains"] < len(row) else None
        regex_value = row[header_map["原始列值_regex"]] if "原始列值_regex" in header_map and header_map["原始列值_regex"] < len(row) else None
        disabled = row[header_map["是否禁用"]] if "是否禁用" in header_map and header_map["是否禁用"] < len(row) else None
        note = row[header_map["备注"]] if "备注" in header_map and header_map["备注"] < len(row) else None
        mode_parts = []
        if contains_value not in (None, ""):
            mode_parts.append("contains")
        if regex_value not in (None, ""):
            mode_parts.append("regex")
        normalized.append(
            [
                seq,
                f"{detail_name}-{field_name}" if detail_name and field_name else None,
                detail_name,
                field_name,
                "",
                ",".join(mode_parts),
                contains_value or regex_value,
                "",
                "",
                0,
                "",
                "",
                "",
                note,
                "是",
                disabled,
                "由旧版关键字映射结构迁移",
            ]
        )
    return normalized


def _unit_scale(source_unit, target_unit):
    unit_to_yuan = {"元": 1.0, "万元": 10000.0, "亿元": 100000000.0}
    source = unit_to_yuan.get(str(source_unit).strip(), unit_to_yuan["万元"])
    target = unit_to_yuan.get(str(target_unit).strip(), unit_to_yuan["万元"])
    return source / target


def _split_rule_terms(text):
    if text is None:
        return []
    return [item.strip() for item in str(text).split("|") if item and item.strip()]


def _clean_text_value(text, stopwords=None):
    if text is None:
        return ""
    value = str(text).strip()
    if not value:
        return ""
    for stopword in stopwords or []:
        value = value.replace(stopword, "")
    return re.sub(r"\s+", " ", value).strip()


def _text_contains_any(text, terms):
    if not text:
        return []
    hits = []
    for term in terms:
        if term and term in text:
            hits.append(term)
    return dedup_keep_order(hits)


def _is_truthy_text_flag(value):
    if value is None:
        return False
    text = str(value).strip()
    if not text:
        return False
    upper = text.upper()
    if upper in {"0", "否", "N", "NO", "FALSE"}:
        return False
    return True


def load_runtime_settings():
    clue_rules = load_clue_rules()
    available_main_labels = dedup_keep_order(
        str(rule.get("主标签") or "").strip() for rule in clue_rules if str(rule.get("主标签") or "").strip()
    )
    default_enabled_result_types = [label for label in available_main_labels if label != "正确"] or RESULT_TYPE_ORDER[:]
    settings = {
        "enabled_bases": ["actual", "customer"],
        "enabled_result_types": default_enabled_result_types,
        "summary_unit": "万元",
        "available_main_labels": available_main_labels,
        "template_row_labels": {
            row_label: _split_csv_values(default_value)
            for row_label, runtime_key in SUMMARY_TEMPLATE_RUNTIME_KEYS.items()
            for default_key, default_value, _ in RUNTIME_DEFAULT_ROWS
            if runtime_key == default_key
        },
    }
    if not os.path.exists(CONFIG_FILE):
        return settings
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, "runtime")
    if ws is not None:
        runtime_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            value = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
            if key:
                runtime_map[key] = value
        bases = []
        for item in _split_csv_values(runtime_map.get("启用核查口径")):
            key = BASIS_RUNTIME_TO_KEY.get(str(item).strip(), BASIS_RUNTIME_TO_KEY.get(str(item).strip().lower()))
            if key in BASIS_NAME_MAP:
                bases.append(key)
        if bases:
            settings["enabled_bases"] = dedup_keep_order(bases)
        result_types = []
        for item in _split_csv_values(runtime_map.get("启用结果类型")):
            if item in KEYWORD_RESULT_ALIASES:
                result_types.append(KEYWORD_RESULT_LABEL)
            elif not available_main_labels or item in available_main_labels:
                result_types.append(item)
        if result_types:
            settings["enabled_result_types"] = dedup_keep_order(result_types)
        if runtime_map.get("汇总报告单位"):
            settings["summary_unit"] = runtime_map["汇总报告单位"]
        for row_label, runtime_key in SUMMARY_TEMPLATE_RUNTIME_KEYS.items():
            if runtime_key in runtime_map:
                settings["template_row_labels"][row_label] = _split_csv_values(runtime_map.get(runtime_key))
    wb.close()
    return settings


def _load_named_sheet_rows(sheet_name):
    if not os.path.exists(CONFIG_FILE):
        return []
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, sheet_name)
    if ws is None:
        wb.close()
        return []
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            if header is None:
                continue
            item[str(header).strip()] = row[idx] if idx < len(row) else None
        disabled = str(item.get("是否禁用") or "").strip()
        if disabled == "1":
            continue
        rows.append(item)
    wb.close()
    return rows


def load_text_analysis_settings():
    def _rows_to_dicts(headers, rows):
        return [{headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))} for row in rows]

    rules = _load_named_sheet_rows("文本规则")
    verdicts = _load_named_sheet_rows("文本结论")
    stopword_rows = _load_named_sheet_rows("文本停用词")
    stopwords = [str(row.get("词语")).strip() for row in stopword_rows if row.get("词语") and str(row.get("是否禁用") or "").strip() != "1"]
    return {
        "rules": rules or _rows_to_dicts(TEXT_RULE_HEADERS, get_default_text_rule_rows()),
        "verdicts": verdicts or _rows_to_dicts(TEXT_VERDICT_HEADERS, get_default_text_verdict_rows()),
        "stopwords": stopwords or [row[1] for row in get_default_text_stopword_rows() if row[1]],
    }


def load_detail_param_lookup():
    rows = _load_named_sheet_rows("明细参数")
    lookup = defaultdict(dict)
    for row in rows:
        detail_name = str(row.get("明细名称") or row.get("适用明细") or "").strip()
        field_name = str(row.get("明细字段") or row.get("字段名称") or "").strip()
        column_letter = str(row.get("字段所在列位置") or row.get("列字母") or "").strip().upper()
        if detail_name and field_name and column_letter:
            lookup[detail_name][field_name] = {
                "column_letter": column_letter,
                "label": field_name,
            }
    return lookup


def load_keyword_rule_rows():
    rows = _load_named_sheet_rows("关键字映射")
    normalized = []
    for row in rows:
        detail_name = str(row.get("明细名称") or "").strip()
        field_name = str(row.get("列字段") or "").strip()
        column_letter = str(row.get("字段所在列位置") or "").strip().upper()
        if not detail_name or (not field_name and not column_letter):
            continue
        row = dict(row)
        row["字段所在列位置"] = column_letter
        normalized.append(row)
    return normalized


def _resolve_keyword_rule_entry(rule, field_catalog):
    field_name = str(rule.get("列字段") or "").strip()
    if field_name and field_name in field_catalog:
        return field_name, field_catalog[field_name]
    column_letter = str(rule.get("字段所在列位置") or "").strip().upper()
    if column_letter:
        explicit_key = f"列位:{column_letter}"
        entry = field_catalog.get(explicit_key)
        if entry is not None:
            return field_name or explicit_key, entry
    return field_name or "", None


def _legacy_text_field_specs(detail_name):
    base = {
        "客户名称": {"column_letter": "E", "label": "贷款客户名称"},
        "贷款类型": {"column_letter": "M", "label": "贷款类型"},
        "贷款用途": {"column_letter": "N", "label": "贷款合同中约定的贷款用途"},
        "贷款客户行业分类": {"column_letter": "F", "label": "贷款客户行业分类"},
        "贷款实际投向行业": {"column_letter": "O", "label": "贷款实际投向行业"},
    }
    if detail_name == "科技产业":
        base.update(
            {
                "佐证摘要": {"aggregate": ["列位:R[纳入高技术制造业贷款依据]", "列位:U[纳入高技术服务业贷款依据]", "列位:X[纳入战略性新兴产业贷款依据]", "列位:AA[纳入知识产权（专利）密集型产业贷款依据]"], "label": "科技产业依据摘要"},
                "是否高技术制造业贷款": {"column_letter": "P", "label": "是否高技术制造业贷款"},
                "高技术制造业贷款类型大类编码": {"column_letter": "Q", "label": "高技术制造业贷款类型大类编码"},
                "是否高技术服务业贷款": {"column_letter": "S", "label": "是否高技术服务业贷款"},
                "高技术服务业贷款类型大类编码": {"column_letter": "T", "label": "高技术服务业贷款类型大类编码"},
                "是否战略性新兴产业贷款": {"column_letter": "V", "label": "是否战略性新兴产业贷款"},
                "战略性新兴产业贷款类型大类编码": {"column_letter": "W", "label": "战略性新兴产业贷款类型大类编码"},
                "是否知识产权（专利）密集型产业贷款": {"column_letter": "Y", "label": "是否知识产权（专利）密集型产业贷款"},
                "知识产权（专利）密集型产业贷款类型大类编码": {"column_letter": "Z", "label": "知识产权（专利）密集型产业贷款类型大类编码"},
            }
        )
    elif detail_name == "数字经济产业":
        base.update(
            {
                "佐证摘要": {"column_letter": "R", "label": "纳入数字经济产业贷款依据"},
                "产业大类编码": {"column_letter": "P", "label": "数字经济产业贷款类型（大类代码-名称）"},
                "产业小类编码": {"column_letter": "Q", "label": "数字经济产业贷款类型（156小类代码-名称）"},
            }
        )
    else:
        base.update(
            {
                "佐证摘要": {"column_letter": "R", "label": "纳入养老产业贷款依据"},
                "产业大类编码": {"column_letter": "P", "label": "养老产业贷款类型（12大类代码-名称）"},
                "产业小类编码": {"column_letter": "Q", "label": "养老产业贷款类型（79小类代码-名称）"},
            }
        )
    return base


def _catalog_entry_value(entry):
    if isinstance(entry, dict):
        return entry.get("value")
    return entry


def _catalog_entry_label(field_name, entry):
    if isinstance(entry, dict):
        return str(entry.get("label") or field_name)
    return str(field_name)


def _resolve_text_rule_field_refs(detail_name, field_catalog, field_text, detail_param_lookup):
    resolved = []
    tokens = _split_rule_terms(field_text)
    detail_param_map = detail_param_lookup.get(detail_name, {})
    for token in tokens:
        if token in field_catalog:
            resolved.append((token, field_catalog[token]))
            continue
        if token in detail_param_map:
            entry = field_catalog.get(token)
            if entry is not None:
                resolved.append((token, entry))
                continue
        match = TEXT_RULE_COLUMN_PATTERN.match(token)
        if match:
            column_letter, label = match.groups()
            alias = f"列位:{column_letter}[{label}]" if label else f"列位:{column_letter}"
            entry = field_catalog.get(alias)
            if entry is None:
                entry = field_catalog.get(f"列位:{column_letter}")
            if entry is not None:
                resolved.append((alias, entry))
                continue
        legacy_spec = _legacy_text_field_specs(detail_name).get(token)
        if legacy_spec:
            alias = token
            entry = field_catalog.get(alias)
            if entry is not None:
                resolved.append((alias, entry))
    return resolved


def _split_keyword_terms(text):
    if text is None:
        return []
    return [item.strip() for item in re.split(r"[，,、]+", str(text)) if item and item.strip()]


def _split_keyword_groups(text):
    groups = []
    for group_text in _split_rule_terms(text):
        terms = _split_keyword_terms(group_text)
        if terms:
            groups.append(terms)
    return groups


def _split_regex_patterns(text):
    if text is None:
        return []
    parts = [item.strip() for item in re.split(r"[\r\n；;]+", str(text)) if item and item.strip()]
    return parts


def _text_regex_hits(text, patterns):
    hits = []
    for pattern in patterns:
        try:
            if pattern and re.search(pattern, text or ""):
                hits.append(pattern)
        except re.error:
            continue
    return dedup_keep_order(hits)


def _is_yes(value):
    return str(value or "").strip() in {"是", "Y", "YES", "1", "true", "True"}


def _industry_terms_from_entry(entry, level_text):
    if not entry:
        return []
    requested_levels = {item.strip() for item in re.split(r"[，,、]+", str(level_text or "小类,说明")) if item and item.strip()}
    level_map = {
        "大类": entry.get("大类名称", ""),
        "中类": entry.get("中类名称", ""),
        "小类": entry.get("小类名称", ""),
        "说明": entry.get("小类说明", ""),
    }
    terms = []
    for level in requested_levels:
        value = _clean_text_value(level_map.get(level, ""))
        if not value:
            continue
        terms.append(value)
        terms.extend([item for item in re.split(r"[，,、；;（）()、\s]+", value) if item and len(item.strip()) >= 2])
    return dedup_keep_order([item.strip() for item in terms if item and item.strip()])


def evaluate_keyword_rules_for_row(detail_name, field_catalog, rules, industry_hierarchy_map):
    hit_rule_names = []
    hit_notes = []
    suspicious = False
    actual_industry_code = extract_industry4(_catalog_entry_value(field_catalog.get("贷款实际投向行业")))
    customer_industry_code = extract_industry4(_catalog_entry_value(field_catalog.get("贷款客户行业分类")))

    for rule in rules:
        if str(rule.get("明细名称") or "").strip() not in {"", detail_name, "全部"}:
            continue
        field_name, entry = _resolve_keyword_rule_entry(rule, field_catalog)
        if entry is None:
            continue
        text = _clean_text_value(_catalog_entry_value(entry))
        if not text:
            continue

        effective_modes = [m.strip().lower() for m in re.split(r"[，,、\s]+", str(rule.get("匹配方式") or "").strip()) if m.strip()]
        if not effective_modes:
            effective_modes = ["contains"]
        if not effective_modes:
            continue

        exclude_hits = _text_contains_any(text, _split_keyword_terms(rule.get("排除关键词")))
        if exclude_hits:
            continue

        group_specs = _split_keyword_groups(rule.get("同时命中关键词组"))
        group_hits = []
        for group in group_specs:
            hits = _text_contains_any(text, group)
            if hits:
                group_hits.append(hits)
        required_group_count = parse_int(rule.get("至少命中组数"), 0)
        if group_specs and required_group_count <= 0:
            required_group_count = len(group_specs)

        matched = False
        mode_summaries = []
        for mode in effective_modes:
            if mode == "contains":
                include_hits = _text_contains_any(text, _split_keyword_terms(rule.get("包含关键词")))
                include_ok = not _split_keyword_terms(rule.get("包含关键词")) or bool(include_hits)
                groups_ok = not group_specs or len(group_hits) >= required_group_count
                if include_ok and groups_ok:
                    matched = True
                    if include_hits:
                        mode_summaries.append(f"{field_name}命中包含词:{','.join(include_hits)}")
                    if group_hits:
                        flat_group_hits = ["+".join(hits) for hits in group_hits]
                        mode_summaries.append(f"{field_name}命中组合词:{'；'.join(flat_group_hits)}")
            elif mode == "regex":
                regex_hits = _text_regex_hits(text, _split_regex_patterns(rule.get("包含关键词")))
                groups_ok = not group_specs or len(group_hits) >= required_group_count
                if regex_hits and groups_ok:
                    matched = True
                    mode_summaries.append(f"{field_name}命中正则:{'；'.join(regex_hits)}")
                    if group_hits:
                        flat_group_hits = ["+".join(hits) for hits in group_hits]
                        mode_summaries.append(f"{field_name}命中组合词:{'；'.join(flat_group_hits)}")
        if not matched:
            continue

        industry_notes = []
        level_text = rule.get("行业比对层级")
        if _is_yes(rule.get("匹配投向行业")) and actual_industry_code:
            industry_entry = industry_hierarchy_map.get(actual_industry_code) or industry_hierarchy_map.get(actual_industry_code[1:])
            industry_hits = _text_contains_any(text, _industry_terms_from_entry(industry_entry, level_text))
            if industry_hits:
                industry_notes.append(f"投向行业重复词:{','.join(industry_hits)}")
        if _is_yes(rule.get("匹配主营行业")) and customer_industry_code:
            industry_entry = industry_hierarchy_map.get(customer_industry_code) or industry_hierarchy_map.get(customer_industry_code[1:])
            industry_hits = _text_contains_any(text, _industry_terms_from_entry(industry_entry, level_text))
            if industry_hits:
                industry_notes.append(f"主营行业重复词:{','.join(industry_hits)}")

        rule_name = str(rule.get("规则名称") or rule.get("序号") or "").strip()
        hit_rule_names.append(rule_name)
        configured_note = str(rule.get("命中说明") or "").strip()
        combined_note = "；".join([note for note in [configured_note] + mode_summaries + industry_notes if note])
        if combined_note:
            hit_notes.append(combined_note)
        if _is_yes(rule.get("是否疑似线索")):
            suspicious = True

    return {
        "是否命中": "是" if hit_rule_names else "",
        "规则名称": "、".join(dedup_keep_order(hit_rule_names)),
        "命中说明": "；".join(dedup_keep_order(hit_notes)),
        "是否疑似线索": "是" if suspicious else "",
    }


def build_result_headers(category_order, runtime_settings):
    enabled_types = [label for label in runtime_settings.get("enabled_result_types", RESULT_TYPE_ORDER) if label != KEYWORD_RESULT_LABEL]
    headers = []
    for basis in runtime_settings.get("enabled_bases", ["actual", "customer"]):
        basis_name = BASIS_NAME_MAP.get(basis, basis)
        for cat in category_order:
            c_name = category_display_name(cat)
            headers.append(f"{basis_name}-报送-{c_name}")
            headers.append(f"{basis_name}-匹配-{c_name}")
            for result_label in enabled_types:
                headers.append(f"{basis_name}-{result_label}-{c_name}")
            headers.append(f"{basis_name}-是否线索-{c_name}")
            headers.append(f"{basis_name}-是否疑似线索-{c_name}")
            headers.append(f"{basis_name}-备注-{c_name}")
    if keyword_detail_enabled(runtime_settings):
        headers.extend(["关键字-是否命中", "关键字-规则名称", "关键字-命中说明", "关键字-是否疑似线索"])
    headers.append("行业小类描述")
    return headers


def basis_result_types(runtime_settings):
    return [label for label in runtime_settings.get("enabled_result_types", RESULT_TYPE_ORDER) if label != KEYWORD_RESULT_LABEL]


def _merge_runtime_rows(existing_rows):
    runtime_map = {}
    runtime_notes = {}
    extra_rows = []
    for row in existing_rows[1:] if existing_rows else []:
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if not key:
            continue
        value = "" if len(row) < 2 or row[1] is None else str(row[1]).strip()
        note = "" if len(row) < 3 or row[2] is None else str(row[2]).strip()
        runtime_map[key] = value
        runtime_notes[key] = note
    merged = [RUNTIME_HEADERS]
    used_keys = set()
    for key, default_value, default_note in RUNTIME_DEFAULT_ROWS:
        merged.append([key, runtime_map.get(key, default_value), runtime_notes.get(key, default_note)])
        used_keys.add(key)
    for row in existing_rows[1:] if existing_rows else []:
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        if not key or key in used_keys:
            continue
        extra = list(row) + [""] * max(0, 3 - len(row))
        merged.append(extra[:3])
    return merged


def _summary_stats_bucket(summary_stats, sheet_name, code):
    key = (sheet_name, code)
    if key not in summary_stats:
        summary_stats[key] = {"amount": defaultdict(float), "count": defaultdict(int)}
    return summary_stats[key]


def _summary_group_key(group_name):
    return f"__GROUP__::{group_name}"


def append_template_summary_stat(summary_stats, sheet_name, code, stat_key, amount=0.0, count=0):
    if not sheet_name or not code or not stat_key:
        return
    bucket = _summary_stats_bucket(summary_stats, sheet_name, code)
    bucket["amount"][stat_key] += amount or 0.0
    bucket["count"][stat_key] += count or 0


def resolve_result_label_display(result, result_label):
    if result_label == result.get("主标签"):
        return result.get("主标签内容", "")
    return result.get(result_label, "")


def _runtime_summary_interest_labels(runtime_settings):
    labels = set(runtime_settings.get("enabled_result_types", []))
    for items in runtime_settings.get("template_row_labels", {}).values():
        labels.update(items or [])
    return labels


def classify_template_stat_keys(actual_label, customer_label, runtime_settings):
    template_labels = runtime_settings.get("template_row_labels", {})
    actual_label = actual_label or ""
    customer_label = customer_label or ""
    interested_labels = _runtime_summary_interest_labels(runtime_settings)
    stat_keys = set()
    if actual_label in set(template_labels.get("线索合计", [])):
        stat_keys.add("line")
    if actual_label in set(template_labels.get("按投向有误", [])):
        stat_keys.add("actual_error")
    if actual_label in set(template_labels.get("按投向疑似无误", [])):
        stat_keys.add("suspect_ok")
    actual_hit_any = actual_label in interested_labels
    if not actual_hit_any:
        if customer_label in set(template_labels.get("线索合计", [])):
            stat_keys.add("line")
        if customer_label in set(template_labels.get("按主营有误", [])):
            stat_keys.add("customer_error")
            stat_keys.add("customer_all")
        if customer_label in set(template_labels.get("按主营疑似无误", [])):
            stat_keys.add("suspect_ok")
            stat_keys.add("customer_all")
    return stat_keys


def classify_basis_template_stat_keys(basis, result_label, runtime_settings):
    template_labels = runtime_settings.get("template_row_labels", {})
    result_label = result_label or ""
    stat_keys = set()
    if not result_label:
        return stat_keys

    line_labels = set(template_labels.get("线索合计", []))
    actual_suspect_labels = set(template_labels.get("按投向疑似无误", []))
    customer_suspect_labels = set(template_labels.get("按主营疑似无误", []))
    line_labels -= actual_suspect_labels
    line_labels -= customer_suspect_labels

    if result_label in line_labels:
        stat_keys.add("line")

    if basis == "actual":
        if result_label in set(template_labels.get("按投向有误", [])):
            stat_keys.add("actual_error")
        if result_label in actual_suspect_labels:
            stat_keys.add("suspect_ok")
    elif basis == "customer":
        if result_label in set(template_labels.get("按主营有误", [])):
            stat_keys.add("customer_error")
            stat_keys.add("customer_all")
        if result_label in customer_suspect_labels:
            stat_keys.add("customer_all")
    return stat_keys


def should_count_keyword_summary(keyword_result, stat_keys):
    if str(keyword_result.get("是否疑似线索") or "").strip() != "是":
        return False
    return not bool(stat_keys)


def keyword_template_stat_keys(runtime_settings):
    stat_keys = set()
    template_labels = runtime_settings.get("template_row_labels", {})
    if KEYWORD_RESULT_LABEL in set(template_labels.get("关键字", [])) or "关键字命中" in set(template_labels.get("关键字", [])):
        stat_keys.add("keyword")
    if KEYWORD_RESULT_LABEL in set(template_labels.get("线索合计", [])) or "关键字命中" in set(template_labels.get("线索合计", [])):
        stat_keys.add("line")
    return stat_keys


def keyword_detail_enabled(runtime_settings):
    enabled_types = set(runtime_settings.get("enabled_result_types", []))
    return bool(enabled_types & KEYWORD_RESULT_ALIASES)


def keyword_summary_enabled(runtime_settings):
    template_labels = runtime_settings.get("template_row_labels", {})
    for items in template_labels.values():
        labels = set(items or [])
        if labels & KEYWORD_RESULT_ALIASES:
            return True
    return False


def _summary_stat_value(summary_stats, sheet_name, codes, stat_key, metric):
    total = 0
    for code, _ in codes:
        bucket = summary_stats.get((sheet_name, code), {})
        metric_bucket = bucket.get(metric, {})
        total += metric_bucket.get(stat_key, 0)
    return total


def _summary_group_stat_value(summary_stats, sheet_name, group_name, stat_key, metric):
    bucket = summary_stats.get((sheet_name, _summary_group_key(group_name)), {})
    metric_bucket = bucket.get(metric, {})
    return metric_bucket.get(stat_key, 0)


def _ratio_value(numerator, denominator):
    if not denominator:
        return None
    return round(numerator / denominator, 6)


def _template_total_rows():
    return sum(len(rows) for _, rows in SUMMARY_TEMPLATE_SECTIONS)


def _write_template_sheet(ws, sheet_name, spec, summary_stats, runtime_settings):
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    ws["A1"] = None
    ws["B1"] = "产业类型"
    ws["B2"] = "大类"
    ws["B3"] = "大类名称"

    col_idx = 3
    for group_name, codes in spec["groups"]:
        group_start = col_idx
        group_end = col_idx + len(codes)
        ws.merge_cells(start_row=1, start_column=group_start, end_row=1, end_column=group_end)
        ws.cell(1, group_start).value = group_name
        ws.merge_cells(start_row=2, start_column=group_start, end_row=3, end_column=group_start)
        ws.cell(2, group_start).value = "合计"
        for code, code_name in codes:
            col_idx += 1
            ws.cell(2, col_idx).value = code
            ws.cell(3, col_idx).value = code_name
        col_idx += 1

    row_offset = 4
    for section_index, (section_name, rows) in enumerate(SUMMARY_TEMPLATE_SECTIONS):
        section_start = row_offset
        section_end = row_offset + len(rows) - 1
        ws.merge_cells(start_row=section_start, start_column=1, end_row=section_end, end_column=1)
        ws.cell(section_start, 1).value = f"{section_name}\n（{runtime_settings['summary_unit']}）" if section_name == "金额" else section_name
        for row_label, stat_key in rows:
            ws.cell(row_offset, 2).value = row_label
            col_idx = 3
            for group_name, codes in spec["groups"]:
                if stat_key == "total_amount":
                    group_total = round(_summary_group_stat_value(summary_stats, sheet_name, group_name, "total_amount", "amount"), 4)
                elif stat_key == "total_count":
                    group_total = _summary_group_stat_value(summary_stats, sheet_name, group_name, "total_count", "count")
                elif section_name == "金额":
                    value = _summary_group_stat_value(summary_stats, sheet_name, group_name, stat_key, "amount")
                    group_total = round(value, 4) if value else 0
                elif section_name == "笔数":
                    group_total = _summary_group_stat_value(summary_stats, sheet_name, group_name, stat_key, "count")
                elif section_index == 1:
                    numerator = _summary_group_stat_value(summary_stats, sheet_name, group_name, stat_key, "amount")
                    denominator = _summary_group_stat_value(summary_stats, sheet_name, group_name, "total_amount", "amount")
                    group_total = _ratio_value(numerator, denominator)
                else:
                    numerator = _summary_group_stat_value(summary_stats, sheet_name, group_name, stat_key, "count")
                    denominator = _summary_group_stat_value(summary_stats, sheet_name, group_name, "total_count", "count")
                    group_total = _ratio_value(numerator, denominator)
                ws.cell(row_offset, col_idx).value = group_total
                for code, _ in codes:
                    col_idx += 1
                    if stat_key == "total_amount":
                        value = summary_stats.get((sheet_name, code), {}).get("amount", {}).get("total_amount", 0.0)
                        ws.cell(row_offset, col_idx).value = round(value, 4)
                    elif stat_key == "total_count":
                        ws.cell(row_offset, col_idx).value = summary_stats.get((sheet_name, code), {}).get("count", {}).get("total_count", 0)
                    elif section_name == "金额":
                        value = summary_stats.get((sheet_name, code), {}).get("amount", {}).get(stat_key, 0.0)
                        ws.cell(row_offset, col_idx).value = round(value, 4) if value else 0
                    elif section_name == "笔数":
                        ws.cell(row_offset, col_idx).value = summary_stats.get((sheet_name, code), {}).get("count", {}).get(stat_key, 0)
                    elif section_index == 1:
                        numerator = summary_stats.get((sheet_name, code), {}).get("amount", {}).get(stat_key, 0.0)
                        denominator = summary_stats.get((sheet_name, code), {}).get("amount", {}).get("total_amount", 0.0)
                        ws.cell(row_offset, col_idx).value = _ratio_value(numerator, denominator)
                    else:
                        numerator = summary_stats.get((sheet_name, code), {}).get("count", {}).get(stat_key, 0)
                        denominator = summary_stats.get((sheet_name, code), {}).get("count", {}).get("total_count", 0)
                        ws.cell(row_offset, col_idx).value = _ratio_value(numerator, denominator)
                col_idx += 1
            row_offset += 1

    for row in ws.iter_rows(min_row=1, max_row=3 + _template_total_rows()):
        for cell in row:
            cell.font = Font(name="Arial", bold=cell.row <= 3 or cell.column <= 2)


def write_template_summary_sheets(workbook, summary_stats, runtime_settings):
    for index, sheet_name in enumerate(SUMMARY_TEMPLATE_SPECS.keys()):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name, index)
        _write_template_sheet(ws, sheet_name, SUMMARY_TEMPLATE_SPECS[sheet_name], summary_stats, runtime_settings)


def detect_text_contradictions(detail_name, field_catalog, rule_rows, detail_param_lookup=None):
    detail_param_lookup = detail_param_lookup or {}
    hits = []
    for rule in rule_rows:
        if rule.get("规则类型") == "语义支持":
            continue
        if rule.get("适用明细") not in {detail_name, "全部", None, ""}:
            continue
        mode = str(rule.get("匹配模式") or "").strip()
        fields = _resolve_text_rule_field_refs(detail_name, field_catalog, rule.get("适用字段"), detail_param_lookup)
        keywords = _split_rule_terms(rule.get("关键词"))
        reverse_keywords = _split_rule_terms(rule.get("反向关键词"))
        matched = False
        field_summary = []
        if mode == "empty_or_short":
            threshold = int(str(rule.get("关键词") or "8").strip() or "8")
            for field_name, entry in fields:
                field_text = _clean_text_value(_catalog_entry_value(entry))
                if len(field_text) < threshold:
                    matched = True
                    field_summary.append(f"{_catalog_entry_label(field_name, entry)}长度{len(field_text)}")
        elif mode == "keyword_pair" and len(fields) >= 2:
            left_name, left_entry = fields[0]
            left_text = _clean_text_value(_catalog_entry_value(left_entry))
            right_text = _clean_text_value(
                " ".join(_clean_text_value(_catalog_entry_value(entry)) for _, entry in fields[1:])
            )
            left_hits = _text_contains_any(left_text, keywords)
            right_hits = _text_contains_any(right_text, reverse_keywords)
            if left_hits and right_hits:
                matched = True
                field_summary.append(f"{_catalog_entry_label(left_name, left_entry)}命中{','.join(left_hits)}")
                field_summary.append(
                    f"{'|'.join(_catalog_entry_label(name, entry) for name, entry in fields[1:])}命中{','.join(right_hits)}"
                )
        elif mode == "negative_keyword":
            merged = " ".join(_clean_text_value(_catalog_entry_value(entry)) for _, entry in fields)
            neg_hits = _text_contains_any(merged, reverse_keywords)
            if neg_hits:
                matched = True
                field_summary.append(
                    f"{'|'.join(_catalog_entry_label(name, entry) for name, entry in fields)}命中{','.join(neg_hits)}"
                )
        elif mode == "flag_requires_keywords" and len(fields) >= 2:
            flag_fields = fields[:-1]
            evidence_name, evidence_entry = fields[-1]
            has_flag = any(_is_truthy_text_flag(_catalog_entry_value(entry)) for _, entry in flag_fields)
            evidence_text = _clean_text_value(_catalog_entry_value(evidence_entry))
            positive_hits = _text_contains_any(evidence_text, keywords)
            if has_flag and not positive_hits:
                matched = True
                field_summary.append(f"{'|'.join(_catalog_entry_label(name, entry) for name, entry in flag_fields)}已标记")
                field_summary.append(f"{_catalog_entry_label(evidence_name, evidence_entry)}缺少关键词")
        if matched:
            hits.append(
                {
                    "规则编号": str(rule.get("规则编号") or ""),
                    "规则类型": str(rule.get("规则类型") or ""),
                    "命中说明": str(rule.get("命中说明") or ""),
                    "风险等级": str(rule.get("风险等级") or "medium").lower(),
                    "命中字段": "；".join(field_summary),
                }
            )
    return hits


def score_text_semantics(target_detail, field_catalog, semantic_rules, stopwords, detail_param_lookup=None):
    detail_param_lookup = detail_param_lookup or {}
    all_details = dedup_keep_order(
        [str(rule.get("适用明细")).strip() for rule in semantic_rules if rule.get("适用明细")]
    )
    scores = {
        detail: {"support_score": 0.0, "positive_hits": [], "negative_hits": []}
        for detail in all_details
    }
    if target_detail not in scores:
        scores[target_detail] = {"support_score": 0.0, "positive_hits": [], "negative_hits": []}
    for rule in semantic_rules:
        if rule.get("规则类型") not in {"语义支持", None, ""}:
            continue
        detail = str(rule.get("适用明细") or "").strip()
        if not detail:
            continue
        fields = _resolve_text_rule_field_refs(target_detail, field_catalog, rule.get("适用字段"), detail_param_lookup)
        merged = " ".join(_clean_text_value(_catalog_entry_value(entry), stopwords) for _, entry in fields)
        if not merged:
            continue
        weight = float(rule.get("字段权重") or 1.0)
        pos_hits = _text_contains_any(merged, _split_rule_terms(rule.get("关键词")))
        neg_hits = _text_contains_any(merged, _split_rule_terms(rule.get("反向关键词")))
        bucket = scores.setdefault(detail, {"support_score": 0.0, "positive_hits": [], "negative_hits": []})
        bucket["support_score"] += len(pos_hits) * weight * 10
        bucket["support_score"] -= len(neg_hits) * weight * 8
        bucket["positive_hits"].extend(pos_hits)
        bucket["negative_hits"].extend(neg_hits)
    for bucket in scores.values():
        bucket["support_score"] = max(0.0, min(100.0, round(bucket["support_score"], 2)))
        bucket["positive_hits"] = dedup_keep_order(bucket["positive_hits"])
        bucket["negative_hits"] = dedup_keep_order(bucket["negative_hits"])
    return scores


def build_text_verdict(target_detail, semantic_scores, contradiction_hits, verdict_rules):
    conflict_rank = {"none": 0, "medium": 1, "high": 2}
    conflict_level = "none"
    if contradiction_hits:
        conflict_level = max(
            (str(item.get("风险等级") or "medium").lower() for item in contradiction_hits),
            key=lambda x: conflict_rank.get(x, 1),
        )
    detail_score = semantic_scores.get(target_detail, {}).get("support_score", 0.0)
    matching_rules = [
        rule
        for rule in verdict_rules
        if str(rule.get("冲突等级") or "").lower() == conflict_level
    ]
    chosen = None
    for rule in matching_rules:
        min_score = float(rule.get("最小分值") or 0)
        max_score = float(rule.get("最大分值") or 100)
        if min_score <= detail_score <= max_score:
            chosen = rule
            break
    if chosen is None and matching_rules:
        chosen = matching_rules[0]
    if chosen is None:
        chosen = {"结论": "建议复核", "复核建议": "补充结构化字段和佐证材料"}
    top_detail = max(semantic_scores.items(), key=lambda item: item[1].get("support_score", 0.0))[0] if semantic_scores else target_detail
    reasons = []
    if contradiction_hits:
        reasons.extend(item["命中说明"] for item in contradiction_hits)
    else:
        reasons.append(f"文本支持度 {detail_score}")
    return {
        "结论": str(chosen.get("结论") or "建议复核"),
        "复核建议": str(chosen.get("复核建议") or ""),
        "冲突等级": conflict_level,
        "目标产业支持度": detail_score,
        "文本支持产业": top_detail,
        "结论原因": "；".join(dedup_keep_order(reasons)),
    }


def parse_amount(value):
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        text = str(value).replace(",", "").strip()
        try:
            return float(text)
        except Exception:
            return 0.0


def resolve_match_map(industry4, mapping_data, industry_desc_map):
    match_map = mapping_data.get(industry4, {})
    if (
        not match_map
        and industry4
        and len(industry4) == 5
        and industry4[0].isalpha()
        and industry4[1:].isdigit()
    ):
        desc_entry = industry_desc_map.get(industry4[1:])
        if isinstance(desc_entry, tuple) and desc_entry[0]:
            canonical = desc_entry[0].upper().replace("*", "")
            if canonical != industry4:
                match_map = mapping_data.get(canonical, match_map)
    return match_map


def evaluate_basis_result(industry4, reported_codes, match_map, cat, clue_rules):
    def _star(mc):
        v = match_map[mc]
        return v[0] if isinstance(v, tuple) else bool(v)

    def _raw_list(mc):
        v = match_map[mc]
        if not (isinstance(v, tuple) and len(v) > 1):
            return []
        raw_v = v[1]
        if isinstance(raw_v, list):
            return [x for x in raw_v if str(x).strip()]
        if isinstance(raw_v, str) and raw_v.strip():
            return [raw_v.strip()]
        return []

    def _star_display_parts(mc):
        raws = _raw_list(mc)
        if raws:
            return [f"【{mc}*:{raw}】" for raw in raws]
        return [f"【{mc}*】"]

    def _nonstar_display_parts(mc):
        raws = _raw_list(mc)
        if raws:
            return [f"【{mc}:{raw}】" for raw in raws]
        return [f"【{mc}】"]

    rpt = dedup_keep_order(reported_codes)
    match_codes = [code for code in match_map.keys() if code.startswith(cat)]
    nonstar_codes = [c for c in match_codes if not _star(c)]
    star_codes = [c for c in match_codes if _star(c)]
    match_display = []
    for code in match_codes:
        match_display.extend(_star_display_parts(code) if _star(code) else _nonstar_display_parts(code))
    reported_set = set(rpt)
    unmatched_reported = [c for c in rpt if c not in match_codes]
    matched_star_reported = [c for c in rpt if c in star_codes]
    missing_nonstar = [c for c in nonstar_codes if c not in reported_set]
    missing_star = [c for c in star_codes if c not in reported_set]

    multi = "、".join(f"【{c}】" for c in unmatched_reported) if unmatched_reported else ""
    suspect_multi_parts = []
    for code in matched_star_reported:
        suspect_multi_parts.extend(_star_display_parts(code))
    suspect_multi = "、".join(suspect_multi_parts)
    miss_parts = []
    for code in missing_nonstar:
        miss_parts.extend(_nonstar_display_parts(code))
    miss = "、".join(miss_parts)
    suspect_miss_parts = []
    for code in missing_star:
        suspect_miss_parts.extend(_star_display_parts(code))
    suspect_miss = "、".join(suspect_miss_parts)

    n_cand = len(match_codes)
    has_star_cat = len(reported_set & set(star_codes)) > 0
    reported_empty_cat = len(rpt) == 0
    hit_non_empty_cat = len(reported_set & set(match_codes)) > 0
    extra_non_empty_cat = len(reported_set - set(match_codes)) > 0
    main_mark, extra_marks, clue_mark, suspect_clue_mark, triggered = apply_clue_rules(
        n_cand, has_star_cat, reported_empty_cat, hit_non_empty_cat, extra_non_empty_cat, clue_rules
    )
    if reported_empty_cat:
        main_mark_display = suspect_miss if has_star_cat else miss
    elif hit_non_empty_cat and not extra_non_empty_cat:
        main_mark_display = "、".join(match_display)
    else:
        main_mark_display = "；".join(s for s in (multi, suspect_multi) if s)
    suspect_right = main_mark == "疑似正确" or "疑似正确" in extra_marks
    if clue_mark in ("是", "否"):
        is_clue = clue_mark == "是"
    else:
        is_clue = main_mark in ("漏报", "多报", "错报") or any(m in ("漏报", "多报", "错报") for m in extra_marks)
    if suspect_clue_mark in ("是", "否"):
        is_suspect_clue = suspect_clue_mark == "是"
    else:
        is_suspect_clue = main_mark in ("疑似漏报", "疑似多报", "疑似错报", "疑似正确") or any(
            m in ("疑似漏报", "疑似多报", "疑似错报", "疑似正确") for m in extra_marks
        )

    note_text = "；".join("触发规则编号 {}：{}".format(rid, note) for rid, note in triggered) if triggered else ""
    return {
        "reported": "、".join(rpt),
        "matched": "、".join(match_display),
        "主标签": main_mark,
        "主标签内容": main_mark_display,
        "副标签": extra_marks,
        "疑似正确": "、".join(match_display) if suspect_right else "",
        "多报": multi,
        "疑似多报": suspect_multi,
        "错报": "；".join(s for s in (multi, suspect_multi) if s),
        "漏报": miss,
        "疑似漏报": suspect_miss,
        "是否线索": "是" if is_clue else "",
        "是否疑似线索": "是" if is_suspect_clue else "",
        "备注": note_text,
    }


def append_summary_stat(summary_stats, sheet_name, basis_name, category_name, result_type, amount):
    key = (sheet_name, basis_name, category_name, result_type)
    bucket = summary_stats.setdefault(key, {"count": 0, "amount": 0.0})
    bucket["count"] += 1
    bucket["amount"] += amount


def find_mapping_file():
    # 1. 优先从config.xlsx的mapping面板读取用户指定路径
    cfg_path = get_mapping_file_from_config()
    if cfg_path:
        return cfg_path

    # 2. 否则在根目录自动识别（工作表最多的xlsx）
    candidate = None
    max_sheets = -1
    for name in os.listdir(ROOT):
        if not name.lower().endswith(".xlsx"):
            continue
        if name.startswith("~$"):
            continue
        if name in {os.path.basename(OUTPUT_FILE), os.path.basename(CONFIG_FILE)}:
            continue
        path = os.path.join(ROOT, name)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()
        if sheet_count > max_sheets:
            max_sheets = sheet_count
            candidate = path
    if not candidate:
        raise FileNotFoundError("未找到匹配表文件。")
    return candidate


def find_ledger_files():
    files = []
    for dp, _, fns in os.walk(ROOT):
        for fn in fns:
            if not fn.lower().endswith(".xlsx"):
                continue
            if fn.startswith("~$"):
                continue
            if fn in {os.path.basename(OUTPUT_FILE), os.path.basename(CONFIG_FILE)}:
                continue
            full = os.path.join(dp, fn)
            rel = os.path.relpath(full, ROOT)
            if rel.count("\\") == 0:
                continue
            if fn.lower() == "22.xlsx":
                continue
            wb = openpyxl.load_workbook(full, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            mc, mr = ws.max_column, ws.max_row
            wb.close()
            if mc in (20, 21, 29) and mr >= 90:
                files.append(full)
    return sorted(files)


def detect_header_row(ws):
    best_row = 1
    best_count = -1
    for r in range(1, min(15, ws.max_row) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        cnt = sum(v is not None and str(v).strip() != "" for v in vals)
        if cnt > best_count:
            best_count = cnt
            best_row = r
    return best_row


def classify_file_type(ws):
    mc = ws.max_column
    header_probe_rows = range(1, min(ws.max_row, 5) + 1)
    text_samples = []
    for r in header_probe_rows:
        row_values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        text_samples.extend(str(v).strip() for v in row_values if v not in (None, ""))
    joined = " ".join(text_samples)
    if any(key in joined for key in ["高技术制造业", "高技术服务业", "战略性新兴产业", "知识产权（专利）密集型产业", "科技产业贷款明细"]):
        return "tech", [17, 20, 23, 26], 15
    if any(key in joined for key in ["数字经济产业贷款类型", "数字经济产业贷款明细", "数字产品制造业", "数字化效率提升业"]):
        return "digital", [16], 15
    if any(key in joined for key in ["养老产业贷款类型", "养老产业贷款明细", "养老照护服务", "养老设施建设"]):
        return "elder", [16], 15
    if mc >= 28:
        return "tech", [17, 20, 23, 26], 15
    if mc == 21:
        return "elder", [16], 15
    if mc == 20:
        return "digital", [16], 15
    raise ValueError(f"无法识别台账类型，工作表={ws.title}，列数={mc}")


def copy_sheet(src_ws, dst_ws):
    for r_idx, row in enumerate(src_ws.iter_rows(values_only=True), start=1):
        for c_idx, v in enumerate(row, start=1):
            dst_ws.cell(r_idx, c_idx).value = v


def ensure_unique_sheet_name(name, used):
    name = (name or "Sheet").strip()[:31] or "Sheet"
    if name not in used:
        used.add(name)
        return name
    base = name[:25]
    i = 2
    while f"{base}_{i}" in used:
        i += 1
    out = f"{base}_{i}"
    used.add(out)
    return out


def _default_col_category_map(file_type, reported_cols):
    if file_type == "tech":
        defaults = ["HTP", "HTS", "SE", "PA"]
        return [(col, defaults[i] if i < len(defaults) else "HTP") for i, col in enumerate(reported_cols)]
    if file_type == "digital":
        return [(col, "DE") for col in reported_cols]
    if file_type == "elder":
        return [(col, "EC") for col in reported_cols]
    return [(col, "ALL") for col in reported_cols]


def category_display_name(cat):
    return CATEGORY_NAME_MAP.get(cat.upper(), cat.upper())


# 已知主/副标签名，用于从备注解析副标签
CLUE_MARK_NAMES = ("正确", "疑似正确", "漏报", "疑似漏报", "多报", "错报", "疑似错报")


def _parse_sub_marks(主标签, 备注):
    """从 备注/副标签 列解析出副标签列表（不含主标签，去重）。"""
    if not 备注:
        return []
    s = str(备注).strip()
    out = []
    for name in CLUE_MARK_NAMES:
        if name != 主标签 and name in s:
            out.append(name)
    return out


def load_clue_rules():
    """
    从 config.xlsx 的 clue 工作表读取规则表。
    返回 list[dict]，每项: n_match, star, empty, hit, need_extra(从 Excel 加载时恒为 False), 主标签, 副标签(从 Excel 加载时恒为 [])。
    第 9 列「备注/副标签」仅供用户说明，不参与逻辑判断。
    """
    if not os.path.exists(CONFIG_FILE):
        return []
    try:
        mtime_ns = os.stat(CONFIG_FILE).st_mtime_ns
    except OSError:
        mtime_ns = None
    cached_rules = _CLUE_RULES_CACHE.get("rules")
    if cached_rules is not None and _CLUE_RULES_CACHE.get("mtime_ns") == mtime_ns:
        return [dict(rule) for rule in cached_rules]
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, "clue")
    if ws is None:
        wb.close()
        return []
    # 表头顺序(9列): 编号(1), 匹配数量(2), 含星(3), 报送空(4), 命中(5), 主标签(6), 是否线索(7), 是否疑似线索(8), 备注/副标签(9，仅用户说明、不参与逻辑，但会写入结果表「备注-产业」列)
    key_n, key_star, key_empty, key_hit, key_main, key_clue, key_suspect, key_note = 1, 2, 3, 4, 5, 6, 7, 8
    rules = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        n_val = row[key_n] if key_n < len(row) else None
        star_val = row[key_star] if key_star < len(row) else None
        empty_val = row[key_empty] if key_empty < len(row) else None
        hit_val = row[key_hit] if key_hit < len(row) else None
        main_val = row[key_main] if key_main < len(row) else None
        clue_val = row[key_clue] if key_clue < len(row) else None
        suspect_val = row[key_suspect] if key_suspect < len(row) else None
        note_val = row[key_note] if key_note < len(row) else None
        if main_val is None or str(main_val).strip() == "":
            continue
        n_str = str(n_val).strip() if n_val is not None else ""
        star_str = str(star_val).strip() if star_val is not None else "-"
        empty_str = str(empty_val).strip() if empty_val is not None else "-"
        hit_str = str(hit_val).strip() if hit_val is not None else "-"
        main_str = str(main_val).strip()
        is_clue_str = str(clue_val).strip() if clue_val is not None else ""
        is_suspect_str = str(suspect_val).strip() if suspect_val is not None else ""
        note_str = str(note_val).strip() if note_val else ""
        # 备注/副标签(I列)不参与逻辑：不解析副标签，不根据「混报」设 need_extra；仅用于结果表「备注-产业」列展示
        need_extra = False
        sub_list = []
        rules.append({
            "n_match": n_str,
            "star": star_str,
            "empty": empty_str,
            "hit": hit_str,
            "need_extra": need_extra,
            "主标签": main_str,
            "副标签": sub_list,
            "是否线索": is_clue_str,
            "是否疑似线索": is_suspect_str,
            "备注": note_str,
        })
    wb.close()
    if rules:
        _CLUE_RULES_CACHE["mtime_ns"] = mtime_ns
        _CLUE_RULES_CACHE["rules"] = [dict(rule) for rule in rules]
        print("已从 config 加载 clue 规则：{} 条（文件：{}）".format(len(rules), CONFIG_FILE))
    else:
        print("已从 config 加载 clue 规则：0 条，使用内置默认规则（请检查 clue 表第6列主标签是否有内容、列顺序是否为 9 列）")
        rules = _default_clue_rules()
        _CLUE_RULES_CACHE["mtime_ns"] = mtime_ns
        _CLUE_RULES_CACHE["rules"] = [dict(rule) for rule in rules]
    return rules


def _default_clue_rules():
    """无 clue 表或解析失败时返回内置默认规则（由 CLUE_DEFAULT_ROWS 生成，与 clue 表实际数据一致）。"""
    rules = []
    for row in CLUE_DEFAULT_ROWS:
        # row: [编号, 匹配数量, 行业是否含星, 机构报送产业为空?, 命中情况, 主标签, 是否线索, 是否疑似线索, 备注]
        if len(row) < 9:
            continue
        rules.append({
            "n_match": str(row[1]),
            "star": str(row[2]) if row[2] is not None else "-",
            "empty": str(row[3]) if row[3] is not None else "-",
            "hit": str(row[4]) if row[4] is not None else "-",
            "need_extra": False,
            "主标签": str(row[5]).strip(),
            "副标签": [],
            "是否线索": str(row[6]).strip() if row[6] is not None else "",
            "是否疑似线索": str(row[7]).strip() if row[7] is not None else "",
            "备注": str(row[8]).strip() if row[8] is not None else "",
        })
    return rules


def apply_clue_rules(n_candidates, has_star, reported_empty, hit_non_empty, extra_non_empty, rules):
    """
    按顺序匹配规则：主标签/是否线索等取第一条命中；若有多条命中则备注拼接展示。
    返回 (主标签, 副标签list, 是否线索, 是否疑似线索, 触发列表)。
    触发列表 = [(规则编号1-based, 备注), ...]，所有命中的规则都会列入，便于在备注列拼接「触发规则编号 1：备注1；触发规则编号 2：备注2」。
    """
    n_str = "0" if n_candidates == 0 else ("1" if n_candidates == 1 else ">1")
    empty_str = "是" if reported_empty else "否"
    hit_str = "是" if hit_non_empty else "否"
    star_ok = lambda rule: rule["star"] in ("-", "是" if has_star else "否", "否/是")
    触发列表 = []
    主标签, 副标签, 是否线索, 是否疑似线索 = "", [], "", ""
    for idx, r in enumerate(rules):
        if r["n_match"] != n_str:
            continue
        if r["empty"] != "-" and r["empty"] != empty_str:
            continue
        if r["hit"] != "-" and r["hit"] != hit_str:
            continue
        if not star_ok(r):
            continue
        if r["need_extra"] and not extra_non_empty:
            continue
        if r["need_extra"] is False and extra_non_empty and r["hit"] == "是" and r["empty"] == "否" and r["n_match"] == ">1":
            continue
        备注 = (r.get("备注") or "").strip()
        触发列表.append((idx + 1, 备注))
        if len(触发列表) == 1:
            是否线索 = (r.get("是否线索") or "").strip()
            是否疑似线索 = (r.get("是否疑似线索") or "").strip()
            主标签, 副标签 = r["主标签"], list(r["副标签"])
    return (主标签, 副标签, 是否线索, 是否疑似线索, 触发列表)


def _create_empty_config_with_schema():
    """新建 config.xlsx：含 config 表头、mapping 全表（A1/B1、A2/B2）、log 表头。"""
    wb = Workbook()
    ws = wb.active
    ws.title = PREFERRED_SHEET_NAMES["config"]
    ws.append(CONFIG_HEADERS)
    for c in range(1, len(CONFIG_HEADERS) + 1):
        ws.cell(1, c).font = Font(name="Arial", bold=True)

    mapping_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["mapping"])
    mapping_ws.append(["参照表路径", ""])
    mapping_ws.append(["映射表路径", ""])

    LOG_HEADERS = [
        "运行时间", "参照表路径", "参照表sheet数量", "参照表sheet列表",
        "结果文件", "明细sheet名称", "来源台账文件", "台账类型",
        "使用参照表sheet序号", "使用参照表sheet名称",
        "多报数量", "漏报数量", "疑似多报数量", "疑似漏报数量",
    ]
    log_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["log"])
    log_ws.append(LOG_HEADERS)
    for c in range(1, len(LOG_HEADERS) + 1):
        log_ws.cell(1, c).font = Font(name="Arial", bold=True)

    clue_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["clue"])
    clue_ws.append(CLUE_HEADERS)
    for c in range(1, len(CLUE_HEADERS) + 1):
        clue_ws.cell(1, c).font = Font(name="Arial", bold=True)
    for row in CLUE_DEFAULT_ROWS:
        clue_ws.append(row)

    runtime_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["runtime"])
    for row in _merge_runtime_rows([]):
        runtime_ws.append(row)
    for c in range(1, len(RUNTIME_HEADERS) + 1):
        runtime_ws.cell(1, c).font = Font(name="Arial", bold=True)

    wb.save(CONFIG_FILE)
    wb.close()


def init_config_file_if_missing():
    if os.path.exists(CONFIG_FILE):
        return

    print("未找到 config.xlsx（程序根目录：{}）".format(ROOT))
    try:
        choice = input("是否新建 config 并初始化？(y/n)：").strip().lower()
    except EOFError:
        choice = "n"
    if choice not in ("y", "yes"):
        raise FileNotFoundError("未找到 config.xlsx。请先新建并初始化，或放入程序根目录。")

    # 先创建空 config（config 表头 + mapping 全表 + log 表头）
    _create_empty_config_with_schema()

    ledger_files = find_ledger_files()
    if not ledger_files:
        print("已创建空的 config.xlsx，请手动在 config 表填写台账配置，或在 mapping 表 B1/B2 填写参照表、映射表路径。")
        return

    # 根据台账文件补全 config 表数据行
    wb = openpyxl.load_workbook(CONFIG_FILE)
    ws = get_sheet_by_logical_name(wb, "config")
    used_names = set()
    for f in ledger_files:
        src_wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        src_ws = src_wb[src_wb.sheetnames[0]]
        header_row = detect_header_row(src_ws)
        file_type, reported_cols, industry_col = classify_file_type(src_ws)
        output_sheet = ensure_unique_sheet_name(src_ws.title, used_names)
        ws.append(
            [
                output_sheet,
                os.path.relpath(f, ROOT),
                TYPE_KEY_TO_LABEL[file_type],
                header_row,
                header_row + 1,
                6,
                industry_col,
                8,
                ",".join(str(x) for x in reported_cols),
                ",".join(f"{x}:{infer}" for x, infer in _default_col_category_map(file_type, reported_cols)),
                {"digital": 2, "elder": 3, "tech": 4}[file_type],
                2,
                13,
                19,
                21,
            ]
        )
        src_wb.close()
    wb.save(CONFIG_FILE)
    wb.close()


def _ensure_mapping_sheet_layout():
    """若 mapping 表为旧格式（B1 为「国民经济行业分类映射表」），则迁移为 A1/B1、A2/B2 新格式。"""
    if not os.path.exists(CONFIG_FILE):
        return
    wb = openpyxl.load_workbook(CONFIG_FILE)
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        wb.close()
        return
    if ws.max_row < 2:
        wb.close()
        return
    b1_val = str(ws.cell(1, 2).value or "").strip()
    if b1_val != "国民经济行业分类映射表":
        wb.close()
        return
    # 旧格式：A2=参照表路径, B2=映射表路径 → 新格式：B1=参照表路径, B2=映射表路径，A1/A2 为表头
    ref_path = str(ws.cell(2, 1).value or "").strip()
    map_path = str(ws.cell(2, 2).value or "").strip()
    ws.cell(1, 1).value = "参照表路径"
    ws.cell(1, 2).value = ref_path
    ws.cell(2, 1).value = "映射表路径"
    ws.cell(2, 2).value = map_path
    wb.save(CONFIG_FILE)
    wb.close()


def ensure_config_schema():
    if not os.path.exists(CONFIG_FILE):
        return
    _ensure_mapping_sheet_layout()
    wb = openpyxl.load_workbook(CONFIG_FILE)
    config_ws = get_sheet_by_logical_name(wb, "config")
    current_headers = [config_ws.cell(1, i).value for i in range(1, config_ws.max_column + 1)] if config_ws is not None else []
    required_sheets = {"mapping", "log", "clue", "关键字映射", "runtime", "明细定义", "明细参数", "字段映射表"}
    runtime_ok = False
    keyword_mapping_ok = False
    detail_ok = False
    runtime_ws_existing = get_sheet_by_logical_name(wb, "runtime")
    if runtime_ws_existing is not None:
        runtime_keys = {
            str(runtime_ws_existing.cell(r, 1).value).strip()
            for r in range(2, runtime_ws_existing.max_row + 1)
            if runtime_ws_existing.cell(r, 1).value is not None
        }
        runtime_ok = all(key in runtime_keys for key, _, _ in RUNTIME_DEFAULT_ROWS)
    keyword_mapping_ws_existing = get_sheet_by_logical_name(wb, "关键字映射")
    if keyword_mapping_ws_existing is not None:
        keyword_mapping_headers = [keyword_mapping_ws_existing.cell(1, c).value for c in range(1, keyword_mapping_ws_existing.max_column + 1)]
        keyword_mapping_ok = keyword_mapping_headers[: len(KEYWORD_MAPPING_HEADERS)] == KEYWORD_MAPPING_HEADERS
    detail_ws_existing = get_sheet_by_logical_name(wb, "明细定义")
    if detail_ws_existing is not None:
        detail_headers = [detail_ws_existing.cell(1, c).value for c in range(1, detail_ws_existing.max_column + 1)]
        detail_ok = all(header in detail_headers for header in DETAIL_DEFINITION_CONFIG_HEADERS)
    has_required_sheets = all(get_sheet_by_logical_name(wb, logical_name) is not None for logical_name in required_sheets)
    if config_ws is None and has_required_sheets and runtime_ok and keyword_mapping_ok and detail_ok:
        wb.close()
        return

    normalized_rows = []
    if config_ws is not None:
        for row in config_ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v is None or str(v).strip() == "" for v in row):
                continue
            row = list(row)
            out_sheet = row[0] if len(row) > 0 else None
            source_file = row[1] if len(row) > 1 else None
            type_label = row[2] if len(row) > 2 else None
            header_row = row[3] if len(row) > 3 else None
            data_start = row[4] if len(row) > 4 else None
            customer_industry_col = row[5] if len(row) > 5 else None
            industry_col = row[6] if len(row) > 6 else None
            balance_col = row[7] if len(row) > 7 else None
            has_balance_unit_col = len(row) >= 16
            balance_unit = row[8] if has_balance_unit_col and len(row) > 8 else "万元"
            reported_cols = row[9] if has_balance_unit_col and len(row) > 9 else (row[8] if len(row) > 8 else None)
            mapping_idx = 10 if has_balance_unit_col else 9
            has_mapping_col = len(row) > mapping_idx and isinstance(row[mapping_idx], str) and ":" in row[mapping_idx]
            if has_mapping_col:
                col_cat_map_text = row[mapping_idx]
                match_sheet = row[mapping_idx + 1] if len(row) > mapping_idx + 1 else None
                class_col = row[mapping_idx + 2] if len(row) > mapping_idx + 2 else None
                ind_col = row[mapping_idx + 3] if len(row) > mapping_idx + 3 else None
                star_col = row[mapping_idx + 4] if len(row) > mapping_idx + 4 else None
                raw_map_col = row[mapping_idx + 5] if len(row) > mapping_idx + 5 else 21
            else:
                type_key = TYPE_LABEL_TO_KEY.get(str(type_label).strip(), "")
                rep_cols = parse_idx_list(reported_cols)
                col_cat_map_text = ",".join(f"{c}:{t}" for c, t in _default_col_category_map(type_key, rep_cols))
                match_sheet = row[mapping_idx + 1] if len(row) > mapping_idx + 1 else None
                class_col = row[mapping_idx + 2] if len(row) > mapping_idx + 2 else None
                ind_col = row[mapping_idx + 3] if len(row) > mapping_idx + 3 else None
                star_col = row[mapping_idx + 4] if len(row) > mapping_idx + 4 else None
                raw_map_col = row[mapping_idx + 5] if len(row) > mapping_idx + 5 else 21
            normalized_rows.append([
                out_sheet, source_file, type_label, header_row, data_start, customer_industry_col or 6,
                industry_col, balance_col or 8, balance_unit or "万元", reported_cols, col_cat_map_text,
                match_sheet, class_col, ind_col, star_col, raw_map_col,
            ])

    mapping_ws = get_sheet_by_logical_name(wb, "mapping")
    if mapping_ws is not None:
        mapping_rows = [
            [cell for cell in row]
            for row in mapping_ws.iter_rows(min_row=1, max_row=mapping_ws.max_row, values_only=True)
        ]
        # 迁移旧格式：原 A1/B1 为表头、A2/B2 为路径 → 现 A1/B1 为参照表、A2/B2 为映射表
        if len(mapping_rows) >= 2 and str(mapping_rows[0][1] or "").strip() == "国民经济行业分类映射表":
            ref_path = str(mapping_rows[1][0] or "").strip() if len(mapping_rows[1]) > 0 else ""
            map_path = str(mapping_rows[1][1] or "").strip() if len(mapping_rows[1]) > 1 else ""
            mapping_rows = [["参照表路径", ref_path], ["映射表路径", map_path]] + mapping_rows[2:]
    else:
        mapping_rows = [["参照表路径", ""], ["映射表路径", ""]]

    log_ws = get_sheet_by_logical_name(wb, "log")
    if log_ws is not None:
        log_rows = [
            [cell for cell in row]
            for row in log_ws.iter_rows(min_row=1, max_row=log_ws.max_row, values_only=True)
        ]
    else:
        log_rows = []

    runtime_rows = []
    runtime_ws_old = get_sheet_by_logical_name(wb, "runtime")
    if runtime_ws_old is not None:
        for row in runtime_ws_old.iter_rows(min_row=1, max_row=runtime_ws_old.max_row, values_only=True):
            runtime_rows.append([cell for cell in row])

    clue_rows = []
    clue_ws_old = get_sheet_by_logical_name(wb, "clue")
    if clue_ws_old is not None:
        for row in clue_ws_old.iter_rows(min_row=1, max_row=clue_ws_old.max_row, values_only=True):
            clue_rows.append([cell for cell in row])

    keyword_mapping_rows = _normalize_keyword_mapping_rows(
        _sheet_rows_or_defaults(wb, "关键字映射", KEYWORD_MAPPING_HEADERS, get_default_keyword_mapping_rows())
    )
    keyword_strategy_ws_old = get_sheet_by_logical_name(wb, "关键字策略")
    if keyword_strategy_ws_old is not None:
        strategy_rows = [
            [cell for cell in row]
            for row in keyword_strategy_ws_old.iter_rows(min_row=1, max_row=keyword_strategy_ws_old.max_row, values_only=True)
        ]
        strategy_header_map = {}
        if strategy_rows:
            strategy_header_map = {
                str(name or "").strip(): idx for idx, name in enumerate(strategy_rows[0]) if str(name or "").strip()
            }
        strategy_lookup = {}
        for row in strategy_rows[1:]:
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            detail_name = row[strategy_header_map["明细名称"]] if "明细名称" in strategy_header_map and strategy_header_map["明细名称"] < len(row) else None
            field_name = row[strategy_header_map["明细字段"]] if "明细字段" in strategy_header_map and strategy_header_map["明细字段"] < len(row) else None
            column_letter = row[strategy_header_map["字段所在列位置"]] if "字段所在列位置" in strategy_header_map and strategy_header_map["字段所在列位置"] < len(row) else None
            match_modes = row[strategy_header_map["匹配方式"]] if "匹配方式" in strategy_header_map and strategy_header_map["匹配方式"] < len(row) else None
            disabled = row[strategy_header_map["是否禁用"]] if "是否禁用" in strategy_header_map and strategy_header_map["是否禁用"] < len(row) else None
            if detail_name and field_name:
                strategy_lookup[(str(detail_name).strip(), str(field_name).strip())] = {
                    "字段所在列位置": str(column_letter or "").strip().upper(),
                    "匹配方式": str(match_modes or "").strip(),
                    "是否禁用": disabled,
                }
        migrated_keyword_rows = [keyword_mapping_rows[0]]
        for row in keyword_mapping_rows[1:]:
            if not row:
                continue
            merged_row = list(row) + [""] * max(0, len(KEYWORD_MAPPING_HEADERS) - len(row))
            key = (str(merged_row[2] or "").strip(), str(merged_row[3] or "").strip())
            strategy_spec = strategy_lookup.get(key)
            if strategy_spec:
                if len(merged_row) > 4 and not str(merged_row[4] or "").strip():
                    merged_row[4] = strategy_spec.get("字段所在列位置", "")
                if len(merged_row) > 5 and not str(merged_row[5] or "").strip():
                    merged_row[5] = strategy_spec.get("匹配方式", "")
                if len(merged_row) > 15 and merged_row[15] in (None, "") and strategy_spec.get("是否禁用") not in (None, ""):
                    merged_row[15] = strategy_spec.get("是否禁用")
            migrated_keyword_rows.append(merged_row[: len(KEYWORD_MAPPING_HEADERS)])
        keyword_mapping_rows = migrated_keyword_rows
    detail_rows = _sheet_rows_or_defaults(wb, "明细定义", DETAIL_DEFINITION_ALL_HEADERS, get_default_detail_definition_rows())
    detail_param_rows = _sheet_rows_or_defaults(wb, "明细参数", DETAIL_PARAM_HEADERS, get_default_detail_param_rows())
    field_mapping_rows = _sheet_rows_or_defaults(wb, "字段映射表", FIELD_MAPPING_HEADERS, get_default_field_mapping_rows())
    detail_rows[0] = _extend_headers_if_needed(detail_rows[0], DETAIL_DEFINITION_CONFIG_HEADERS)
    if normalized_rows:
        config_dicts = []
        for row in normalized_rows:
            config_dicts.append(
                {
                    "输出工作表名称": row[0],
                    "来源台账文件": row[1],
                    "台账类型": TYPE_LABEL_TO_KEY.get(str(row[2]).strip(), ""),
                    "表头行号": parse_int(row[3], 1),
                    "数据起始行号": parse_int(row[4], 2),
                    "贷款客户行业列序号": parse_int(row[5], 6),
                    "贷款投向行业列序号": parse_int(row[6], 15),
                    "贷款余额列序号": parse_int(row[7], 8),
                    "贷款余额原始单位": str(row[8] or "万元").strip() or "万元",
                    "机构报送产业分类列序号": parse_idx_list(row[9]),
                    "报送列-类别映射": dedup_pairs_keep_order(parse_col_category_map(row[10])),
                    "参照表工作表序号": parse_sheet_selector(row[11], 4),
                    "参照表产业分类代码列序号": parse_int(row[12], 2),
                    "参照表行业4位码列序号": parse_int(row[13], 13),
                    "参照表星标列序号": parse_int(row[14], 19),
                    "参照表原始映射列序号": parse_int(row[15], 21),
                }
            )
        detail_rows, detail_param_rows = _runtime_detail_rows_from_config_rows(config_dicts)
    elif not detail_ok:
        detail_rows[0] = _extend_headers_if_needed(detail_rows[0], DETAIL_DEFINITION_CONFIG_HEADERS)

    # 清空原工作簿，重建
    while wb.worksheets:
        wb.remove(wb.worksheets[0])

    mapping_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["mapping"])
    for row in mapping_rows:
        mapping_ws.append(row)

    log_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["log"])
    if log_rows:
        for row in log_rows:
            log_ws.append(row)
    else:
        log_headers = [
            "运行时间", "参照表路径", "参照表sheet数量", "参照表sheet列表",
            "结果文件", "明细sheet名称", "来源台账文件", "台账类型",
            "使用参照表sheet序号", "使用参照表sheet名称",
            "多报数量", "漏报数量", "疑似多报数量", "疑似漏报数量",
        ]
        log_ws.append(log_headers)

    clue_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["clue"])
    if clue_rows and len(clue_rows) >= 1:
        for row in clue_rows:
            clue_ws.append(row)
        for c in range(1, len(CLUE_HEADERS) + 1):
            clue_ws.cell(1, c).font = Font(name="Arial", bold=True)
    else:
        _append_sheet_with_headers(clue_ws, CLUE_HEADERS)
        for row in CLUE_DEFAULT_ROWS:
            clue_ws.append(row)

    keyword_mapping_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["关键字映射"])
    for row in keyword_mapping_rows:
        keyword_mapping_ws.append(row)
    for c in range(1, len(KEYWORD_MAPPING_HEADERS) + 1):
        keyword_mapping_ws.cell(1, c).font = Font(name="Arial", bold=True)

    runtime_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["runtime"])
    for row in _merge_runtime_rows(runtime_rows):
        runtime_ws.append(row)
    for c in range(1, len(RUNTIME_HEADERS) + 1):
        runtime_ws.cell(1, c).font = Font(name="Arial", bold=True)

    detail_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["明细定义"])
    for row in detail_rows:
        detail_ws.append(row)
    for c in range(1, len(detail_rows[0]) + 1):
        detail_ws.cell(1, c).font = Font(name="Arial", bold=True)

    detail_param_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["明细参数"])
    for row in detail_param_rows:
        detail_param_ws.append(row)
    for c in range(1, len(DETAIL_PARAM_HEADERS) + 1):
        detail_param_ws.cell(1, c).font = Font(name="Arial", bold=True)

    field_mapping_ws = wb.create_sheet(PREFERRED_SHEET_NAMES["字段映射表"])
    for row in field_mapping_rows:
        field_mapping_ws.append(row)
    for c in range(1, len(FIELD_MAPPING_HEADERS) + 1):
        field_mapping_ws.cell(1, c).font = Font(name="Arial", bold=True)

    wb.save(CONFIG_FILE)
    wb.close()


def get_mapping_file_from_config():
    """从 config.xlsx 的 mapping 工作表读取 B1，即参照表（五篇大文章与国民经济行业分类对应参照表）的文件路径。"""
    if not os.path.exists(CONFIG_FILE):
        return None
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        wb.close()
        return None
    path_cell = ws.cell(1, 2).value  # B1 = 参照表路径
    wb.close()
    if not path_cell:
        return None
    raw = str(path_cell).strip()
    if not raw:
        return None
    # 允许绝对或相对路径
    if os.path.isabs(raw):
        p = raw
    else:
        p = os.path.join(ROOT, raw)
    return p if os.path.exists(p) else None


def get_source_file_map_from_mapping():
    """
    从 config.xlsx 的 mapping 工作表中读取源文件 → 源文件路径映射。
    期望结构：
    A1: 参照表路径   B1: 参照表（五篇大文章…）文件路径
    A2: 映射表路径   B2: 映射表（国民经济行业分类映射表）文件路径
    空一行（可选）
    A4: 源文件   B4: 源文件路径
    A5..: 源文件名称（或相对路径） B5..: 源文件完整路径
    """
    if not os.path.exists(CONFIG_FILE):
        return {}
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        wb.close()
        return {}
    mapping = {}
    # 寻找“源文件”表头所在行
    header_row = None
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        if str(v1).strip() == "源文件" and str(v2).strip() == "源文件路径":
            header_row = r
            break
    if not header_row:
        wb.close()
        return {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        path = ws.cell(r, 2).value
        if not name or not path:
            continue
        name_s = str(name).strip()
        path_s = str(path).strip()
        if not name_s or not path_s:
            continue
        mapping[name_s] = path_s
    wb.close()
    return mapping


def resolve_source_workbook_path(source_key, source_file_map):
    if source_key is None or not str(source_key).strip():
        return None
    src_key = str(source_key).strip()
    if os.path.isabs(src_key):
        return src_key
    if src_key in source_file_map:
        return source_file_map[src_key]
    base_name = os.path.basename(src_key)
    for key, value in source_file_map.items():
        if os.path.basename(str(key).strip()) == base_name:
            return value
    return os.path.join(ROOT, src_key)


def _normalized_path_key(path):
    return os.path.normcase(os.path.abspath(str(path)))


def standardize_source_workbooks(config_rows, source_file_map):
    standardize_module = _load_standardize_module()
    standardized_map = {}
    processed = set()
    output_dir = Path(STANDARDIZE_OUTPUT_DIR)
    config_path = Path(CONFIG_FILE)
    output_dir.mkdir(parents=True, exist_ok=True)

    for cfg in config_rows:
        source_abs = resolve_source_workbook_path(cfg.get("来源台账文件"), source_file_map)
        if not source_abs or not os.path.exists(source_abs):
            continue
        path_key = _normalized_path_key(source_abs)
        if path_key in processed:
            continue
        processed.add(path_key)
        outputs = standardize_module.standardize_matching_workbook(Path(source_abs), config_path, output_dir)
        if outputs:
            standardized_map[path_key] = outputs
    return standardized_map


def prepare_standardized_sources_for_strategy(config_rows, source_file_map):
    temp_dir_obj = tempfile.TemporaryDirectory(prefix="strategy1_std_", ignore_cleanup_errors=True)
    original_output_dir = STANDARDIZE_OUTPUT_DIR
    try:
        globals()["STANDARDIZE_OUTPUT_DIR"] = temp_dir_obj.name
        standardized_map = standardize_source_workbooks(config_rows, source_file_map)
    finally:
        globals()["STANDARDIZE_OUTPUT_DIR"] = original_output_dir
    return standardized_map, temp_dir_obj


def standardize_only_from_mapping():
    init_config_file_if_missing()
    ensure_config_schema()
    source_file_map = get_source_file_map_from_mapping()
    config_rows = load_config_rows()
    if not config_rows and source_file_map:
        config_rows = build_config_rows_from_files(source_file_map.values())
    config_rows = reconcile_config_rows_with_mapping(config_rows, source_file_map)
    if not config_rows:
        raise ValueError("未能从config.xlsx或mapping中推断任何可标准化的台账配置行。")
    standardized_map = standardize_source_workbooks(config_rows, source_file_map)
    output_paths = []
    for entries in standardized_map.values():
        for item in entries:
            output_paths.append(str(item["output_path"]))
    return dedup_keep_order(output_paths)


def reconcile_config_rows_with_mapping(config_rows, source_file_map):
    if not source_file_map:
        return config_rows
    inferred_rows = build_config_rows_from_files(source_file_map.values())
    if not config_rows:
        return inferred_rows

    valid_rows = []
    covered_sources = set()
    for cfg in config_rows:
        source_abs = resolve_source_workbook_path(cfg.get("来源台账文件"), source_file_map)
        if source_abs and os.path.exists(source_abs):
            valid_rows.append(cfg)
            covered_sources.add(_normalized_path_key(source_abs))
    if not valid_rows:
        return inferred_rows or config_rows

    merged_rows = list(valid_rows)
    for cfg in inferred_rows:
        source_abs = resolve_source_workbook_path(cfg.get("来源台账文件"), source_file_map)
        if not source_abs:
            continue
        source_key = _normalized_path_key(source_abs)
        if source_key in covered_sources:
            continue
        merged_rows.append(cfg)
        covered_sources.add(source_key)
    return merged_rows


def resolve_standardized_workbook_path(source_abs, preferred_sheet_name, standardized_map):
    entries = standardized_map.get(_normalized_path_key(source_abs), [])
    if not entries:
        return source_abs
    for item in entries:
        if item.get("sheet_name") == preferred_sheet_name:
            return str(item["output_path"])
    if len(entries) == 1:
        return str(entries[0]["output_path"])
    return source_abs


def pick_source_worksheet(workbook, preferred_sheet_name):
    if preferred_sheet_name in workbook.sheetnames:
        return workbook[preferred_sheet_name]
    for sheet_name in workbook.sheetnames:
        if sheet_name not in STANDARDIZE_REPORT_SHEETS:
            return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def build_strategy_output_path():
    os.makedirs(STANDARDIZE_OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(STANDARDIZE_OUTPUT_DIR, f"策略一核查结果_{ts}.xlsx")


def build_column_validation_output_path():
    os.makedirs(STANDARDIZE_OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(STANDARDIZE_OUTPUT_DIR, f"配置取列校验_{ts}.xlsx")


# openpyxl 仅支持以下格式，.xls 等需用 xlrd 或先转为 xlsx
OPENPYXL_SUPPORTED_SUFFIXES = (".xlsx", ".xlsm", ".xltx", ".xltm")


def build_config_rows_from_files(file_paths):
    """
    根据给定的源文件路径列表，自动推断config行（台账类型、列序号等）。
    仅处理 openpyxl 支持的格式（.xlsx/.xlsm/.xltx/.xltm），跳过 .xls 等并提示。
    """
    rows = []
    for p in sorted(file_paths):
        if not p:
            continue
        path = str(p).strip()
        if not path:
            continue
        if not os.path.isabs(path):
            path_abs = os.path.join(ROOT, path)
        else:
            path_abs = path
        if not os.path.exists(path_abs):
            continue
        if not path_abs.lower().endswith(OPENPYXL_SUPPORTED_SUFFIXES):
            print("跳过（openpyxl 不支持该格式，仅支持 .xlsx/.xlsm/.xltx/.xltm）：{}".format(path_abs))
            continue
        try:
            wb = openpyxl.load_workbook(path_abs, read_only=True, data_only=True)
        except InvalidFileException:
            print("跳过（无法用 openpyxl 打开，请用 Excel 另存为 .xlsx 后重试）：{}".format(path_abs))
            continue
        matched_any = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name in STANDARDIZE_REPORT_SHEETS:
                continue
            header_row = detect_header_row(ws)
            try:
                file_type, reported_cols, industry_col = classify_file_type(ws)
            except ValueError:
                continue
            output_sheet = ws.title
            rows.append(
                {
                    "输出工作表名称": output_sheet,
                    "来源台账文件": os.path.relpath(path_abs, ROOT),
                    "台账类型": file_type,
                    "表头行号": header_row,
                    "数据起始行号": header_row + 1,
                    "贷款客户行业列序号": 6,
                    "贷款投向行业列序号": industry_col,
                    "贷款余额列序号": 8,
                    "贷款余额原始单位": "万元",
                    "机构报送产业分类列序号": reported_cols,
                    "报送列-类别映射": _default_col_category_map(file_type, reported_cols),
                    "参照表工作表序号": {"digital": 2, "elder": 3, "tech": 4}[file_type],
                    "参照表产业分类代码列序号": 2,
                    "参照表行业4位码列序号": 13,
                    "参照表星标列序号": 19,
                    "参照表原始映射列序号": 21,
                }
            )
            matched_any = True
        wb.close()
        if not matched_any:
            print(f"跳过（无法识别台账类型，请检查表头或在 config 中手工配置）：{path_abs}")
    return rows

def append_run_log(mapping_file, mapping_sheetnames, executed_entries, output_path):
    if not executed_entries or not os.path.exists(CONFIG_FILE):
        return
    wb = openpyxl.load_workbook(CONFIG_FILE)
    ws = get_sheet_by_logical_name(wb, "log")
    if ws is not None:
        start_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(PREFERRED_SHEET_NAMES["log"])
        ws.append(
            [
                "运行时间",
                "参照表路径",
                "参照表sheet数量",
                "参照表sheet列表",
                "结果文件",
                "明细sheet名称",
                "来源台账文件",
                "台账类型",
                "使用参照表sheet序号",
                "使用参照表sheet名称",
                "多报数量",
                "漏报数量",
                "疑似多报数量",
                "疑似漏报数量",
            ]
        )
        start_row = 2

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet_list_str = ",".join(mapping_sheetnames)
    for entry in executed_entries:
        sel = entry["参照表工作表序号"]
        if isinstance(sel, int) and 1 <= sel <= len(mapping_sheetnames):
            mname = mapping_sheetnames[sel - 1]
        elif isinstance(sel, str) and sel.strip():
            mname = sel.strip()
        else:
            mname = ""
        ws.append(
            [
                ts,
                mapping_file,
                len(mapping_sheetnames),
                sheet_list_str,
                output_path,
                entry["输出工作表名称"],
                entry["来源台账文件"],
                TYPE_KEY_TO_LABEL.get(entry["台账类型"], entry["台账类型"]),
                sel,
                mname,
                entry.get("多报数量", ""),
                entry.get("漏报数量", ""),
                entry.get("疑似多报数量", ""),
                entry.get("疑似漏报数量", ""),
            ]
        )

    wb.save(CONFIG_FILE)
    wb.close()


def load_config_rows():
    if not os.path.exists(CONFIG_FILE):
        raise FileNotFoundError("未找到config.xlsx。")
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    rows = _extract_runtime_rows_from_detail_sheets(wb)
    if not rows:
        source_map = get_source_file_map_from_mapping()
        inferred_rows = build_config_rows_from_files(source_map.values()) if source_map else []
        inferred_by_type = {row["台账类型"]: row for row in inferred_rows}

        detail_defs = _load_named_sheet_rows("明细定义")
        source_field_rows = _load_field_config_rows("明细参数")
        ref_field_rows = _load_field_config_rows("参照表字段配置")

        source_lookup = defaultdict(dict)
        for row in source_field_rows:
            detail_name = str(row.get("适用明细") or row.get("明细名称") or "").strip()
            field_name = str(row.get("字段名称") or row.get("明细字段") or "").strip()
            if not detail_name or not field_name:
                continue
            source_lookup[detail_name][field_name] = row

        ref_sheet_name = ""
        for row in ref_field_rows:
            ref_sheet_name = str(row.get("来源Sheet") or "").strip()
            if ref_sheet_name:
                break
        ref_sheet_selector = ref_sheet_name or "汇总"
        ref_code_col = _field_config_column_index(ref_field_rows, "产业大类编码", ref_sheet_name) or 2
        ref_ind4_col = _field_config_column_index(ref_field_rows, "对应国名经济行业编码小类", ref_sheet_name) or _field_config_column_index(ref_field_rows, "对应国民经济行业小类名称", ref_sheet_name) or 13
        # 汇总 sheet 中“小类行业编码”仍按原逻辑使用第 13 列
        if ref_ind4_col == _field_config_column_index(ref_field_rows, "对应国民经济行业小类名称", ref_sheet_name):
            ref_ind4_col = 13
        ref_star_col = _field_config_column_index(ref_field_rows, "是否带星号", ref_sheet_name) or 19
        ref_raw_col = _field_config_column_index(ref_field_rows, "原始映射类别", ref_sheet_name) or 21

        detail_name_to_type = {"科技产业": "tech", "数字经济产业": "digital", "养老产业": "elder"}

        for row in detail_defs:
            detail_name = str(row.get("明细名称") or "").strip()
            file_type = detail_name_to_type.get(detail_name, "")
            if not file_type:
                continue
            base = dict(inferred_by_type.get(file_type, {}))
            detail_fields = source_lookup.get(detail_name, {})
            reported_cols = []
            col_cat_map = []
            if file_type == "tech":
                tech_map = [
                    ("列位置_高技术制造业贷款类型大类编码", "HTP"),
                    ("列位置_高技术服务业贷款类型大类编码", "HTS"),
                    ("列位置_战略性新兴产业贷款类型大类编码", "SE"),
                    ("列位置_知识产权（专利）密集型产业贷款类型大类编码", "PA"),
                ]
                for field_name, cat in tech_map:
                    item = detail_fields.get(field_name)
                    if item and item.get("列序号"):
                        col_idx = parse_int(item.get("列序号"), 0)
                        if col_idx:
                            reported_cols.append(col_idx)
                            col_cat_map.append((col_idx, cat))
            elif file_type == "digital":
                item = detail_fields.get("列位置_产业大类编码")
                col_idx = parse_int(item.get("列序号") if item else None, 0)
                if col_idx:
                    reported_cols = [col_idx]
                    col_cat_map = [(col_idx, "DE")]
            elif file_type == "elder":
                item = detail_fields.get("列位置_产业大类编码")
                col_idx = parse_int(item.get("列序号") if item else None, 0)
                if col_idx:
                    reported_cols = [col_idx]
                    col_cat_map = [(col_idx, "EC")]

            rows.append(
                {
                    "输出工作表名称": str(row.get("输出工作表名称") or base.get("输出工作表名称") or "").strip(),
                    "来源台账文件": base.get("来源台账文件", ""),
                    "台账类型": file_type,
                    "表头行号": parse_int(row.get("表头行号"), base.get("表头行号", 1)),
                    "数据起始行号": parse_int(row.get("数据起始行号"), base.get("数据起始行号", 2)),
                    "贷款客户行业列序号": parse_int(detail_fields.get("列位置_贷款客户行业分类", {}).get("列序号"), base.get("贷款客户行业列序号", 6)),
                    "贷款投向行业列序号": parse_int(detail_fields.get("列位置_贷款实际投向行业分类", {}).get("列序号"), base.get("贷款投向行业列序号", 15)),
                    "贷款余额列序号": base.get("贷款余额列序号", 8),
                    "贷款余额原始单位": base.get("贷款余额原始单位", "万元"),
                    "机构报送产业分类列序号": reported_cols if file_type in {"digital", "elder"} else (reported_cols or base.get("机构报送产业分类列序号", [])),
                    "报送列-类别映射": col_cat_map if file_type in {"digital", "elder"} else (col_cat_map or base.get("报送列-类别映射", [])),
                    "参照表工作表序号": ref_sheet_selector,
                    "参照表产业分类代码列序号": ref_code_col,
                    "参照表行业4位码列序号": ref_ind4_col,
                    "参照表星标列序号": ref_star_col,
                    "参照表原始映射列序号": ref_raw_col,
                }
            )
        if not rows:
            rows = _load_legacy_config_rows()
    wb.close()
    return rows


def _load_legacy_config_rows():
    if not os.path.exists(CONFIG_FILE):
        return []
    wb = openpyxl.load_workbook(CONFIG_FILE, read_only=True, data_only=True)
    ws = get_sheet_by_logical_name(wb, "config") or wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        row = list(row)
        has_balance_unit_col = len(row) >= 16
        row = row + [None] * max(0, (16 if has_balance_unit_col else 15) - len(row))
        output_sheet = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        source_file = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        file_type_label = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        file_type = TYPE_LABEL_TO_KEY.get(file_type_label, "")
        if not output_sheet or not source_file or not file_type:
            continue
        customer_industry_col = parse_int(row[5], 6)
        actual_industry_col = parse_int(row[6], 15)
        balance_col = parse_int(row[7], 8)
        balance_unit = str(row[8]).strip() if has_balance_unit_col and row[8] not in {None, ""} else "万元"
        reported_cols = parse_idx_list(row[9] if has_balance_unit_col else row[8])
        col_cat_map = dedup_pairs_keep_order(parse_col_category_map(row[10] if has_balance_unit_col else row[9]))
        if not col_cat_map:
            col_cat_map = _default_col_category_map(file_type, reported_cols)
        if file_type in {"digital", "elder"} and len(reported_cols) > 1:
            reported_cols = [reported_cols[-1]]
        ref_sheet_idx = 11 if has_balance_unit_col else 10
        rows.append(
            {
                "输出工作表名称": output_sheet,
                "来源台账文件": source_file,
                "台账类型": file_type,
                "表头行号": parse_int(row[3], 1),
                "数据起始行号": parse_int(row[4], 2),
                "贷款客户行业列序号": customer_industry_col,
                "贷款投向行业列序号": actual_industry_col,
                "贷款余额列序号": balance_col,
                "贷款余额原始单位": balance_unit or "万元",
                "机构报送产业分类列序号": reported_cols,
                "报送列-类别映射": col_cat_map,
                "参照表工作表序号": parse_sheet_selector(row[ref_sheet_idx], 4),
                "参照表产业分类代码列序号": parse_int(row[ref_sheet_idx + 1], 2),
                "参照表行业4位码列序号": parse_int(row[ref_sheet_idx + 2], 13),
                "参照表星标列序号": parse_int(row[ref_sheet_idx + 3], 19),
                "参照表原始映射列序号": parse_int(row[ref_sheet_idx + 4], 21),
            }
        )
    wb.close()
    return rows


def build_mapping_by_config(mapping_file, cfg):
    wb = openpyxl.load_workbook(mapping_file, read_only=True, data_only=True)
    sel = cfg["参照表工作表序号"]
    sheetnames = wb.sheetnames
    # 支持数字序号或工作表名称
    if isinstance(sel, int):
        sheet_idx = sel - 1
        if sheet_idx < 0 or sheet_idx >= len(sheetnames):
            wb.close()
            return defaultdict(dict)
        ws = wb[sheetnames[sheet_idx]]
    else:
        name = str(sel).strip()
        if not name or name not in sheetnames:
            wb.close()
            return defaultdict(dict)
        ws = wb[name]
    code_idx = cfg["参照表产业分类代码列序号"] - 1
    ind_idx = cfg["参照表行业4位码列序号"] - 1
    star_idx = cfg["参照表星标列序号"] - 1
    raw_idx = cfg.get("参照表原始映射列序号", 21) - 1  # 参照表 U 列=21，存小类\中类\大类\门类等原始映射内容（与映射表 B2 无关）
    d = defaultdict(dict)
    target_industry_type = {
        "tech": "科技相关产业",
        "digital": "数字经济产业",
        "elder": "养老产业",
    }.get(cfg.get("台账类型"), "")
    industry_type_idx = 0 if str(cfg.get("参照表工作表序号") or "").strip() == "汇总" else None

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) if row is not None else []
        if industry_type_idx is not None:
            industry_type = str(row[industry_type_idx] or "").strip() if industry_type_idx < len(row) else ""
            if target_industry_type and industry_type != target_industry_type:
                continue
        if max(code_idx, ind_idx, star_idx) >= len(row):
            continue
        class_code = row[code_idx]
        ind4 = row[ind_idx]
        star_raw = row[star_idx]
        raw_mapping = ""
        if raw_idx < len(row) and row[raw_idx] is not None:
            raw_mapping = str(row[raw_idx]).strip()
        class_code = str(class_code).strip().upper() if class_code is not None else ""
        ind4 = str(ind4).strip().upper() if ind4 is not None else ""
        if not class_code or not re.fullmatch(r"[A-Z]{2,4}\d{2,6}", class_code):
            continue
        if not re.fullmatch(r"[A-Z]\d{4}", ind4):
            m = re.search(r"[A-Z]\d{4}", ind4)
            if not m:
                continue
            ind4 = m.group(0)
        star = is_star_value(star_raw)
        old = d[ind4].get(class_code)
        # 存 (是否星标, 原始映射内容列表)，同代码多条时保留全部层级映射（如小类/中类/大类）
        if old is None:
            raws = [raw_mapping] if raw_mapping else []
            d[ind4][class_code] = (star, raws)
        else:
            old_star, old_raws = old
            if isinstance(old_raws, list):
                raws = old_raws[:]
            elif isinstance(old_raws, str) and old_raws.strip():
                raws = [old_raws.strip()]
            else:
                raws = []
            if raw_mapping and raw_mapping not in raws:
                raws.append(raw_mapping)
            d[ind4][class_code] = (old_star or star, raws)

    wb.close()
    return d


def detail_name_from_cfg(cfg):
    return {"tech": "科技产业", "digital": "数字经济产业", "elder": "养老产业"}.get(cfg.get("台账类型"), cfg.get("台账类型", ""))


def _detect_validation_issue(field_name, header_value, sample_values):
    header_text = str(header_value or "").strip()
    samples = [str(v).strip() for v in sample_values if v not in (None, "")]
    joined_samples = " ".join(samples)
    if not header_text:
        return "是", "表头为空"
    if not samples:
        return "是", "抽样值为空"
    if "贷款余额" in field_name:
        if not any(re.search(r"-?\d+(?:\.\d+)?", text) for text in samples):
            return "是", "金额字段未抽到数字"
    if "行业" in field_name or "编码" in field_name or "报送" in field_name:
        if not re.search(r"[A-Z]{1,4}\d{2,6}", joined_samples.upper()):
            return "是", "编码字段未抽到类似 C1442/DE05/EC09 的值"
    if len(max(samples, key=len, default="")) > 120 and ("编码" in field_name or "列位置_" in field_name):
        return "是", "字段疑似命中长文本说明列"
    return "", ""


def _collect_non_empty_samples(ws, col_idx, start_row, max_samples=3):
    samples = []
    if not col_idx or col_idx > ws.max_column:
        return samples
    for r in range(max(1, start_row), ws.max_row + 1):
        value = ws.cell(r, col_idx).value
        if value in (None, ""):
            continue
        if isinstance(value, str) and not value.strip():
            continue
        samples.append(value)
        if len(samples) >= max_samples:
            break
    return samples


def write_column_validation_report(config_rows, source_file_map, mapping_file, detail_param_lookup):
    report_wb = Workbook()
    source_ws = report_wb.active
    source_ws.title = "源文件字段校验"
    source_headers = [
        "明细名称", "来源文件", "来源工作表", "字段名称", "配置列字母", "配置列序号",
        "实际命中表头", "抽样值1", "抽样值2", "抽样值3", "是否疑似异常", "异常说明",
    ]
    source_ws.append(source_headers)

    ref_ws = report_wb.create_sheet("参照表字段校验")
    ref_headers = [
        "适用对象", "来源文件", "来源Sheet", "字段名称", "配置列字母", "配置列序号",
        "实际命中表头", "抽样值1", "抽样值2", "抽样值3", "是否疑似异常", "异常说明",
    ]
    ref_ws.append(ref_headers)

    for ws in (source_ws, ref_ws):
        for c in range(1, ws.max_column + 1):
            ws.cell(1, c).font = Font(name="Arial", bold=True)

    for cfg in config_rows:
        detail_name = detail_name_from_cfg(cfg)
        source_abs = resolve_source_workbook_path(cfg.get("来源台账文件"), source_file_map)
        if not source_abs or not os.path.exists(source_abs):
            continue
        wb = openpyxl.load_workbook(source_abs, read_only=True, data_only=True)
        try:
            ws = pick_source_worksheet(wb, cfg["输出工作表名称"])
            header_row = cfg.get("表头行号", 1)
            data_start_row = cfg.get("数据起始行号", header_row + 1)
            fields = detail_param_lookup.get(detail_name, {})
            for field_name, spec in fields.items():
                col_letter = str(spec.get("column_letter") or "").upper()
                if not col_letter:
                    continue
                col_idx = column_index_from_string(col_letter)
                header_value = ws.cell(header_row, col_idx).value if col_idx <= ws.max_column else None
                samples = _collect_non_empty_samples(ws, col_idx, data_start_row, max_samples=3)
                abnormal, note = _detect_validation_issue(field_name, header_value, samples)
                source_ws.append([
                    detail_name,
                    source_abs,
                    ws.title,
                    field_name,
                    col_letter,
                    col_idx,
                    header_value,
                    samples[0] if len(samples) > 0 else None,
                    samples[1] if len(samples) > 1 else None,
                    samples[2] if len(samples) > 2 else None,
                    abnormal,
                    note,
                ])
        finally:
            wb.close()

    ref_rows = _load_field_config_rows("参照表字段配置")
    if not ref_rows:
        synthesized = []
        seen = set()
        for cfg in config_rows:
            detail_name = detail_name_from_cfg(cfg)
            items = [
                ("列位置_产业大类编码", cfg.get("参照表产业分类代码列序号")),
                ("列位置_对应国名经济行业编码小类", cfg.get("参照表行业4位码列序号")),
                ("列位置_是否带星号", cfg.get("参照表星标列序号")),
                ("列位置_原始映射类别", cfg.get("参照表原始映射列序号")),
            ]
            for field_name, col_idx in items:
                col_idx = parse_int(col_idx, 0)
                key = (detail_name, str(cfg.get("参照表工作表序号") or "").strip(), field_name, col_idx)
                if not col_idx or key in seen:
                    continue
                seen.add(key)
                synthesized.append(
                    {
                        "适用明细": detail_name,
                        "来源Sheet": str(cfg.get("参照表工作表序号") or "").strip(),
                        "字段名称": field_name,
                        "列字母": get_column_letter(col_idx),
                        "列序号": col_idx,
                    }
                )
        ref_rows = synthesized
    if mapping_file and os.path.exists(mapping_file):
        wb = openpyxl.load_workbook(mapping_file, read_only=True, data_only=True)
        try:
            grouped_rows = defaultdict(list)
            for row in ref_rows:
                grouped_rows[str(row.get("来源Sheet") or "").strip() or wb.sheetnames[0]].append(row)
            for sheet_name, rows in grouped_rows.items():
                if sheet_name not in wb.sheetnames:
                    for row in rows:
                        ref_ws.append([
                            str(row.get("适用明细") or "").strip(),
                            mapping_file,
                            sheet_name,
                            str(row.get("字段名称") or "").strip(),
                            str(row.get("列字母") or "").strip().upper(),
                            parse_int(row.get("列序号"), 0),
                            None,
                            None,
                            None,
                            None,
                            "是",
                            "来源Sheet不存在",
                        ])
                    continue
                ws = wb[sheet_name]
                for row in rows:
                    field_name = str(row.get("字段名称") or "").strip()
                    col_letter = str(row.get("列字母") or "").strip().upper()
                    col_idx = parse_int(row.get("列序号"), 0)
                    header_value = ws.cell(1, col_idx).value if col_idx and col_idx <= ws.max_column else None
                    samples = [
                        ws.cell(r, col_idx).value if col_idx and col_idx <= ws.max_column else None
                        for r in range(2, min(ws.max_row, 4) + 1)
                    ]
                    abnormal, note = _detect_validation_issue(field_name, header_value, samples)
                    ref_ws.append([
                        str(row.get("适用明细") or "").strip(),
                        mapping_file,
                        sheet_name,
                    field_name,
                    col_letter,
                    col_idx,
                    header_value,
                    samples[0] if len(samples) > 0 else None,
                    samples[1] if len(samples) > 1 else None,
                    samples[2] if len(samples) > 2 else None,
                    abnormal,
                    note,
                ])
        finally:
            wb.close()

    output_path = build_column_validation_output_path()
    report_wb.save(output_path)
    report_wb.close()
    return output_path


def build_text_field_catalog(cfg, src_ws, row_idx, detail_param_lookup):
    detail_name = detail_name_from_cfg(cfg)

    def getv(col_idx):
        return src_ws.cell(row_idx, col_idx).value if col_idx <= src_ws.max_column else None

    def add_entry(catalog, key, value, label, column_letter=None):
        entry = {"value": value, "label": label}
        if column_letter:
            entry["column_letter"] = column_letter
        catalog[key] = entry

    catalog = {}
    detail_param_map = detail_param_lookup.get(detail_name, {})
    for field_name, spec in detail_param_map.items():
        col_letter = spec.get("column_letter")
        if not col_letter:
            continue
        col_idx = column_index_from_string(col_letter)
        add_entry(catalog, field_name, getv(col_idx), f"{field_name}({col_letter})", col_letter)

    legacy_specs = _legacy_text_field_specs(detail_name)
    for key, spec in legacy_specs.items():
        if spec.get("aggregate"):
            parts = []
            for token in spec["aggregate"]:
                match = TEXT_RULE_COLUMN_PATTERN.match(token)
                if not match:
                    continue
                col_letter, _ = match.groups()
                value = getv(column_index_from_string(col_letter))
                if value not in (None, ""):
                    parts.append(str(value).strip())
            add_entry(catalog, key, "；".join(parts), spec.get("label") or key)
            continue
        col_letter = spec.get("column_letter")
        if not col_letter:
            continue
        value = getv(column_index_from_string(col_letter))
        label = spec.get("label") or key
        add_entry(catalog, key, value, label, col_letter)
        explicit_key = f"列位:{col_letter}[{label}]"
        add_entry(catalog, explicit_key, value, label, col_letter)
        short_key = f"列位:{col_letter}"
        if short_key not in catalog:
            add_entry(catalog, short_key, value, label, col_letter)

    return catalog


def extract_text_fields_for_row(cfg, src_ws, row_idx, detail_param_lookup):
    field_catalog = build_text_field_catalog(cfg, src_ws, row_idx, detail_param_lookup)
    text_fields = {
        "客户名称": _catalog_entry_value(field_catalog.get("客户名称")),
        "贷款类型": _catalog_entry_value(field_catalog.get("贷款类型")),
        "贷款用途": _catalog_entry_value(field_catalog.get("贷款用途")),
        "贷款客户行业分类": _catalog_entry_value(field_catalog.get("贷款客户行业分类")),
        "贷款实际投向行业": _catalog_entry_value(field_catalog.get("贷款实际投向行业")),
        "佐证摘要": _catalog_entry_value(field_catalog.get("佐证摘要")),
    }
    return text_fields, field_catalog


def build_structured_text_summary(cfg, src_ws, row_idx, mapping_data, clue_rules, runtime_settings, industry_desc_map):
    reported_summary = []
    issue_summary = []
    support_any = False
    basis_columns = {
        "actual": cfg.get("贷款投向行业列序号", 15),
        "customer": cfg.get("贷款客户行业列序号", cfg.get("贷款投向行业列序号", 15)),
    }
    col_cat_map = dedup_pairs_keep_order(cfg.get("报送列-类别映射", []))
    reported_cols = cfg["机构报送产业分类列序号"]
    if not col_cat_map:
        col_cat_map = _default_col_category_map(cfg["台账类型"], reported_cols)
    category_order = dedup_keep_order([cat for _, cat in col_cat_map])
    for basis in runtime_settings["enabled_bases"]:
        industry_col = basis_columns.get(basis, 0)
        industry4 = extract_industry4(src_ws.cell(row_idx, industry_col).value) if industry_col <= src_ws.max_column else ""
        match_map = resolve_match_map(industry4, mapping_data, industry_desc_map)
        for cat in category_order:
            reported_cat = []
            for col, mapped_cat in col_cat_map:
                if mapped_cat == cat and col <= src_ws.max_column:
                    reported_cat.extend(extract_codes(src_ws.cell(row_idx, col).value))
            result = evaluate_basis_result(industry4, reported_cat, match_map, cat, clue_rules)
            if result["matched"]:
                support_any = True
                reported_summary.append(f"{BASIS_NAME_MAP.get(basis, basis)}-{category_display_name(cat)}:{result['matched']}")
            problems = "；".join(item for item in [result["错报"], result["漏报"], result["疑似漏报"]] if item)
            if problems:
                issue_summary.append(f"{BASIS_NAME_MAP.get(basis, basis)}-{category_display_name(cat)}:{problems}")
    return {
        "结构化支持": support_any,
        "结构化匹配摘要": " | ".join(reported_summary),
        "结构化问题摘要": " | ".join(issue_summary),
    }


def build_text_analysis_output_path():
    os.makedirs(LOCAL_OUTPUT_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(LOCAL_OUTPUT_DIR, f"文本辅助核查结果_{ts}.xlsx")


def write_text_analysis_results():
    init_config_file_if_missing()
    ensure_config_schema()
    source_file_map = get_source_file_map_from_mapping()
    config_rows = load_config_rows()
    if not config_rows and source_file_map:
        config_rows = build_config_rows_from_files(source_file_map.values())
    config_rows = reconcile_config_rows_with_mapping(config_rows, source_file_map)
    if not config_rows:
        raise ValueError("未能从config.xlsx或目录中推断任何台账配置行。")
    standardized_map = standardize_source_workbooks(config_rows, source_file_map)
    mapping_file = find_mapping_file()
    mapping_cache = {}
    clue_rules = load_clue_rules()
    runtime_settings = load_runtime_settings()
    industry_desc_map = load_industry_desc_map()
    text_settings = load_text_analysis_settings()
    detail_param_lookup = load_detail_param_lookup()
    text_rules = text_settings["rules"]
    verdict_rules = text_settings["verdicts"]
    stopwords = text_settings["stopwords"]

    summary_rows = []
    rule_hit_rows = []
    output_wb = Workbook()
    summary_ws = output_wb.active
    summary_ws.title = "文本辅助核查"
    summary_headers = [
        "明细名称", "来源工作表", "序号", "贷款合同编码", "贷款借据编码", "客户名称",
        "结构化匹配摘要", "结构化问题摘要", "贷款类型", "贷款用途", "佐证摘要",
        "文本支持产业", "目标产业支持度", "命中矛盾规则", "文本问题摘要", "结论", "结论原因", "复核建议",
    ]
    summary_ws.append(summary_headers)
    for c in range(1, len(summary_headers) + 1):
        summary_ws.cell(1, c).font = Font(name="Arial", bold=True)

    for cfg in config_rows:
        source_abs = resolve_source_workbook_path(cfg.get("来源台账文件"), source_file_map)
        if not source_abs or not os.path.exists(source_abs):
            continue
        source_for_analysis = resolve_standardized_workbook_path(source_abs, cfg["输出工作表名称"], standardized_map)
        src_wb = openpyxl.load_workbook(source_for_analysis, read_only=True, data_only=True)
        src_ws = pick_source_worksheet(src_wb, cfg["输出工作表名称"])
        cache_key = (
            cfg.get("台账类型"),
            cfg["参照表工作表序号"],
            cfg["参照表产业分类代码列序号"],
            cfg["参照表行业4位码列序号"],
            cfg["参照表星标列序号"],
            cfg.get("参照表原始映射列序号", 21),
        )
        if cache_key not in mapping_cache:
            mapping_cache[cache_key] = build_mapping_by_config(mapping_file, cfg)
        mapping_data = mapping_cache[cache_key]
        detail_name = detail_name_from_cfg(cfg)
        for row_idx in range(cfg["数据起始行号"], src_ws.max_row + 1):
            seq_value = src_ws.cell(row_idx, 1).value
            if seq_value in (None, ""):
                continue
            text_fields, field_catalog = extract_text_fields_for_row(cfg, src_ws, row_idx, detail_param_lookup)
            structured = build_structured_text_summary(cfg, src_ws, row_idx, mapping_data, clue_rules, runtime_settings, industry_desc_map)
            contradiction_hits = detect_text_contradictions(detail_name, field_catalog, text_rules, detail_param_lookup)
            semantic_scores = score_text_semantics(detail_name, field_catalog, text_rules, stopwords, detail_param_lookup)
            verdict = build_text_verdict(detail_name, semantic_scores, contradiction_hits, verdict_rules)
            hit_ids = "、".join(item["规则编号"] for item in contradiction_hits)
            problem_text = "；".join(item["命中说明"] for item in contradiction_hits)
            summary_row = [
                detail_name,
                src_ws.title,
                seq_value,
                src_ws.cell(row_idx, 2).value,
                src_ws.cell(row_idx, 3).value,
                text_fields.get("客户名称"),
                structured["结构化匹配摘要"],
                structured["结构化问题摘要"],
                text_fields.get("贷款类型"),
                text_fields.get("贷款用途"),
                text_fields.get("佐证摘要"),
                verdict["文本支持产业"],
                verdict["目标产业支持度"],
                hit_ids,
                problem_text,
                verdict["结论"],
                verdict["结论原因"],
                verdict["复核建议"],
            ]
            summary_rows.append(summary_row)
            summary_ws.append(summary_row)
            for hit in contradiction_hits:
                rule_hit_rows.append([
                    detail_name,
                    src_ws.title,
                    seq_value,
                    text_fields.get("客户名称"),
                    hit["规则编号"],
                    hit["规则类型"],
                    hit["风险等级"],
                    hit["命中说明"],
                    hit["命中字段"],
                ])
        src_wb.close()

    hit_ws = output_wb.create_sheet("规则命中明细")
    hit_headers = ["明细名称", "来源工作表", "序号", "客户名称", "规则编号", "规则类型", "风险等级", "命中说明", "命中字段"]
    hit_ws.append(hit_headers)
    for c in range(1, len(hit_headers) + 1):
        hit_ws.cell(1, c).font = Font(name="Arial", bold=True)
    for row in rule_hit_rows:
        hit_ws.append(row)

    snapshot_ws = output_wb.create_sheet("词典配置快照")
    snapshot_ws.append(["文本规则"])
    snapshot_ws.append(TEXT_RULE_HEADERS)
    for row in text_rules:
        snapshot_ws.append([row.get(header) for header in TEXT_RULE_HEADERS])
    snapshot_ws.append([])
    snapshot_ws.append(["文本结论"])
    snapshot_ws.append(TEXT_VERDICT_HEADERS)
    for row in verdict_rules:
        snapshot_ws.append([row.get(header) for header in TEXT_VERDICT_HEADERS])
    snapshot_ws.append([])
    snapshot_ws.append(["文本停用词"])
    snapshot_ws.append(["词语"])
    for item in stopwords:
        snapshot_ws.append([item])

    output_path = build_text_analysis_output_path()
    output_wb.save(output_path)
    output_wb.close()
    return output_path


def write_results():
    init_config_file_if_missing()
    ensure_config_schema()
    source_file_map = get_source_file_map_from_mapping()
    config_rows = load_config_rows()
    # 若config表中没有有效行，则尝试根据mapping中的源文件路径自动构造
    if not config_rows and source_file_map:
        config_rows = build_config_rows_from_files(source_file_map.values())
    config_rows = reconcile_config_rows_with_mapping(config_rows, source_file_map)
    # 仍然没有，则退回扫描目录自动识别
    if not config_rows:
        auto_files = find_ledger_files()
        config_rows = build_config_rows_from_files(auto_files)
    if not config_rows:
        raise ValueError("未能从config.xlsx或目录中推断任何台账配置行。")
    temp_standardized_dir = None
    try:
        standardized_map, temp_standardized_dir = prepare_standardized_sources_for_strategy(config_rows, source_file_map)

        mapping_file = find_mapping_file()
        map_wb = openpyxl.load_workbook(mapping_file, read_only=True, data_only=True)
        mapping_sheetnames = map_wb.sheetnames
        map_wb.close()

        out_wb = Workbook()
        out_wb.remove(out_wb.active)
        used_names = set()
        mapping_cache = {}
        executed_entries = []
        industry_desc_map = load_industry_desc_map()
        industry_hierarchy_map = load_industry_hierarchy_map()
        clue_rules = load_clue_rules()
        runtime_settings = load_runtime_settings()
        detail_param_lookup = load_detail_param_lookup()
        keyword_rules = load_keyword_rule_rows()
        write_column_validation_report(config_rows, source_file_map, mapping_file, detail_param_lookup)
        template_summary_amounts = {}

        for cfg in config_rows:
            source_abs = resolve_source_workbook_path(cfg.get("来源台账文件"), source_file_map)
            if source_abs is None or not str(source_abs).strip():
                print(f"跳过：来源台账文件为空 -> 输出工作表名称「{cfg.get('输出工作表名称', '')}」")
                continue
            if not os.path.exists(source_abs):
                print(f"跳过：来源文件不存在 -> {cfg['来源台账文件']}")
                continue

            source_for_strategy = resolve_standardized_workbook_path(source_abs, cfg["输出工作表名称"], standardized_map)
            src_wb = openpyxl.load_workbook(source_for_strategy, read_only=True, data_only=True)
            try:
                src_ws = pick_source_worksheet(src_wb, cfg["输出工作表名称"])
                dst_name = ensure_unique_sheet_name(cfg["输出工作表名称"], used_names)
                dst_ws = out_wb.create_sheet(dst_name)
                copy_sheet(src_ws, dst_ws)

                working_ws = dst_ws
                working_max_column = working_ws.max_column
                working_max_row = working_ws.max_row

                header_row = cfg["表头行号"]
                data_start_row = cfg["数据起始行号"] or (header_row + 1)
                amount_col = cfg.get("贷款余额列序号", 8)
                reported_cols = cfg["机构报送产业分类列序号"]
                col_cat_map = dedup_pairs_keep_order(cfg.get("报送列-类别映射", []))
                if not col_cat_map:
                    col_cat_map = _default_col_category_map(cfg["台账类型"], reported_cols)
                category_order = dedup_keep_order([cat for _, cat in col_cat_map])
                summary_sheet_name = detail_name_from_cfg(cfg)

                start_col = working_max_column + 1
                cache_key = (
                    cfg.get("台账类型"),
                    cfg["参照表工作表序号"],
                    cfg["参照表产业分类代码列序号"],
                    cfg["参照表行业4位码列序号"],
                    cfg["参照表星标列序号"],
                    cfg.get("参照表原始映射列序号", 21),
                )
                if cache_key not in mapping_cache:
                    mapping_cache[cache_key] = build_mapping_by_config(mapping_file, cfg)
                mp = mapping_cache[cache_key]

                headers = build_result_headers(category_order, runtime_settings)
                for i, h in enumerate(headers):
                    cell = dst_ws.cell(header_row, start_col + i)
                    cell.value = h
                    cell.font = Font(name="Arial", bold=True)

                count_by_cat = {
                    basis: {cat: {"多报": 0, "漏报": 0, "疑似多报": 0, "疑似漏报": 0} for cat in category_order}
                    for basis in runtime_settings["enabled_bases"]
                }
                basis_columns = {
                    "actual": cfg.get("贷款投向行业列序号", 15),
                    "customer": cfg.get("贷款客户行业列序号", cfg.get("贷款投向行业列序号", 15)),
                }

                for r in range(data_start_row, working_max_row + 1):
                    amount_value = parse_amount(working_ws.cell(r, amount_col).value) if amount_col <= working_max_column else 0.0
                    amount_value_for_summary = amount_value * _unit_scale(
                        cfg.get("贷款余额原始单位", "万元"), runtime_settings["summary_unit"]
                    )
                    col_offset = 0
                    desc_source_code = ""
                    basis_results = {}
                    for basis in runtime_settings["enabled_bases"]:
                        industry_col = basis_columns.get(basis, 0)
                        industry4 = extract_industry4(working_ws.cell(r, industry_col).value) if industry_col <= working_max_column else ""
                        if not desc_source_code:
                            desc_source_code = industry4
                        match_map = resolve_match_map(industry4, mp, industry_desc_map)
                        for cat in category_order:
                            reported_cat = []
                            for col, mapped_cat in col_cat_map:
                                if mapped_cat != cat:
                                    continue
                                if col and col > 0 and col <= working_max_column:
                                    reported_cat.extend(extract_codes(working_ws.cell(r, col).value))
                            reported_cat = dedup_keep_order(reported_cat)
                            result = evaluate_basis_result(industry4, reported_cat, match_map, cat, clue_rules)
                            basis_results[(basis, cat)] = result
                            dst_ws.cell(r, start_col + col_offset).value = result["reported"]
                            col_offset += 1
                            dst_ws.cell(r, start_col + col_offset).value = result["matched"]
                            col_offset += 1
                            for result_label in basis_result_types(runtime_settings):
                                dst_ws.cell(r, start_col + col_offset).value = resolve_result_label_display(result, result_label)
                                col_offset += 1
                            dst_ws.cell(r, start_col + col_offset).value = result["是否线索"]
                            col_offset += 1
                            dst_ws.cell(r, start_col + col_offset).value = result["是否疑似线索"]
                            col_offset += 1
                            dst_ws.cell(r, start_col + col_offset).value = result["备注"]
                            col_offset += 1
                            if result["多报"]:
                                count_by_cat[basis][cat]["多报"] += 1
                            if result["疑似多报"]:
                                count_by_cat[basis][cat]["疑似多报"] += 1
                            if result["漏报"]:
                                count_by_cat[basis][cat]["漏报"] += 1
                            if result["疑似漏报"]:
                                count_by_cat[basis][cat]["疑似漏报"] += 1

                    field_catalog = build_text_field_catalog(cfg, working_ws, r, detail_param_lookup)
                    keyword_result = {
                        "是否命中": "",
                        "规则名称": "",
                        "命中说明": "",
                        "是否疑似线索": "",
                    }
                    if keyword_detail_enabled(runtime_settings) or keyword_summary_enabled(runtime_settings):
                        keyword_result = evaluate_keyword_rules_for_row(
                            detail_name_from_cfg(cfg),
                            field_catalog,
                            keyword_rules,
                            industry_hierarchy_map,
                        )

                    for cat in category_order:
                        valid_reported_codes = [
                            reported_code
                            for reported_code in dedup_keep_order(
                                [
                                    code
                                    for col, mapped_cat in col_cat_map
                                    if mapped_cat == cat and col and col > 0 and col <= working_max_column
                                    for code in extract_codes(working_ws.cell(r, col).value)
                                ]
                            )
                            if re.fullmatch(r"[A-Z]{2,4}\d{2,6}", reported_code)
                        ]
                        if not valid_reported_codes:
                            continue

                        actual_label = basis_results.get(("actual", cat), {}).get("主标签")
                        customer_label = basis_results.get(("customer", cat), {}).get("主标签")
                        stat_keys = classify_template_stat_keys(actual_label, customer_label, runtime_settings)
                        group_name = resolve_summary_group_name(summary_sheet_name, cat)

                        for _ in runtime_settings["enabled_bases"]:
                            if group_name:
                                append_template_summary_stat(
                                    template_summary_amounts,
                                    summary_sheet_name,
                                    _summary_group_key(group_name),
                                    "total_amount",
                                    amount=amount_value_for_summary,
                                )
                                append_template_summary_stat(
                                    template_summary_amounts,
                                    summary_sheet_name,
                                    _summary_group_key(group_name),
                                    "total_count",
                                    count=1,
                                )
                            for reported_code in valid_reported_codes:
                                append_template_summary_stat(
                                    template_summary_amounts,
                                    summary_sheet_name,
                                    reported_code,
                                    "total_amount",
                                    amount=amount_value_for_summary,
                                )
                                append_template_summary_stat(
                                    template_summary_amounts,
                                    summary_sheet_name,
                                    reported_code,
                                    "total_count",
                                    count=1,
                                )

                        if group_name:
                            for stat_key in stat_keys:
                                append_template_summary_stat(
                                    template_summary_amounts,
                                    summary_sheet_name,
                                    _summary_group_key(group_name),
                                    stat_key,
                                    amount=amount_value_for_summary,
                                    count=1,
                                )
                            if should_count_keyword_summary(keyword_result, stat_keys):
                                for keyword_stat_key in keyword_template_stat_keys(runtime_settings):
                                    append_template_summary_stat(
                                        template_summary_amounts,
                                        summary_sheet_name,
                                        _summary_group_key(group_name),
                                        keyword_stat_key,
                                        amount=amount_value_for_summary,
                                        count=1,
                                    )
                        for reported_code in valid_reported_codes:
                            for stat_key in stat_keys:
                                append_template_summary_stat(
                                    template_summary_amounts,
                                    summary_sheet_name,
                                    reported_code,
                                    stat_key,
                                    amount=amount_value_for_summary,
                                    count=1,
                                )
                            if should_count_keyword_summary(keyword_result, stat_keys):
                                for keyword_stat_key in keyword_template_stat_keys(runtime_settings):
                                    append_template_summary_stat(
                                        template_summary_amounts,
                                        summary_sheet_name,
                                        reported_code,
                                        keyword_stat_key,
                                        amount=amount_value_for_summary,
                                        count=1,
                                    )

                    if keyword_detail_enabled(runtime_settings):
                        dst_ws.cell(r, start_col + col_offset).value = keyword_result["是否命中"]
                        col_offset += 1
                        dst_ws.cell(r, start_col + col_offset).value = keyword_result["规则名称"]
                        col_offset += 1
                        dst_ws.cell(r, start_col + col_offset).value = keyword_result["命中说明"]
                        col_offset += 1
                        dst_ws.cell(r, start_col + col_offset).value = keyword_result["是否疑似线索"]
                        col_offset += 1

                    desc_entry = industry_desc_map.get(desc_source_code) or (
                        industry_desc_map.get(desc_source_code[1:])
                        if (
                            desc_source_code
                            and len(desc_source_code) == 5
                            and desc_source_code[0].isalpha()
                            and desc_source_code[1:].isdigit()
                        )
                        else None
                    )
                    if isinstance(desc_entry, tuple):
                        code_disp, desc_text = desc_entry
                        industry_desc = f"{code_disp}：{desc_text}" if desc_text else code_disp
                    else:
                        industry_desc = desc_entry or ""
                    dst_ws.cell(r, start_col + col_offset).value = industry_desc

                executed_entries.append(
                    {
                        "输出工作表名称": dst_name,
                        "来源台账文件": cfg["来源台账文件"],
                        "台账类型": cfg["台账类型"],
                        "参照表工作表序号": cfg["参照表工作表序号"],
                        "多报数量": "；".join(
                            f"{BASIS_NAME_MAP.get(basis, basis)}-{category_display_name(cat)}：{count_by_cat[basis][cat]['多报']}"
                            for basis in runtime_settings["enabled_bases"]
                            for cat in category_order
                        ),
                        "漏报数量": "；".join(
                            f"{BASIS_NAME_MAP.get(basis, basis)}-{category_display_name(cat)}：{count_by_cat[basis][cat]['漏报']}"
                            for basis in runtime_settings["enabled_bases"]
                            for cat in category_order
                        ),
                        "疑似多报数量": "；".join(
                            f"{BASIS_NAME_MAP.get(basis, basis)}-{category_display_name(cat)}：{count_by_cat[basis][cat]['疑似多报']}"
                            for basis in runtime_settings["enabled_bases"]
                            for cat in category_order
                        ),
                        "疑似漏报数量": "；".join(
                            f"{BASIS_NAME_MAP.get(basis, basis)}-{category_display_name(cat)}：{count_by_cat[basis][cat]['疑似漏报']}"
                            for basis in runtime_settings["enabled_bases"]
                            for cat in category_order
                        ),
                    }
                )
            finally:
                src_wb.close()

        if not out_wb.worksheets:
            ws_info = out_wb.create_sheet("说明", 0)
            ws_info["A1"] = "所有配置的台账来源文件均不存在或已跳过，未生成任何明细表。"
            ws_info["A2"] = "请检查 config 表「来源台账文件」及 mapping 表「源文件路径」，确保文件存在后再执行策略一。"
        else:
            write_template_summary_sheets(out_wb, template_summary_amounts, runtime_settings)

        for ws in out_wb.worksheets:
            ws.sheet_view.zoomScale = 90

        actual_output = build_strategy_output_path()
        out_wb.save(actual_output)
        out_wb.close()
        append_run_log(mapping_file, mapping_sheetnames, executed_entries, actual_output)
        return mapping_file, actual_output, CONFIG_FILE
    finally:
        if temp_standardized_dir is not None:
            temp_standardized_dir.cleanup()
    

def cli_set_mapping_path():
    """配置参照表路径，写入 mapping 表 B1。"""
    path = _ask_open_file("选择参照表文件（五篇大文章与国民经济行业分类对应参照表）")
    if path is None and _tk is None:
        path = input("当前环境无法打开文件选择器，请输入参照表路径（留空取消）：").strip()
    if not path:
        print("已取消设置参照表路径。")
        return
    if not os.path.isabs(path):
        candidate = os.path.join(ROOT, path)
        if os.path.exists(candidate):
            path = candidate
    wb = openpyxl.load_workbook(CONFIG_FILE) if os.path.exists(CONFIG_FILE) else Workbook()
    if not wb.sheetnames:
        wb.create_sheet(PREFERRED_SHEET_NAMES["config"])
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        ws = wb.create_sheet(PREFERRED_SHEET_NAMES["mapping"])
        ws.append(["参照表路径", path])
        ws.append(["映射表路径", ""])
    else:
        if ws.max_row < 1:
            ws.append(["参照表路径", path])
            ws.append(["映射表路径", ""])
        elif ws.max_row < 2:
            ws.cell(1, 1).value = "参照表路径"
            ws.cell(1, 2).value = path
            ws.append(["映射表路径", ""])
        else:
            ws.cell(1, 1).value = "参照表路径"
            ws.cell(1, 2).value = path
            if str(ws.cell(2, 1).value or "").strip() != "映射表路径":
                ws.cell(2, 1).value = "映射表路径"
    try:
        wb.save(CONFIG_FILE)
    except PermissionError as exc:
        wb.close()
        raise RuntimeError("config.xlsx 当前被占用，请先关闭 Excel 中的 config.xlsx 后重试。") from exc
    wb.close()
    print("参照表路径已更新为：", path)


def cli_set_source_files():
    paths = _ask_open_files("选择源文件（可多选）")
    if not paths and _tk is None:
        print("请输入源文件路径，多个用分号 ; 分隔。")
        raw = input("源文件路径列表：").strip()
        paths = [p.strip() for p in raw.split(";") if p.strip()]
    if not paths:
        print("已取消或未选择任何源文件。")
        return
    wb = openpyxl.load_workbook(CONFIG_FILE) if os.path.exists(CONFIG_FILE) else Workbook()
    if not wb.sheetnames:
        wb.create_sheet(PREFERRED_SHEET_NAMES["config"])
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        ws = wb.create_sheet(PREFERRED_SHEET_NAMES["mapping"])
        ws.append(["参照表路径", ""])
        ws.append(["映射表路径", ""])
        ws.append(["源文件", "源文件路径"])
    else:
        # 查找或创建 “源文件 / 源文件路径” 表头
        header_row = None
        for r in range(1, ws.max_row + 1):
            v1 = ws.cell(r, 1).value
            v2 = ws.cell(r, 2).value
            if str(v1).strip() == "源文件" and str(v2).strip() == "源文件路径":
                header_row = r
                break
        if header_row is None:
            header_row = ws.max_row + 1
            ws.cell(header_row, 1).value = "源文件"
            ws.cell(header_row, 2).value = "源文件路径"
        # 清理旧的源文件行
        for r in range(header_row + 1, ws.max_row + 1):
            ws.cell(r, 1).value = None
            ws.cell(r, 2).value = None
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    # 重新填充源文件列表
    ws = get_sheet_by_logical_name(wb, "mapping")
    header_row = None
    for r in range(1, ws.max_row + 1):
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        if str(v1).strip() == "源文件" and str(v2).strip() == "源文件路径":
            header_row = r
            break
    row = header_row + 1
    for p in paths:
        ws.cell(row, 1).value = os.path.basename(p)
        ws.cell(row, 2).value = p
        row += 1
    try:
        wb.save(CONFIG_FILE)
    except PermissionError as exc:
        wb.close()
        raise RuntimeError("config.xlsx 当前被占用，请先关闭 Excel 中的 config.xlsx 后重试。") from exc
    wb.close()
    print(f"已更新 {len(paths)} 个源文件路径。")


def cli_set_industry_map_path():
    """配置映射表（国民经济行业分类映射表）路径，写入 mapping 表 B2。"""
    path = _ask_open_file("选择国民经济行业分类映射表文件")
    if path is None and _tk is None:
        path = input("当前环境无法打开文件选择器，请输入映射表路径（留空取消）：").strip()
    if not path:
        print("已取消设置国民经济行业分类映射表路径。")
        return
    if not os.path.isabs(path):
        candidate = os.path.join(ROOT, path)
        if os.path.exists(candidate):
            path = candidate
    wb = openpyxl.load_workbook(CONFIG_FILE) if os.path.exists(CONFIG_FILE) else Workbook()
    if not wb.sheetnames:
        wb.create_sheet(PREFERRED_SHEET_NAMES["config"])
    ws = get_sheet_by_logical_name(wb, "mapping")
    if ws is None:
        ws = wb.create_sheet(PREFERRED_SHEET_NAMES["mapping"])
        ws.append(["参照表路径", ""])
        ws.append(["映射表路径", path])
    else:
        if ws.max_row < 1:
            ws.append(["参照表路径", ""])
            ws.append(["映射表路径", path])
        elif ws.max_row < 2:
            ws.cell(1, 1).value = "参照表路径"
            ws.cell(1, 2).value = ws.cell(1, 2).value or ""
            ws.append(["映射表路径", path])
        else:
            ws.cell(2, 1).value = "映射表路径"
            ws.cell(2, 2).value = path
    try:
        wb.save(CONFIG_FILE)
    except PermissionError as exc:
        wb.close()
        raise RuntimeError("config.xlsx 当前被占用，请先关闭 Excel 中的 config.xlsx 后重试。") from exc
    wb.close()
    print("国民经济行业分类映射表路径已更新为：", path)


def reset_config_sheet():
    rows = initialize_config_workbook(CONFIG_FILE)
    print(f"config.xlsx 已初始化完成，共写入 {len(rows)} 行策略一基准配置。")


def main():
    while True:
        print("\n====== 策略一工具面板 ======")
        print("1. 执行策略一（读取config.xlsx和参照表）")
        print("2. 配置参照表路径（mapping面板）")
        print("3. 批量配置源文件路径（mapping面板）")
        print("4. 配置国民经济行业分类映射表路径（mapping面板）")
        print("5. 初始化config.xlsx")
        print("7. 仅执行源文件标准化（按 mapping 源文件路径输出到当前目录 output）")
        print("0. 退出")
        choice = input("请输入序号并回车：").strip()
        try:
            if choice == "1":
                mapping, out, cfg = write_results()
                print("策略一执行完成。")
                print("参照表：", mapping)
                print("结果文件：", out)
            elif choice == "7":
                outputs = standardize_only_from_mapping()
                if outputs:
                    print("源文件标准化完成，输出文件：")
                    for path in outputs:
                        print(path)
                else:
                    print("未命中任何可标准化的源文件，请检查 mapping、明细定义 和 明细参数 配置。")
            elif choice == "2":
                cli_set_mapping_path()
            elif choice == "3":
                cli_set_source_files()
            elif choice == "4":
                cli_set_industry_map_path()
            elif choice == "5":
                reset_config_sheet()
            elif choice in ("0", ""):
                print("已退出。")
                break
            else:
                print("无效输入，请重新选择。")
        except RuntimeError as exc:
            print(str(exc))
        except PermissionError:
            print("config.xlsx 当前被占用，请先关闭 Excel 中的 config.xlsx 后重试。")


if __name__ == "__main__":
    main()
